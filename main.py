
from fastapi import FastAPI, UploadFile, File, Form, Query, Request, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, Response, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from datetime import datetime, timedelta
from fastapi.middleware.cors import CORSMiddleware
from openai import OpenAI
from dotenv import load_dotenv
from cryptography.fernet import Fernet, InvalidToken
import hashlib
import io
import json
import os
import secrets
import sqlite3
import time

BASE_DIR = os.path.dirname(__file__)
load_dotenv(os.path.join(BASE_DIR, ".env"))
api_key = os.getenv("OPENAI_API_KEY", "").strip()
client = OpenAI(api_key=api_key) if api_key else None

# Datos persistentes (SQLite y documentos subidos). En Render, DATA_DIR debe
# apuntar al mountPath del disco persistente para que sobrevivan a los deploys.
# En local (sin DATA_DIR configurada) se usa la misma carpeta del proyecto.
DATA_DIR = os.getenv("DATA_DIR", "").strip() or BASE_DIR
os.makedirs(DATA_DIR, exist_ok=True)

DASHBOARD_PATH = os.path.join(BASE_DIR, "dashboard.html")
MENU_PRINCIPAL_PATH = os.path.join(BASE_DIR, "menu_principal.html")
LOGIN_PATH = os.path.join(BASE_DIR, "login.html")
ADMINISTRACION_PATH = os.path.join(BASE_DIR, "administracion.html")
FORM_VIAJE_PATH = os.path.join(BASE_DIR, "form_viaje.html")
FORM_RECURSOS_PATH = os.path.join(BASE_DIR, "form_recursos.html")
PERSONAL_FORM_PATH = os.path.join(BASE_DIR, "personal.html")
COMPRAS_PATH = os.path.join(BASE_DIR, "compras.html")
PRINT_VIAJE_PATH = os.path.join(BASE_DIR, "print_viaje.html")
PRINT_ORDEN_SALIDA_PATH = os.path.join(BASE_DIR, "print_orden_salida.html")
ORDENES_VIEW_PATH = os.path.join(BASE_DIR, "ordenes_view.html")
GESTION_OPERATIVA_PATH = os.path.join(BASE_DIR, "gestion_operativa.html")
ALMACEN_V2_PATH = os.path.join(BASE_DIR, "almacen_v2.html")
FLEETCARE_PATH = os.path.join(BASE_DIR, "fleetcare.html")
BASE_DATOS_PATH = os.path.join(BASE_DIR, "base_datos.html")
DOC_LOG_VIAJES_DIR = os.path.join(DATA_DIR, "Doc_Log_Viajes")
DOC_ALMACEN_DIR = os.path.join(DATA_DIR, "Doc_Almacen")
DOC_ALMACEN_ADJ_DIR = os.path.join(DOC_ALMACEN_DIR, "adjuntos")
SQLITE_DB_PATH = os.path.join(DATA_DIR, "dashboard.db")
SCHEMA_PATH = os.path.join(BASE_DIR, "schema.sql")
MEMBRETE_LOGO_PATH = os.path.join(BASE_DIR, "Imagenes", "09-smg.png")

EMPRESA_MEMBRETE = {
    "razon_social": "S.M.G. S.R.L.",
    "telefono": "0388-4052871",
    "codigo_postal": "4612",
    "direccion": "Santa Catalina N°551 B° (4612) Palpala Jujuy",
    "cuit": "30-71102383-2",
}

os.makedirs(DOC_LOG_VIAJES_DIR, exist_ok=True)
os.makedirs(DOC_ALMACEN_DIR, exist_ok=True)
os.makedirs(DOC_ALMACEN_ADJ_DIR, exist_ok=True)

app = FastAPI()
app.mount("/Imagenes", StaticFiles(directory=os.path.join(BASE_DIR, "Imagenes")), name="Imagenes")
app.mount("/Doc_Log_Viajes", StaticFiles(directory=DOC_LOG_VIAJES_DIR), name="Doc_Log_Viajes")
app.mount("/Doc_Almacen", StaticFiles(directory=DOC_ALMACEN_DIR), name="Doc_Almacen")
app.mount("/Doc_Almacen_Adjuntos", StaticFiles(directory=DOC_ALMACEN_ADJ_DIR), name="Doc_Almacen_Adjuntos")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

AUTH_USER = os.getenv("APP_LOGIN_USER", "admin")
AUTH_PASS = os.getenv("APP_LOGIN_PASSWORD", "admin123")
AUTH_COOKIE_NAME = "dash_auth_user"

# Cifrado reversible de contraseñas para que un ADMINISTRADOR pueda consultar
# la contraseña original de un usuario desde el módulo Administración. Esto es
# ADEMÁS del hash SHA-256 (que sigue siendo lo único usado para validar el
# login); si PASSWORD_ENCRYPTION_KEY no está configurada, la función de "ver
# contraseña" queda deshabilitada pero el login sigue funcionando normalmente.
PASSWORD_ENCRYPTION_KEY = os.getenv("PASSWORD_ENCRYPTION_KEY", "").strip()
try:
    _password_cipher = Fernet(PASSWORD_ENCRYPTION_KEY.encode("utf-8")) if PASSWORD_ENCRYPTION_KEY else None
except ValueError:
    _password_cipher = None


def _cifrar_password(password: str) -> str:
    if _password_cipher is None:
        return ""
    return _password_cipher.encrypt(password.encode("utf-8")).decode("utf-8")


def _descifrar_password(password_cifrada: str):
    if _password_cipher is None or not password_cifrada:
        return None
    try:
        return _password_cipher.decrypt(password_cifrada.encode("utf-8")).decode("utf-8")
    except InvalidToken:
        return None


SESSION_IDLE_SECONDS = 60 * 60
ALL_MODULES = [
    "logistica", "almacen", "operaciones", "compras", "rrhh",
    "mantenimiento", "dashboard_ejecutivo", "administracion", "base_datos",
]


def _usuario_autenticado(request: Request):
    """Devuelve el perfil de la sesión vigente o None si no es válida."""
    usuario = str(request.cookies.get(AUTH_COOKIE_NAME, "")).strip()
    if not usuario:
        return None

    # Cuenta de contingencia configurada por variables de entorno.
    if usuario == AUTH_USER:
        return {
            "usuario": AUTH_USER,
            "nombre_apellido": "Administrador",
            "tipo_usuario": "ADMINISTRADOR",
            "modulos": ALL_MODULES,
        }

    try:
        with get_sqlite_connection() as conn:
            fila = conn.execute(
                """
                SELECT correo, nombre_apellido, tipo_usuario, modulos_json
                FROM usuarios
                WHERE lower(correo) = lower(?)
                  AND upper(COALESCE(estado, 'ACTIVO')) = 'ACTIVO'
                  AND COALESCE(bloqueado, 0) = 0
                """,
                (usuario,),
            ).fetchone()
    except sqlite3.Error:
        return None

    if not fila:
        return None
    return {
        "usuario": fila["correo"],
        "nombre_apellido": fila["nombre_apellido"],
        "tipo_usuario": fila["tipo_usuario"],
        "modulos": parse_json_list(fila["modulos_json"]),
    }


def _requiere_login(request: Request):
    if _usuario_autenticado(request) is not None:
        return None
    return RedirectResponse("/login", status_code=302)


def _requiere_administrador(request: Request):
    perfil = _usuario_autenticado(request)
    if perfil is None or str(perfil.get("tipo_usuario", "")).upper() != "ADMINISTRADOR":
        raise HTTPException(status_code=403, detail="Se requiere perfil ADMINISTRADOR")
    return perfil


def _leer_html(path: str):
    with open(path, "r", encoding="utf-8") as f:
        html = f.read()
    if path == LOGIN_PATH:
        return html
    session_guard = f"""
<script>
(() => {{
  const idleMs = {SESSION_IDLE_SECONDS * 1000};
  const refreshMs = 60 * 1000;
  let timeoutId;
  let lastServerTouch = 0;
  let closed = false;
  const cerrarSesion = () => {{
    if (closed) return;
    closed = true;
    window.location.replace('/logout');
  }};
  const registrarActividad = () => {{
    if (closed) return;
    clearTimeout(timeoutId);
    timeoutId = window.setTimeout(cerrarSesion, idleMs);
    if (Date.now() - lastServerTouch >= refreshMs) {{
      lastServerTouch = Date.now();
      fetch('/session/actividad', {{ credentials: 'same-origin' }})
        .then((response) => {{ if (!response.ok) cerrarSesion(); }})
        .catch(() => {{}});
    }}
  }};
  ['pointerdown', 'mousemove', 'keydown', 'scroll', 'touchstart'].forEach((eventName) =>
    window.addEventListener(eventName, registrarActividad, {{ passive: true }}));
  registrarActividad();
}})();
</script>
"""
    return html.replace("</body>", f"{session_guard}</body>")


@app.middleware("http")
async def renovar_sesion_por_actividad(request: Request, call_next):
    response = await call_next(request)
    if request.url.path != "/logout" and _usuario_autenticado(request) is not None:
        response.set_cookie(
            AUTH_COOKIE_NAME,
            request.cookies.get(AUTH_COOKIE_NAME, ""),
            max_age=SESSION_IDLE_SECONDS,
            httponly=True,
            samesite="lax",
        )
    return response

def normalizar_texto(texto):
    if texto is None:
        return ""
    tabla = str.maketrans("áéíóúÁÉÍÓÚñÑ", "aeiouAEIOUnN")
    return str(texto).strip().translate(tabla).lower()


def obtener_campo_habilitacion(destino):
    destino_norm = normalizar_texto(destino)
    mapa = {
        "pirquitas": "habilitacion_pirquitas",
        "exar": "habilitacion_exar",
        "sdj": "habilitacion_sdj",
        "rincon": "habilitacion_rincon",
        "arli": "habilitacion_arli",
    }
    for clave, campo in mapa.items():
        if clave in destino_norm:
            return campo
    return None


def esta_habilitado(persona, campo_habilitacion):
    if campo_habilitacion is None:
        return True
    valor = str(persona.get(campo_habilitacion, "")).strip().upper()
    return valor in {"SI", "S", "OK", "X", "1", "TRUE", "HABILITADO"}


def parse_json_dict(texto, default=None):
    if default is None:
        default = {}
    if not texto:
        return default
    try:
        data = json.loads(texto)
        return data if isinstance(data, dict) else default
    except Exception:
        return default


def parse_json_list(texto):
    if not texto:
        return []
    try:
        data = json.loads(texto)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def get_sqlite_connection():
    conn = sqlite3.connect(SQLITE_DB_PATH, timeout=20)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 20000")
    conn.execute("PRAGMA synchronous = NORMAL")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _payload_value(payload, key, default=""):
    if payload is None:
        return default
    if isinstance(payload, dict):
        return payload.get(key, default)
    return getattr(payload, key, default)


def admin_crear_usuario(payload):
    conn = get_sqlite_connection()
    try:
        nombre_apellido = str(_payload_value(payload, "nombre_apellido", "")).strip()
        dni = str(_payload_value(payload, "dni", "")).strip()
        legajo = str(_payload_value(payload, "legajo", "")).strip()
        correo = str(_payload_value(payload, "correo", "")).strip()
        password = str(_payload_value(payload, "password", "")).strip()
        estado = str(_payload_value(payload, "estado", "ACTIVO")).strip().upper() or "ACTIVO"
        tipo_usuario = str(_payload_value(payload, "tipo_usuario", "CONSULTOR")).strip().upper() or "CONSULTOR"
        modulos = _payload_value(payload, "modulos", [])
        paneles = _payload_value(payload, "paneles", {})
        acciones = _payload_value(payload, "acciones", {})
        roles = _payload_value(payload, "roles", [])

        if not nombre_apellido or not correo:
            return {"ok": False, "error": "nombre_apellido y correo son obligatorios"}

        modulos_json = json.dumps(modulos if isinstance(modulos, list) else [], ensure_ascii=False)
        paneles_json = json.dumps(paneles if isinstance(paneles, dict) else {}, ensure_ascii=False)
        acciones_json = json.dumps(acciones if isinstance(acciones, dict) else {}, ensure_ascii=False)
        password_hash = hashlib.sha256(password.encode("utf-8")).hexdigest()
        password_cifrada = _cifrar_password(password)

        cur = conn.execute(
            """
            INSERT INTO usuarios (
                nombre_apellido,
                dni,
                legajo,
                correo,
                password_hash,
                password_cifrada,
                estado,
                tipo_usuario,
                modulos_json,
                paneles_json,
                acciones_json,
                bloqueado,
                intentos_fallidos,
                password_temporal,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, 1, datetime('now'), datetime('now'))
            """,
            (
                nombre_apellido,
                dni,
                legajo,
                correo,
                password_hash,
                password_cifrada,
                estado,
                tipo_usuario,
                modulos_json,
                paneles_json,
                acciones_json,
            ),
        )
        usuario_id = cur.lastrowid
        for rol_id in roles if isinstance(roles, list) else []:
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO usuario_roles_funcionales (usuario_id, rol_id) VALUES (?, ?)",
                    (usuario_id, int(rol_id)),
                )
            except (TypeError, ValueError):
                continue
        conn.commit()
        return {"ok": True, "id": usuario_id}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    finally:
        conn.close()


def migrar_tabla_viajes(conn):
    columnas_esperadas = [
        "id",
        "solicitante",
        "area",
        "origen",
        "destino",
        "motivo",
        "fecha_salida",
        "fecha_regreso",
        "chofer",
        "vehiculo",
        "acompanantes",
        "alojamiento",
        "estado",
        "fecha_creacion",
        "orden_salida_generada",
        "nro_orden_salida",
        "raw_json",
    ]
    cursor_columnas = conn.execute("PRAGMA table_info(viajes)")
    columnas_actuales = [row[1] for row in cursor_columnas.fetchall()]
    cursor_columnas.close()
    if columnas_actuales == columnas_esperadas:
        return

    cursor_filas = conn.execute("SELECT * FROM viajes")
    filas = cursor_filas.fetchall()
    cursor_filas.close()
    registros = []
    for fila in filas:
        registro = dict(fila)
        viaje_raw = parse_json_dict(registro.get("raw_json"), default={})

        acompanantes = registro.get("acompanantes")
        if acompanantes is None:
            acompanantes = viaje_raw.get("acompanantes")
        if isinstance(acompanantes, list):
            acompanantes = json.dumps(acompanantes, ensure_ascii=False)

        registros.append(
            (
                registro.get("id"),
                registro.get("solicitante"),
                registro.get("area"),
                registro.get("origen"),
                registro.get("destino"),
                registro.get("motivo"),
                registro.get("fecha_salida"),
                registro.get("fecha_regreso"),
                registro.get("chofer") if "chofer" in registro else viaje_raw.get("chofer"),
                registro.get("vehiculo") if "vehiculo" in registro else viaje_raw.get("vehiculo"),
                acompanantes,
                registro.get("alojamiento") if "alojamiento" in registro else viaje_raw.get("alojamiento"),
                registro.get("estado"),
                registro.get("fecha_creacion"),
                registro.get("orden_salida_generada", 0),
                registro.get("nro_orden_salida"),
                registro.get("raw_json"),
            )
        )

    conn.commit()
    conn.execute("PRAGMA foreign_keys = OFF")
    conn.execute(
        """
        CREATE TABLE viajes_new (
            id INTEGER PRIMARY KEY,
            solicitante TEXT,
            area TEXT,
            origen TEXT,
            destino TEXT,
            motivo TEXT,
            fecha_salida TEXT,
            fecha_regreso TEXT,
            chofer TEXT,
            vehiculo TEXT,
            acompanantes TEXT,
            alojamiento TEXT,
            estado TEXT,
            fecha_creacion TEXT,
            orden_salida_generada INTEGER DEFAULT 0,
            nro_orden_salida TEXT,
            raw_json TEXT
        )
        """
    )
    conn.executemany(
        """
        INSERT INTO viajes_new (
            id, solicitante, area, origen, destino, motivo,
            fecha_salida, fecha_regreso, chofer, vehiculo, acompanantes, alojamiento,
            estado, fecha_creacion, orden_salida_generada, nro_orden_salida, raw_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        registros,
    )
    conn.execute("DROP TABLE viajes")
    conn.execute("ALTER TABLE viajes_new RENAME TO viajes")
    conn.execute("PRAGMA foreign_keys = ON")


def sembrar_configuracion_almacen(conn):
    categorias = [
        "Tecnologia",
        "Herramientas",
        "EPP",
        "Vehiculos",
        "Repuestos",
        "Seguridad",
        "Oficina",
        "Consumibles",
    ]
    tipos = [
        "Consumible",
        "Permanente",
        "Herramienta",
        "Electronico",
        "Activo Fijo",
        "Repuesto",
    ]
    unidades = ["Unidad", "Caja", "Bolsa", "Par", "Metro", "Litro", "Kg", "Juego"]
    roles = [
        ("Administrador", '{"all": true}'),
        ("Supervisor", '{"aprobar_ajustes": true, "inventarios": true, "reportes": true}'),
        ("Operador Almacen", '{"ri": true, "re": true, "stock": true, "movimientos": true}'),
        ("Consulta", '{"solo_lectura": true}'),
    ]
    familias = [
        "Monitoreo",
        "Video",
        "Conectividad",
        "Electrico",
        "Seguridad",
    ]
    marcas = ["Hikvision", "Samsung", "Dell", "Lenovo", "Genrica"]
    modelos = [
        ("Hikvision", "DS-7108"),
        ("Samsung", "SSD 1TB"),
        ("Dell", "Latitude 5420"),
        ("Lenovo", "Tab M10"),
    ]
    proyectos = [
        ("EXAR", "Exar", "Exar"),
        ("PIRQUITAS", "Pirquitas", "Pirquitas"),
        ("SDJ", "SDJ", "SDJ"),
        ("ARLI", "ARLI", "ARLI"),
        ("RINCON", "Rincon", "Rincon"),
    ]
    instalaciones = [
        ("EXAR", "PLANTA", "Planta"),
        ("EXAR", "TALLER", "Taller"),
        ("PIRQUITAS", "BASE", "Base Operativa"),
        ("SDJ", "CAMP", "Campamento"),
    ]

    for nombre in categorias:
        conn.execute("INSERT OR IGNORE INTO categorias (nombre) VALUES (?)", (nombre,))
    for nombre in tipos:
        conn.execute("INSERT OR IGNORE INTO tipos_producto (nombre, activo) VALUES (?, 1)", (nombre,))
    for nombre in unidades:
        conn.execute("INSERT OR IGNORE INTO unidades_medida (nombre, activo) VALUES (?, 1)", (nombre,))
    for nombre, permisos_json in roles:
        conn.execute(
            "INSERT OR IGNORE INTO roles_almacen (nombre, permisos_json, activo) VALUES (?, ?, 1)",
            (nombre, permisos_json),
        )

    for familia in familias:
        conn.execute("INSERT OR IGNORE INTO familias (nombre) VALUES (?)", (familia,))

    for nombre in marcas:
        conn.execute("INSERT OR IGNORE INTO marcas (nombre, activo) VALUES (?, 1)", (nombre,))

    marca_map = {
        r["nombre"]: r["id"]
        for r in conn.execute("SELECT id, nombre FROM marcas").fetchall()
    }
    for marca, modelo in modelos:
        conn.execute(
            "INSERT OR IGNORE INTO modelos (id_marca, nombre, activo) VALUES (?, ?, 1)",
            (marca_map.get(marca), modelo),
        )

    for codigo, nombre, cliente in proyectos:
        conn.execute(
            "INSERT OR IGNORE INTO proyectos (codigo, nombre, cliente, activo) VALUES (?, ?, ?, 1)",
            (codigo, nombre, cliente),
        )

    proyecto_map = {
        r["codigo"]: r["id"]
        for r in conn.execute("SELECT id, codigo FROM proyectos").fetchall()
    }
    for codigo_proy, codigo_inst, nombre_inst in instalaciones:
        conn.execute(
            """
            INSERT OR IGNORE INTO instalaciones (id_proyecto, codigo, nombre, activo)
            VALUES (?, ?, ?, 1)
            """,
            (proyecto_map.get(codigo_proy), codigo_inst, nombre_inst),
        )


def migrar_almacen_v2(conn):
    columnas = {r[1] for r in conn.execute("PRAGMA table_info(productos)").fetchall()}
    columnas_nuevas = [
        ("id_familia", "INTEGER"),
        ("id_marca", "INTEGER"),
        ("id_modelo", "INTEGER"),
        ("id_unidad", "INTEGER"),
        ("tipo_control", "TEXT DEFAULT 'CONSUMIBLE'"),
    ]
    for nombre_col, tipo_col in columnas_nuevas:
        if nombre_col not in columnas:
            conn.execute(f"ALTER TABLE productos ADD COLUMN {nombre_col} {tipo_col}")

    # Compatibilidad hacia nuevo campo id_unidad
    if "unidad_medida_id" in columnas and "id_unidad" in {r[1] for r in conn.execute("PRAGMA table_info(productos)").fetchall()}:
        conn.execute(
            """
            UPDATE productos
               SET id_unidad = COALESCE(id_unidad, unidad_medida_id)
             WHERE id_unidad IS NULL
            """
        )

    conn.execute(
        """
        UPDATE productos
           SET tipo_control = COALESCE(tipo_control, 'CONSUMIBLE')
        """
    )

    columnas_actuales = [r[1] for r in conn.execute("PRAGMA table_info(productos)").fetchall()]
    columnas_a_eliminar = {"controla_serie", "activo", "compat_vehiculos", "compat_proyectos"}
    if any(c in columnas_a_eliminar for c in columnas_actuales):
        conn.commit()
        conn.execute("PRAGMA foreign_keys = OFF")
        conn.execute(
            """
            CREATE TABLE productos_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL UNIQUE,
                descripcion TEXT NOT NULL,
                marca TEXT,
                modelo TEXT,
                categoria_id INTEGER,
                tipo_producto_id INTEGER,
                unidad_medida_id INTEGER,
                stock_minimo REAL DEFAULT 0,
                stock_maximo REAL DEFAULT 0,
                punto_reposicion REAL DEFAULT 0,
                ubicacion_id INTEGER,
                observaciones TEXT,
                fecha_alta TEXT,
                usuario_alta TEXT,
                raw_json TEXT,
                id_familia INTEGER,
                id_marca INTEGER,
                id_modelo INTEGER,
                id_unidad INTEGER,
                tipo_control TEXT DEFAULT 'CONSUMIBLE',
                FOREIGN KEY (categoria_id) REFERENCES categorias (id) ON DELETE SET NULL,
                FOREIGN KEY (tipo_producto_id) REFERENCES tipos_producto (id) ON DELETE SET NULL,
                FOREIGN KEY (unidad_medida_id) REFERENCES unidades_medida (id) ON DELETE SET NULL,
                FOREIGN KEY (ubicacion_id) REFERENCES ubicaciones (id) ON DELETE SET NULL
            )
            """
        )
        conn.execute(
            """
            INSERT INTO productos_new (
                id, codigo, descripcion, marca, modelo,
                categoria_id, tipo_producto_id, unidad_medida_id,
                stock_minimo, stock_maximo, punto_reposicion,
                ubicacion_id, observaciones, fecha_alta, usuario_alta, raw_json,
                id_familia, id_marca, id_modelo, id_unidad, tipo_control
            )
            SELECT
                id, codigo, descripcion, marca, modelo,
                categoria_id, tipo_producto_id, unidad_medida_id,
                COALESCE(stock_minimo, 0), COALESCE(stock_maximo, 0), COALESCE(punto_reposicion, 0),
                ubicacion_id, observaciones, fecha_alta, usuario_alta, raw_json,
                id_familia, id_marca, id_modelo, COALESCE(id_unidad, unidad_medida_id), COALESCE(tipo_control, 'CONSUMIBLE')
            FROM productos
            """
        )
        conn.execute("DROP TABLE productos")
        conn.execute("ALTER TABLE productos_new RENAME TO productos")
        conn.execute("PRAGMA foreign_keys = ON")


def migrar_catalogos_v2(conn):
    conn.commit()
    conn.execute("PRAGMA foreign_keys = OFF")

    cols_cat = [r[1] for r in conn.execute("PRAGMA table_info(categorias)").fetchall()]
    if cols_cat != ["id", "nombre"]:
        rows = conn.execute("SELECT id, nombre FROM categorias ORDER BY id").fetchall()
        conn.execute(
            """
            CREATE TABLE categorias_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE
            )
            """
        )
        conn.executemany(
            "INSERT OR IGNORE INTO categorias_new (id, nombre) VALUES (?, ?)",
            [(r["id"], r["nombre"]) for r in rows],
        )
        conn.execute("DROP TABLE categorias")
        conn.execute("ALTER TABLE categorias_new RENAME TO categorias")

    cols_fam = [r[1] for r in conn.execute("PRAGMA table_info(familias)").fetchall()]
    if cols_fam != ["id", "nombre"]:
        rows = conn.execute("SELECT id, nombre FROM familias ORDER BY id").fetchall()
        conn.execute(
            """
            CREATE TABLE familias_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE
            )
            """
        )
        conn.executemany(
            "INSERT OR IGNORE INTO familias_new (id, nombre) VALUES (?, ?)",
            [(r["id"], r["nombre"]) for r in rows],
        )
        conn.execute("DROP TABLE familias")
        conn.execute("ALTER TABLE familias_new RENAME TO familias")

    conn.execute("DROP INDEX IF EXISTS idx_familias_categoria")
    conn.execute("PRAGMA foreign_keys = ON")


def migrar_eliminar_activos_v2(conn):
    tablas = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    if "activos" not in tablas and "instalaciones_activos" not in tablas and "retiros_activos" not in tablas:
        return

    conn.commit()
    conn.execute("PRAGMA foreign_keys = OFF")

    cols_doc = {r[1] for r in conn.execute("PRAGMA table_info(documentos_detalle)").fetchall()}
    if "id_activo" in cols_doc:
        conn.execute(
            """
            CREATE TABLE documentos_detalle_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_documento INTEGER NOT NULL,
                id_producto INTEGER,
                cantidad REAL DEFAULT 0,
                id_vehiculo TEXT,
                id_proyecto INTEGER,
                id_instalacion INTEGER,
                observaciones TEXT,
                FOREIGN KEY (id_documento) REFERENCES documentos (id) ON DELETE CASCADE,
                FOREIGN KEY (id_producto) REFERENCES productos (id) ON DELETE RESTRICT,
                FOREIGN KEY (id_vehiculo) REFERENCES vehiculos (codigo) ON DELETE SET NULL,
                FOREIGN KEY (id_proyecto) REFERENCES proyectos (id) ON DELETE SET NULL,
                FOREIGN KEY (id_instalacion) REFERENCES instalaciones (id) ON DELETE SET NULL
            )
            """
        )
        conn.execute(
            """
            INSERT INTO documentos_detalle_new (
                id, id_documento, id_producto, cantidad,
                id_vehiculo, id_proyecto, id_instalacion, observaciones
            )
            SELECT
                id, id_documento, id_producto, cantidad,
                id_vehiculo, id_proyecto, id_instalacion, observaciones
            FROM documentos_detalle
            """
        )
        conn.execute("DROP TABLE documentos_detalle")
        conn.execute("ALTER TABLE documentos_detalle_new RENAME TO documentos_detalle")

    cols_mov = {r[1] for r in conn.execute("PRAGMA table_info(movimientos)").fetchall()}
    if "id_activo" in cols_mov:
        conn.execute(
            """
            CREATE TABLE movimientos_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                tipo TEXT NOT NULL,
                documento TEXT,
                id_documento INTEGER,
                id_producto INTEGER,
                cantidad REAL DEFAULT 0,
                stock_anterior REAL DEFAULT 0,
                stock_nuevo REAL DEFAULT 0,
                id_personal TEXT,
                id_vehiculo TEXT,
                id_proyecto INTEGER,
                id_instalacion INTEGER,
                observaciones TEXT,
                raw_json TEXT,
                FOREIGN KEY (id_documento) REFERENCES documentos (id) ON DELETE SET NULL,
                FOREIGN KEY (id_producto) REFERENCES productos (id) ON DELETE RESTRICT,
                FOREIGN KEY (id_personal) REFERENCES personal (legajo) ON DELETE SET NULL,
                FOREIGN KEY (id_vehiculo) REFERENCES vehiculos (codigo) ON DELETE SET NULL,
                FOREIGN KEY (id_proyecto) REFERENCES proyectos (id) ON DELETE SET NULL,
                FOREIGN KEY (id_instalacion) REFERENCES instalaciones (id) ON DELETE SET NULL
            )
            """
        )
        conn.execute(
            """
            INSERT INTO movimientos_new (
                id, fecha, tipo, documento, id_documento, id_producto,
                cantidad, stock_anterior, stock_nuevo,
                id_personal, id_vehiculo, id_proyecto, id_instalacion,
                observaciones, raw_json
            )
            SELECT
                id, fecha, tipo, documento, id_documento, id_producto,
                cantidad, stock_anterior, stock_nuevo,
                id_personal, id_vehiculo, id_proyecto, id_instalacion,
                observaciones, raw_json
            FROM movimientos_stock
            """
        )
        conn.execute("DROP TABLE movimientos")
        conn.execute("ALTER TABLE movimientos_new RENAME TO movimientos")

    cols_adj = {r[1] for r in conn.execute("PRAGMA table_info(adjuntos)").fetchall()}
    if "id_activo" in cols_adj:
        conn.execute(
            """
            CREATE TABLE adjuntos_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                tipo TEXT NOT NULL,
                archivo TEXT NOT NULL,
                id_documento INTEGER,
                id_inventario INTEGER,
                id_ajuste INTEGER,
                FOREIGN KEY (id_documento) REFERENCES documentos (id) ON DELETE CASCADE,
                FOREIGN KEY (id_inventario) REFERENCES inventarios (id) ON DELETE CASCADE,
                FOREIGN KEY (id_ajuste) REFERENCES ajustes_stock (id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            INSERT INTO adjuntos_new (
                id, fecha, tipo, archivo, id_documento, id_inventario, id_ajuste
            )
            SELECT
                id, fecha, tipo, archivo, id_documento, id_inventario, id_ajuste
            FROM adjuntos
            """
        )
        conn.execute("DROP TABLE adjuntos")
        conn.execute("ALTER TABLE adjuntos_new RENAME TO adjuntos")

    conn.execute("DROP TABLE IF EXISTS retiros_activos")
    conn.execute("DROP TABLE IF EXISTS instalaciones_activos")
    conn.execute("DROP TABLE IF EXISTS activos")
    conn.execute("DROP TABLE IF EXISTS productos_series")

    conn.execute("DROP INDEX IF EXISTS idx_activos_producto")
    conn.execute("DROP INDEX IF EXISTS idx_activos_estado")
    conn.execute("DROP INDEX IF EXISTS idx_mov_v2_activo")
    conn.execute("DROP INDEX IF EXISTS idx_inst_activo")
    conn.execute("DROP INDEX IF EXISTS idx_retiros_inst")
    conn.execute("DROP INDEX IF EXISTS idx_series_producto")
    conn.execute("DROP INDEX IF EXISTS idx_series_estado")

    conn.execute("CREATE INDEX IF NOT EXISTS idx_doc_detalle_doc ON documentos_detalle (id_documento)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mov_v2_fecha ON movimientos (fecha)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mov_v2_producto ON movimientos (id_producto)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_adj_doc ON adjuntos (id_documento)")

    conn.execute("PRAGMA foreign_keys = ON")


def migrar_remitos_ingreso_refs(conn):
    columnas = {r[1] for r in conn.execute("PRAGMA table_info(remitos_ingreso)").fetchall()}
    if "nro_remito_referencia" not in columnas:
        conn.execute("ALTER TABLE remitos_ingreso ADD COLUMN nro_remito_referencia TEXT")


def migrar_remitos_entrega_autorizacion(conn):
    columnas = {r[1] for r in conn.execute("PRAGMA table_info(remitos_entrega)").fetchall()}
    nuevas = {
        "estado_autorizacion": "TEXT DEFAULT 'AUTORIZADO'",
        "fecha_autorizacion": "TEXT",
        "autorizado_por": "TEXT",
        "observaciones_autorizacion": "TEXT",
    }
    for nombre_col, tipo_col in nuevas.items():
        if nombre_col not in columnas:
            conn.execute(f"ALTER TABLE remitos_entrega ADD COLUMN {nombre_col} {tipo_col}")

    conn.execute(
        """
        UPDATE remitos_entrega
           SET estado_autorizacion = COALESCE(NULLIF(TRIM(estado_autorizacion), ''), 'AUTORIZADO')
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_re_estado_autorizacion ON remitos_entrega (estado_autorizacion)")


def migrar_gestion_operativa(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestion_operativa (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nro_orden TEXT,
            fecha_orden TEXT,
            estado_orden TEXT,
            id_viaje INTEGER,
            centro_costo TEXT,
            proyecto TEXT,
            origen TEXT,
            destino TEXT,
            fecha_salida TEXT,
            fecha_regreso TEXT,
            jornadas INTEGER,
            legajo TEXT,
            nombre TEXT,
            rol TEXT,
            vehiculo TEXT,
            chofer TEXT,
            viatico REAL DEFAULT 0,
            horas_totales REAL DEFAULT 0,
            horas_normales REAL DEFAULT 0,
            horas_compensables REAL DEFAULT 0,
            costo_total REAL DEFAULT 0,
            fecha_cierre TEXT
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_go_nro_orden ON gestion_operativa (nro_orden)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_go_fecha_orden ON gestion_operativa (fecha_orden)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_go_nombre ON gestion_operativa (nombre)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_go_proyecto ON gestion_operativa (proyecto)")


def migrar_usuarios_administracion(conn):
    """Añade la configuración granular usada por el módulo Administración."""
    columnas = {r[1] for r in conn.execute("PRAGMA table_info(usuarios)").fetchall()}
    for nombre, definicion in {
        "paneles_json": "TEXT",
        "acciones_json": "TEXT",
        "password_cifrada": "TEXT",
    }.items():
        if nombre not in columnas:
            conn.execute(f"ALTER TABLE usuarios ADD COLUMN {nombre} {definicion}")
    for nombre in ("ADMINISTRADOR", "SUPERVISOR", "ASISTENTE", "OPERADOR", "CONSULTOR"):
        conn.execute(
            "INSERT OR IGNORE INTO roles_funcionales (nombre, activo) VALUES (?, 1)",
            (nombre,),
        )


def migrar_personal_dotacion(conn):
    """Incorpora los campos de dotación para instalaciones ya existentes."""
    columnas = {r[1] for r in conn.execute("PRAGMA table_info(personal)").fetchall()}
    columnas_nuevas = {
        "fecha_nacimiento": "TEXT",
        "activo": "TEXT DEFAULT 'ACTIVO'",
    }
    agregadas = []
    for nombre, definicion in columnas_nuevas.items():
        if nombre not in columnas:
            conn.execute(f"ALTER TABLE personal ADD COLUMN {nombre} {definicion}")
            agregadas.append(nombre)

    if not agregadas:
        return

    # Si estos campos ya se habian guardado antes (por ejemplo, via una version
    # anterior que solo los persistia dentro de raw_json), se recuperan de ahi
    # para no perder datos al agregar recien ahora las columnas dedicadas.
    filas = conn.execute("SELECT legajo, raw_json FROM personal").fetchall()
    for fila in filas:
        try:
            datos = json.loads(fila["raw_json"] or "{}")
        except Exception:
            continue
        if not isinstance(datos, dict):
            continue
        sets, valores = [], []
        for campo in agregadas:
            if datos.get(campo):
                sets.append(f"{campo} = ?")
                valores.append(datos[campo])
        if sets:
            valores.append(fila["legajo"])
            conn.execute(f"UPDATE personal SET {', '.join(sets)} WHERE legajo = ?", valores)


PROYECTOS_HABILITACION_DEFAULT = [
    ("Pirquitas", "pirquitas"),
    ("Exar", "exar"),
    ("SDJ", "sdj"),
    ("Rincón", "rincon"),
    ("ARLI", "arli"),
]


def migrar_habilitaciones_personal(conn):
    """Crea el catalogo de proyectos y migra las habilitaciones SI/NO existentes
    a personal_habilitaciones, con una fecha de vencimiento placeholder, la
    primera vez que se detecta que la tabla nueva esta vacia."""
    for nombre, clave in PROYECTOS_HABILITACION_DEFAULT:
        conn.execute(
            "INSERT OR IGNORE INTO proyectos_habilitacion (nombre, clave) VALUES (?, ?)",
            (nombre, clave),
        )

    ya_migrado = conn.execute("SELECT COUNT(*) FROM personal_habilitaciones").fetchone()[0]
    if ya_migrado:
        return

    proyectos = {
        clave: pid for pid, clave in conn.execute(
            "SELECT id, clave FROM proyectos_habilitacion"
        ).fetchall()
    }
    columnas_legacy = [
        ("habilitacion_pirquitas", "pirquitas"),
        ("habilitacion_exar", "exar"),
        ("habilitacion_sdj", "sdj"),
        ("habilitacion_rincon", "rincon"),
        ("habilitacion_arli", "arli"),
    ]
    filas = conn.execute(
        "SELECT legajo, habilitacion_pirquitas, habilitacion_exar, habilitacion_sdj, "
        "habilitacion_rincon, habilitacion_arli FROM personal"
    ).fetchall()
    for fila in filas:
        for indice, (_columna, clave) in enumerate(columnas_legacy, start=1):
            valor = str(fila[indice] or "").strip().upper()
            if valor == "SI" and clave in proyectos:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO personal_habilitaciones (legajo, proyecto_id, fecha_vencimiento)
                    VALUES (?, ?, ?)
                    """,
                    (fila["legajo"], proyectos[clave], "2026-08-30"),
                )


def init_sqlite():
    if not os.path.exists(SCHEMA_PATH):
        return
    max_intentos = 6
    espera_segundos = 0.6
    for intento in range(1, max_intentos + 1):
        try:
            with get_sqlite_connection() as conn:
                with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
                    conn.executescript(f.read())
                try:
                    migrar_tabla_viajes(conn)
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        raise
                    print("No se pudo migrar la tabla viajes porque dashboard.db esta bloqueada por otro proceso.")
                migrar_almacen_v2(conn)
                try:
                    migrar_catalogos_v2(conn)
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        raise
                    print("No se pudo migrar categorias porque dashboard.db esta bloqueada por otro proceso.")
                try:
                    migrar_eliminar_activos_v2(conn)
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        raise
                    print("No se pudo eliminar tablas de activos porque dashboard.db esta bloqueada por otro proceso.")
                try:
                    migrar_remitos_ingreso_refs(conn)
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        raise
                    print("No se pudo migrar remitos de ingreso porque dashboard.db esta bloqueada por otro proceso.")
                try:
                    migrar_remitos_entrega_autorizacion(conn)
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        raise
                    print("No se pudo migrar autorizacion de remitos de entrega porque dashboard.db esta bloqueada por otro proceso.")
                try:
                    migrar_gestion_operativa(conn)
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        raise
                    print("No se pudo migrar gestion_operativa porque dashboard.db esta bloqueada por otro proceso.")
                migrar_usuarios_administracion(conn)
                migrar_personal_dotacion(conn)
                migrar_habilitaciones_personal(conn)
                sembrar_configuracion_almacen(conn)
                try:
                    conn.commit()
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        raise
                    print("Init SQLite: commit omitido temporalmente por bloqueo de dashboard.db; reintentara en el siguiente reload.")
            return
        except sqlite3.OperationalError as exc:
            if "database is locked" not in str(exc).lower():
                raise
            if intento == max_intentos:
                print("Init SQLite: dashboard.db sigue bloqueada tras varios reintentos; se continuara sin abortar el arranque.")
                return
            time.sleep(espera_segundos)


def obtener_vehiculos_data():
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT codigo, propiedad, marca, tipo, modelo, dominio, anio, motor, chasis,
                   sector, proyecto, operativo,
                   habilitacion_pirquitas, habilitacion_exar, habilitacion_sdj,
                   habilitacion_rincon, habilitacion_arli
            FROM vehiculos
            ORDER BY codigo
            """
        ).fetchall()
    return [dict(r) for r in rows]


def obtener_choferes_data():
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            "SELECT nombre, estado FROM choferes WHERE UPPER(COALESCE(estado, '')) = 'ACTIVO' ORDER BY nombre"
        ).fetchall()
    return [dict(r) for r in rows]


def _estado_habilitacion(fecha_vencimiento, dias_por_vencer=30):
    if not fecha_vencimiento:
        return None
    try:
        vencimiento = datetime.strptime(str(fecha_vencimiento)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None
    hoy = datetime.now().date()
    if vencimiento < hoy:
        return "VENCIDO"
    if (vencimiento - hoy).days <= dias_por_vencer:
        return "POR_VENCER"
    return "VIGENTE"


def obtener_proyectos_habilitacion(conn):
    return [
        dict(r) for r in conn.execute(
            "SELECT id, nombre, clave FROM proyectos_habilitacion WHERE COALESCE(activo, 1) = 1 ORDER BY nombre"
        ).fetchall()
    ]


def obtener_habilitaciones_por_legajo(conn, legajo=None):
    query = """
        SELECT ph.legajo, p.clave, ph.fecha_vencimiento
        FROM personal_habilitaciones ph
        JOIN proyectos_habilitacion p ON p.id = ph.proyecto_id
    """
    params = ()
    if legajo is not None:
        query += " WHERE ph.legajo = ?"
        params = (str(legajo),)
    resultado = {}
    for fila in conn.execute(query, params).fetchall():
        resultado.setdefault(fila["legajo"], {})[fila["clave"]] = {
            "vencimiento": fila["fecha_vencimiento"],
            "estado": _estado_habilitacion(fila["fecha_vencimiento"]),
        }
    return resultado


def obtener_personal_data():
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT legajo, nombre, cuil, fecha_nacimiento, activo
            FROM personal
            ORDER BY CAST(legajo AS INTEGER), legajo
            """
        ).fetchall()
        habilitaciones = obtener_habilitaciones_por_legajo(conn)
    personal = [dict(r) for r in rows]
    for persona in personal:
        persona["habilitaciones"] = habilitaciones.get(str(persona["legajo"]), {})
    return personal


def esta_habilitado_personal(persona, clave_proyecto):
    hab = (persona.get("habilitaciones") or {}).get(clave_proyecto)
    if not hab:
        return False
    return hab.get("estado") in {"VIGENTE", "POR_VENCER"}


def obtener_viajes_data():
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, solicitante, area, origen, destino, motivo,
                   fecha_salida, fecha_regreso, estado, fecha_creacion,
                   orden_salida_generada, nro_orden_salida, raw_json
            FROM viajes
            ORDER BY id
            """
        ).fetchall()
        viajes = []
        for row in rows:
            viaje = parse_json_dict(row["raw_json"], default={})
            if not viaje:
                viaje = {
                    "id": row["id"],
                    "solicitante": row["solicitante"],
                    "area": row["area"],
                    "origen": row["origen"],
                    "destino": row["destino"],
                    "motivo": row["motivo"],
                    "fecha_salida": row["fecha_salida"],
                    "fecha_regreso": row["fecha_regreso"],
                    "estado": row["estado"],
                    "fecha_creacion": row["fecha_creacion"],
                    "orden_salida_generada": bool(row["orden_salida_generada"]),
                    "nro_orden_salida": row["nro_orden_salida"],
                }
            viajes.append(viaje)
        return viajes


def obtener_ordenes_data():
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT nro_orden, fecha_orden, id_viaje, estado, cierre_logistica_json, raw_json
            FROM ordenes_salida
            ORDER BY fecha_orden, nro_orden
            """
        ).fetchall()
        ordenes = []
        for row in rows:
            orden = parse_json_dict(row["raw_json"], default={})
            if not orden:
                orden = {
                    "nro_orden": row["nro_orden"],
                    "fecha_orden": row["fecha_orden"],
                    "id_viaje": row["id_viaje"],
                    "estado": row["estado"],
                }
            if row["cierre_logistica_json"]:
                orden["cierre_logistica"] = parse_json_dict(row["cierre_logistica_json"], default={})
            ordenes.append(orden)
        return ordenes


def generar_numero_orden_sql(conn):
    rows = conn.execute("SELECT nro_orden FROM ordenes_salida").fetchall()
    max_num = 0
    for row in rows:
        nro = str(row["nro_orden"] or "")
        if nro.startswith("OS-"):
            try:
                valor = int(nro.split("-")[1])
                max_num = max(max_num, valor)
            except (ValueError, IndexError):
                continue
    return f"OS-{max_num + 1:06d}"


def guardar_viaje_sql(conn, viaje_data):
    acompanantes = viaje_data.get("acompanantes")
    if isinstance(acompanantes, list):
        acompanantes = json.dumps(acompanantes, ensure_ascii=False)
    conn.execute(
        """
        INSERT INTO viajes (
            id, solicitante, area, origen, destino, motivo,
            fecha_salida, fecha_regreso, chofer, vehiculo, acompanantes, alojamiento,
            estado, fecha_creacion,
            orden_salida_generada, nro_orden_salida, raw_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            solicitante = excluded.solicitante,
            area = excluded.area,
            origen = excluded.origen,
            destino = excluded.destino,
            motivo = excluded.motivo,
            fecha_salida = excluded.fecha_salida,
            fecha_regreso = excluded.fecha_regreso,
            chofer = excluded.chofer,
            vehiculo = excluded.vehiculo,
            acompanantes = excluded.acompanantes,
            alojamiento = excluded.alojamiento,
            estado = excluded.estado,
            fecha_creacion = excluded.fecha_creacion,
            orden_salida_generada = excluded.orden_salida_generada,
            nro_orden_salida = excluded.nro_orden_salida,
            raw_json = excluded.raw_json
        """,
        (
            viaje_data.get("id"),
            viaje_data.get("solicitante"),
            viaje_data.get("area"),
            viaje_data.get("origen"),
            viaje_data.get("destino"),
            viaje_data.get("motivo"),
            viaje_data.get("fecha_salida"),
            viaje_data.get("fecha_regreso"),
            viaje_data.get("chofer"),
            viaje_data.get("vehiculo"),
            acompanantes,
            viaje_data.get("alojamiento"),
            viaje_data.get("estado"),
            viaje_data.get("fecha_creacion"),
            1 if viaje_data.get("orden_salida_generada") else 0,
            viaje_data.get("nro_orden_salida"),
            json.dumps(viaje_data, ensure_ascii=False),
        ),
    )


def guardar_recursos_sql(conn, id_viaje, recursos_data):
    if not isinstance(recursos_data, dict):
        recursos_data = {}

    conn.execute(
        """
        INSERT INTO recursos_viaje (
            id_viaje, fecha, centro_costo, datos_solicitante, area_solicitante, partida,
            destino, motivo_viaje, fecha_salida_viaje, fecha_regreso_viaje,
            hora_ingreso_base, hora_salida, hora_regreso, duracion_jornadas,
            itinerario, rutas, paradas, chofer, chofer_viatico, vehiculo,
            vehiculo_fuera_flota, viaticos, medio_pago, alojamiento, otros_gastos,
            verificado_administracion, comprobacion_operaciones_logistica_json, raw_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id_viaje) DO UPDATE SET
            fecha = excluded.fecha,
            centro_costo = excluded.centro_costo,
            datos_solicitante = excluded.datos_solicitante,
            area_solicitante = excluded.area_solicitante,
            partida = excluded.partida,
            destino = excluded.destino,
            motivo_viaje = excluded.motivo_viaje,
            fecha_salida_viaje = excluded.fecha_salida_viaje,
            fecha_regreso_viaje = excluded.fecha_regreso_viaje,
            hora_ingreso_base = excluded.hora_ingreso_base,
            hora_salida = excluded.hora_salida,
            hora_regreso = excluded.hora_regreso,
            duracion_jornadas = excluded.duracion_jornadas,
            itinerario = excluded.itinerario,
            rutas = excluded.rutas,
            paradas = excluded.paradas,
            chofer = excluded.chofer,
            chofer_viatico = excluded.chofer_viatico,
            vehiculo = excluded.vehiculo,
            vehiculo_fuera_flota = excluded.vehiculo_fuera_flota,
            viaticos = excluded.viaticos,
            medio_pago = excluded.medio_pago,
            alojamiento = excluded.alojamiento,
            otros_gastos = excluded.otros_gastos,
            verificado_administracion = excluded.verificado_administracion,
            comprobacion_operaciones_logistica_json = excluded.comprobacion_operaciones_logistica_json,
            raw_json = excluded.raw_json
        """,
        (
            id_viaje,
            recursos_data.get("fecha"),
            recursos_data.get("centro_costo"),
            recursos_data.get("datos_solicitante"),
            recursos_data.get("area_solicitante"),
            recursos_data.get("partida"),
            recursos_data.get("destino"),
            recursos_data.get("motivo_viaje"),
            recursos_data.get("fecha_salida_viaje"),
            recursos_data.get("fecha_regreso_viaje"),
            recursos_data.get("hora_ingreso_base"),
            recursos_data.get("hora_salida"),
            recursos_data.get("hora_regreso"),
            recursos_data.get("duracion_jornadas"),
            recursos_data.get("itinerario"),
            recursos_data.get("rutas"),
            recursos_data.get("paradas"),
            recursos_data.get("chofer"),
            float(recursos_data.get("chofer_viatico", 0) or 0),
            recursos_data.get("vehiculo"),
            1 if recursos_data.get("vehiculo_fuera_flota") else 0,
            float(recursos_data.get("viaticos", 0) or 0),
            recursos_data.get("medio_pago"),
            recursos_data.get("alojamiento"),
            float(recursos_data.get("otros_gastos", 0) or 0),
            recursos_data.get("verificado_administracion"),
            json.dumps(recursos_data.get("comprobacion_operaciones_logistica", {}), ensure_ascii=False),
            json.dumps(recursos_data, ensure_ascii=False),
        ),
    )

    conn.execute("DELETE FROM recurso_acompanantes WHERE id_viaje = ?", (id_viaje,))
    nombres = recursos_data.get("acompanantes") or []
    viaticos = recursos_data.get("acompanantes_con_viatico") or []
    viaticos_map = {}
    for item in viaticos:
        if isinstance(item, dict):
            nombre = str(item.get("nombre", "")).strip()
            if nombre:
                viaticos_map[nombre] = float(item.get("viatico", 0) or 0)
    for nombre in nombres:
        nombre_txt = str(nombre or "").strip()
        if not nombre_txt:
            continue
        conn.execute(
            "INSERT INTO recurso_acompanantes (id_viaje, nombre, viatico) VALUES (?, ?, ?)",
            (id_viaje, nombre_txt, viaticos_map.get(nombre_txt, 0.0)),
        )


async def guardar_adjunto_logistico(upload: UploadFile | None, nro_orden: str, nombre_base: str):
    if upload is None or not upload.filename:
        return None

    ext = os.path.splitext(upload.filename)[1]
    nombre_archivo = f"{nro_orden}-{nombre_base}{ext}"
    ruta_archivo = os.path.join(DOC_LOG_VIAJES_DIR, nombre_archivo)

    contenido = await upload.read()
    with open(ruta_archivo, "wb") as f:
        f.write(contenido)

    return nombre_archivo


def _calcular_horas_control(control_data: dict):
    def _to_minutes(hhmm: str):
        texto = str(hhmm or "").strip()
        if ":" not in texto:
            return None
        partes = texto.split(":")
        if len(partes) < 2:
            return None
        try:
            horas = int(partes[0])
            minutos = int(partes[1])
        except Exception:
            return None
        if horas < 0 or horas > 23 or minutos < 0 or minutos > 59:
            return None
        return (horas * 60) + minutos

    detalle = control_data.get("detalle_dias", []) if isinstance(control_data, dict) else []
    if not isinstance(detalle, list):
        return 0.0, 0.0, 0.0

    jornada_normal_hs = 8.0
    jornada_compensable_hs = 10.0
    total_horas = 0.0
    total_normales = 0.0
    total_compensables = 0.0
    for fila in detalle:
        if not isinstance(fila, dict):
            continue
        ini = _to_minutes(fila.get("hs_inicio", ""))
        fin = _to_minutes(fila.get("hs_final", ""))
        if ini is None or fin is None:
            continue
        tramo = fin - ini
        if tramo < 0:
            tramo += 24 * 60
        horas_dia = tramo / 60.0
        normales_dia = min(horas_dia, jornada_normal_hs)
        # Las compensables se miden contra 10 hs objetivo: puede dar positivo o negativo (horas a devolver).
        compensables_dia = horas_dia - jornada_compensable_hs
        total_horas += horas_dia
        total_normales += normales_dia
        total_compensables += compensables_dia

    return round(total_horas, 2), round(total_normales, 2), round(total_compensables, 2)


def _legajos_por_nombre(conn):
    mapa = {}
    rows = conn.execute("SELECT legajo, nombre FROM personal").fetchall()
    for row in rows:
        nombre = str(row["nombre"] or "").strip()
        if not nombre:
            continue
        clave = normalizar_texto(nombre)
        if clave and clave not in mapa:
            mapa[clave] = str(row["legajo"] or "").strip()
    return mapa


def _participantes_operativos(recursos_data: dict, conn):
    recursos = recursos_data if isinstance(recursos_data, dict) else {}
    legajos = _legajos_por_nombre(conn)
    chofer = str(recursos.get("chofer") or "").strip()
    acompanantes = recursos.get("acompanantes") or []
    if not isinstance(acompanantes, list):
        acompanantes = []
    acomp_viaticos = recursos.get("acompanantes_con_viatico") or []
    if not isinstance(acomp_viaticos, list):
        acomp_viaticos = []

    map_viaticos = {}
    for item in acomp_viaticos:
        if not isinstance(item, dict):
            continue
        nombre = str(item.get("nombre") or "").strip()
        if not nombre:
            continue
        map_viaticos[normalizar_texto(nombre)] = float(item.get("viatico") or 0)

    out = []
    if chofer:
        out.append(
            {
                "nombre": chofer,
                "rol": "CHOFER",
                "legajo": legajos.get(normalizar_texto(chofer), ""),
                "viatico": float(recursos.get("chofer_viatico") or 0),
                "chofer": chofer,
            }
        )

    for nombre in acompanantes:
        nombre_txt = str(nombre or "").strip()
        if not nombre_txt:
            continue
        clave = normalizar_texto(nombre_txt)
        out.append(
            {
                "nombre": nombre_txt,
                "rol": "ACOMPANANTE",
                "legajo": legajos.get(clave, ""),
                "viatico": float(map_viaticos.get(clave, 0.0)),
                "chofer": chofer,
            }
        )

    return out


def sincronizar_gestion_operativa_asignado(conn, orden: dict, recursos_data: dict):
    if not isinstance(orden, dict):
        return
    nro_orden = str(orden.get("nro_orden") or "").strip()
    if not nro_orden:
        return

    viaje = orden.get("viaje") if isinstance(orden.get("viaje"), dict) else {}
    recursos = recursos_data if isinstance(recursos_data, dict) else {}
    participantes = _participantes_operativos(recursos, conn)

    conn.execute("DELETE FROM gestion_operativa WHERE nro_orden = ?", (nro_orden,))

    for persona in participantes:
        jornadas_raw = str(recursos.get("duracion_jornadas") or "").strip()
        jornadas = int(float(jornadas_raw)) if jornadas_raw else 0
        conn.execute(
            """
            INSERT INTO gestion_operativa (
                nro_orden, fecha_orden, estado_orden, id_viaje,
                centro_costo, proyecto, origen, destino,
                fecha_salida, fecha_regreso, jornadas,
                legajo, nombre, rol,
                vehiculo, chofer,
                viatico, horas_totales, horas_normales, horas_compensables,
                costo_total, fecha_cierre
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, NULL)
            """,
            (
                nro_orden,
                str(orden.get("fecha_orden") or ""),
                "ASIGNADO",
                orden.get("id_viaje"),
                str(recursos.get("centro_costo") or ""),
                str(recursos.get("proyecto") or viaje.get("proyecto") or ""),
                str(recursos.get("partida") or viaje.get("origen") or ""),
                str(recursos.get("destino") or viaje.get("destino") or ""),
                str(recursos.get("fecha_salida_viaje") or viaje.get("fecha_salida") or ""),
                str(recursos.get("fecha_regreso_viaje") or viaje.get("fecha_regreso") or ""),
                jornadas,
                persona.get("legajo", ""),
                persona.get("nombre", ""),
                persona.get("rol", ""),
                str(recursos.get("vehiculo") or viaje.get("vehiculo") or ""),
                persona.get("chofer", ""),
                float(persona.get("viatico") or 0),
            ),
        )


def sincronizar_gestion_operativa_cierre(conn, nro_orden: str, orden: dict, payload_data: dict):
    nro = str(nro_orden or "").strip()
    if not nro:
        return

    orden_data = orden if isinstance(orden, dict) else {}
    recursos = orden_data.get("recursos") if isinstance(orden_data.get("recursos"), dict) else {}
    asignacion = payload_data.get("asignacion", {}) if isinstance(payload_data, dict) else {}
    solicitud = payload_data.get("solicitud", {}) if isinstance(payload_data, dict) else {}
    control = payload_data.get("control_y_seguimiento", {}) if isinstance(payload_data, dict) else {}

    # Reconstituye la nómina final para no arrastrar participantes tentativos del ASIGNADO.
    sincronizar_gestion_operativa_asignado(conn, orden_data, recursos)

    horas_totales, horas_normales, horas_compensables = _calcular_horas_control(control)

    acomp_viaticos = asignacion.get("acompanantes_con_viatico", recursos.get("acompanantes_con_viatico", []))
    if not isinstance(acomp_viaticos, list):
        acomp_viaticos = []
    map_viaticos = {}
    for item in acomp_viaticos:
        if not isinstance(item, dict):
            continue
        nombre = str(item.get("nombre") or "").strip()
        if not nombre:
            continue
        map_viaticos[normalizar_texto(nombre)] = float(item.get("viatico") or 0)

    chofer = str(asignacion.get("chofer") or recursos.get("chofer") or "").strip()
    vehiculo = str(asignacion.get("vehiculo") or recursos.get("vehiculo") or "").strip()
    viatico_chofer = float(asignacion.get("viatico_chofer", recursos.get("chofer_viatico", 0)) or 0)
    otros_gastos = float(asignacion.get("otros_gastos", recursos.get("otros_gastos", 0)) or 0)

    rows = conn.execute(
        "SELECT id, nombre, rol FROM gestion_operativa WHERE nro_orden = ?",
        (nro,),
    ).fetchall()
    personas = len(rows)
    prorrateo_otros = (otros_gastos / personas) if personas > 0 else 0.0
    fecha_cierre = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    jornadas_raw = str(solicitud.get("duracion_jornadas") or recursos.get("duracion_jornadas") or "").strip()
    jornadas = int(float(jornadas_raw)) if jornadas_raw else 0

    for row in rows:
        nombre = str(row["nombre"] or "").strip()
        rol = str(row["rol"] or "").strip().upper()
        if rol == "CHOFER" or (chofer and normalizar_texto(nombre) == normalizar_texto(chofer)):
            viatico = viatico_chofer
        else:
            viatico = float(map_viaticos.get(normalizar_texto(nombre), 0.0))

        costo_total = round(viatico + prorrateo_otros, 2)

        conn.execute(
            """
            UPDATE gestion_operativa
               SET estado_orden = 'CERRADO',
                   fecha_cierre = ?,
                   centro_costo = ?,
                   proyecto = ?,
                   origen = ?,
                   destino = ?,
                   fecha_salida = ?,
                   fecha_regreso = ?,
                   jornadas = ?,
                   vehiculo = ?,
                   chofer = ?,
                   viatico = ?,
                   horas_totales = ?,
                   horas_normales = ?,
                   horas_compensables = ?,
                   costo_total = ?
             WHERE id = ?
            """,
            (
                fecha_cierre,
                str(solicitud.get("centro_costos") or recursos.get("centro_costo") or ""),
                str(solicitud.get("proyecto") or recursos.get("proyecto") or ""),
                str(solicitud.get("partida") or recursos.get("partida") or ""),
                str(solicitud.get("destino") or recursos.get("destino") or ""),
                str(solicitud.get("fecha_salida") or recursos.get("fecha_salida_viaje") or ""),
                str(solicitud.get("fecha_llegada") or recursos.get("fecha_regreso_viaje") or ""),
                jornadas,
                vehiculo,
                chofer,
                round(viatico, 2),
                horas_totales,
                horas_normales,
                horas_compensables,
                costo_total,
                row["id"],
            ),
        )


def reconciliar_gestion_operativa_cerradas(conn):
    rows = conn.execute(
        """
        SELECT nro_orden, raw_json, cierre_logistica_json
        FROM ordenes_salida
        WHERE UPPER(COALESCE(estado, '')) = 'CERRADO'
           OR cierre_logistica_json IS NOT NULL
        """
    ).fetchall()

    for row in rows:
        orden = parse_json_dict(row["raw_json"], default={})
        if not isinstance(orden, dict) or not orden:
            continue
        cierre = parse_json_dict(row["cierre_logistica_json"], default={})
        if not isinstance(cierre, dict):
            continue
        payload_data = cierre.get("datos", {}) if isinstance(cierre.get("datos", {}), dict) else {}
        if not payload_data:
            continue
        sincronizar_gestion_operativa_cierre(conn, row["nro_orden"], orden, payload_data)


init_sqlite()


class Viaje(BaseModel):
    solicitante: str
    area: str
    origen: str
    destino: str
    motivo: str
    fecha_salida: str
    fecha_regreso: str
    chofer: str = ""
    vehiculo: str = ""
    acompanantes: list[str] = []
    alojamiento: str = "NO"

@app.post("/viajes")
def crear_viaje(viaje: Viaje):
    nuevo = viaje.dict()
    with get_sqlite_connection() as conn:
        row = conn.execute("SELECT COALESCE(MAX(id), 0) AS max_id FROM viajes").fetchone()
        nuevo_id = int(row["max_id"] or 0) + 1

    nuevo["id"] = nuevo_id
    nuevo["estado"] = "PENDIENTE"
    nuevo["fecha_creacion"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_sqlite_connection() as conn:
        guardar_viaje_sql(conn, nuevo)
        conn.commit()

    return {"mensaje": "Viaje creado", "id": nuevo["id"]}

@app.get("/viajes")
def listar_viajes():
    return obtener_viajes_data()

class Recursos(BaseModel):
    id_viaje: int
    fecha: str = ""
    centro_costo: str = ""
    datos_solicitante: str = ""
    area_solicitante: str = ""
    partida: str = ""
    destino: str = ""
    motivo_viaje: str = ""
    fecha_salida_viaje: str = ""
    fecha_regreso_viaje: str = ""
    hora_ingreso_base: str = ""
    hora_salida: str = ""
    hora_regreso: str = ""
    duracion_jornadas: str = ""
    itinerario: str = ""
    rutas: str = ""
    paradas: str = ""
    chofer: str
    chofer_viatico: float = 0
    vehiculo: str
    vehiculo_fuera_flota: bool = False
    acompanantes: list[str] = []
    acompanantes_con_viatico: list[dict] = []
    viaticos: float = 0
    medio_pago: str = "Caja"
    alojamiento: str = "NO"
    otros_gastos: float = 0
    verificado_administracion: str = "NO"
    comprobacion_operaciones_logistica: dict = {}


@app.post("/analisis")
def analisis(data: dict):
    viajes = float(data.get("viajes", 0) or 0)
    aprobados = float(data.get("aprobados", 0) or 0)
    ordenes = float(data.get("ordenes", 0) or 0)
    ordenes_cerradas = float(data.get("ordenes_cerradas", 0) or 0)
    flota = float(data.get("flota", 0) or 0)

    def conclusion_local():
        texto = ""
        if flota < 60:
            texto += "Baja disponibilidad de flota. "
        if aprobados < viajes and ordenes_cerradas <= 0:
            texto += "Hay solicitudes pendientes de aprobación. "
        if ordenes_cerradas > 0:
            texto += "Se registran ordenes cerradas, lo que indica viajes ya ejecutados y cierre operativo completado. "
            if aprobados <= 0:
                texto += "Puede haber desfase de carga entre aprobaciones y cierres; conviene auditar la trazabilidad de estados. "
        if ordenes < aprobados:
            texto += "No todos los viajes aprobados tienen orden generada. "
        if flota >= 70:
            texto += "La flota se encuentra en buen estado operativo. "
        if texto == "":
            texto = "Operación estable sin alertas relevantes."
        return texto.strip()

    def _normalizar_conclusion(texto: str, fallback: str):
        limpio = str(texto or "").strip()
        if not limpio:
            return fallback
        # Evita markdown en la UI y mantiene lectura simple.
        limpio = limpio.replace("**", "").replace("__", "")
        if len(limpio) < 20:
            return fallback
        # Si no termina en puntuacion pero ya hay frases completas, corta en la ultima frase cerrada.
        if not limpio.endswith((".", "!", "?")):
            ult = max(limpio.rfind("."), limpio.rfind("!"), limpio.rfind("?"))
            if ult >= 0:
                limpio = limpio[: ult + 1].strip()
        if len(limpio) < 20:
            return fallback
        return limpio

    if client is None:
        print("USANDO FALLBACK LOCAL")
        return {
            "conclusion": conclusion_local()
        }

    prompt = f"""
    Analizar estos datos operativos de logistica:

    - Viajes totales: {viajes}
    - Viajes aprobados: {aprobados}
    - Ordenes generadas: {ordenes}
    - Ordenes cerradas: {ordenes_cerradas}
    - Flota operativa: {flota}%

    Regla clave: si hay ordenes cerradas, considerar que esos viajes ya completaron proceso operativo; no concluir ineficiencia solo por "aprobados = 0".

    Generar una conclusion breve, clara y profesional.
    Incluir posibles riesgos operativos si los hay.
    """

    try:
        print("USANDO IA REAL")
        respuesta = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "Sos un analista logistico senior. Detecta problemas, riesgos y recomenda acciones concretas. Responde en texto plano, sin markdown.",
                },
                {"role": "user", "content": prompt},
            ],
            max_tokens=220,
        )

        choice = respuesta.choices[0]
        finish_reason = str(getattr(choice, "finish_reason", "") or "").lower()
        if finish_reason == "length":
            return {"conclusion": conclusion_local()}

        conclusion = _normalizar_conclusion(
            getattr(choice.message, "content", "") or "",
            conclusion_local(),
        )

        return {"conclusion": conclusion}
    except Exception:
        print("USANDO FALLBACK LOCAL")
        return {
            "conclusion": conclusion_local()
        }


@app.post("/analisis_vehiculos")
def analisis_vehiculos(data: dict):
    total = float(data.get("total", 0) or 0)
    operativos = float(data.get("operativos", 0) or 0)
    no_operativos = float(data.get("no_operativos", 0) or 0)
    flota = float(data.get("flota", 0) or 0)
    proyecto_principal = str(data.get("proyecto_principal", "SIN DATO") or "SIN DATO")
    sector_principal = str(data.get("sector_principal", "SIN DATO") or "SIN DATO")
    pct_habilitaciones = float(data.get("pct_habilitaciones", 0) or 0)

    def conclusion_local_vehiculos():
        texto = ""
        if flota < 60:
            texto += "Baja disponibilidad de flota vehicular. "
        if no_operativos > operativos:
            texto += "Hay mas equipos no operativos que operativos. "
        if pct_habilitaciones < 50:
            texto += "El nivel de habilitaciones es bajo para la demanda operativa. "
        if not texto:
            texto = "Flota vehicular estable, con indicadores operativos controlados."
        return texto.strip()

    def _normalizar_conclusion(texto: str, fallback: str):
        limpio = str(texto or "").strip().replace("**", "").replace("__", "")
        if len(limpio) < 20:
            return fallback
        if not limpio.endswith((".", "!", "?")):
            ult = max(limpio.rfind("."), limpio.rfind("!"), limpio.rfind("?"))
            if ult >= 0:
                limpio = limpio[: ult + 1].strip()
        return limpio if len(limpio) >= 20 else fallback

    if client is None:
        return {"conclusion": conclusion_local_vehiculos()}

    prompt = f"""
    Analizar SOLO estos datos de flota vehicular. No mezclar con viajes, ordenes u otras interfaces.

    - Total de equipos: {total}
    - Equipos operativos: {operativos}
    - Equipos no operativos: {no_operativos}
    - Flota operativa: {flota}%
    - Proyecto con mayor concentracion: {proyecto_principal}
    - Sector con mayor concentracion: {sector_principal}
    - Nivel global de habilitaciones: {pct_habilitaciones}%

    Generar una conclusion breve, profesional y accionable (2 a 4 oraciones),
    enfocada unicamente en estado de vehiculos y habilitaciones.
    """

    try:
        respuesta = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "Sos un analista de flota vehicular y mantenimiento. Responde solo sobre vehiculos y en texto plano, sin markdown.",
                },
                {"role": "user", "content": prompt},
            ],
            max_tokens=220,
        )

        choice = respuesta.choices[0]
        finish_reason = str(getattr(choice, "finish_reason", "") or "").lower()
        if finish_reason == "length":
            return {"conclusion": conclusion_local_vehiculos()}

        conclusion = _normalizar_conclusion(
            getattr(choice.message, "content", "") or "",
            conclusion_local_vehiculos(),
        )
        return {"conclusion": conclusion}
    except Exception:
        return {"conclusion": conclusion_local_vehiculos()}


@app.post("/analisis_personal")
def analisis_personal(data: dict):
    total = float(data.get("total", 0) or 0)
    con_alguna = float(data.get("con_alguna", 0) or 0)
    sin_habilitaciones = float(data.get("sin_habilitaciones", 0) or 0)
    pct_cobertura = float(data.get("pct_cobertura", 0) or 0)
    proyecto_principal = str(data.get("proyecto_principal", "SIN DATO") or "SIN DATO")

    def conclusion_local_personal():
        texto = ""
        if pct_cobertura < 50:
            texto += "Cobertura de habilitaciones baja para la dotacion actual. "
        if sin_habilitaciones > con_alguna:
            texto += "Predomina personal sin habilitaciones registradas. "
        if not texto:
            texto = "La dotacion presenta un nivel de habilitacion aceptable para la operacion."
        return texto.strip()

    def _normalizar_conclusion(texto: str, fallback: str):
        limpio = str(texto or "").strip().replace("**", "").replace("__", "")
        if len(limpio) < 20:
            return fallback
        if not limpio.endswith((".", "!", "?")):
            ult = max(limpio.rfind("."), limpio.rfind("!"), limpio.rfind("?"))
            if ult >= 0:
                limpio = limpio[: ult + 1].strip()
        return limpio if len(limpio) >= 20 else fallback

    if client is None:
        return {"conclusion": conclusion_local_personal()}

    prompt = f"""
    Analizar SOLO estos datos de gestion de personal y habilitaciones. No mezclar con viajes, ordenes ni flota vehicular.

    - Total de personal: {total}
    - Personal con al menos una habilitacion: {con_alguna}
    - Personal sin habilitaciones: {sin_habilitaciones}
    - Cobertura global de habilitaciones: {pct_cobertura}%
    - Proyecto con mayor dotacion habilitada: {proyecto_principal}

    Generar una conclusion breve, profesional y accionable (2 a 4 oraciones),
    enfocada exclusivamente en habilitaciones del personal.
    """

    try:
        respuesta = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "Sos un analista de RRHH operativo senior. Detecta riesgos de habilitaciones y recomienda acciones concretas. Responde en texto plano, sin markdown.",
                },
                {"role": "user", "content": prompt},
            ],
            max_tokens=220,
        )

        choice = respuesta.choices[0]
        finish_reason = str(getattr(choice, "finish_reason", "") or "").lower()
        if finish_reason == "length":
            return {"conclusion": conclusion_local_personal()}

        conclusion = _normalizar_conclusion(
            getattr(choice.message, "content", "") or "",
            conclusion_local_personal(),
        )
        return {"conclusion": conclusion}
    except Exception:
        return {"conclusion": conclusion_local_personal()}

@app.post("/recursos")
def asignar_recursos(data: Recursos):
    with get_sqlite_connection() as conn:
        existe_orden = conn.execute(
            "SELECT 1 FROM ordenes_salida WHERE id_viaje = ?",
            (data.id_viaje,),
        ).fetchone()
        if existe_orden:
            return {"error": f"El viaje ID {data.id_viaje} ya tiene Orden de Salida generada"}

    vehiculo_ingresado = (data.vehiculo or "").strip()
    if not vehiculo_ingresado:
        return {"error": "Debe indicar un vehiculo"}

    vehiculos = obtener_vehiculos_data()
    codigo_vehiculo = vehiculo_ingresado.split(" - ")[0].strip()
    vehiculo_encontrado = next((v for v in vehiculos if str(v.get("codigo", "")).strip() == codigo_vehiculo), None)

    if vehiculo_encontrado is not None and str(vehiculo_encontrado.get("operativo", "NO")).strip().upper() != "SI":
        return {"error": f"El vehiculo {codigo_vehiculo} no esta operativo"}

    campo_habilitacion_vehiculo = obtener_campo_habilitacion(data.destino)
    if vehiculo_encontrado is not None and campo_habilitacion_vehiculo is not None:
        if not esta_habilitado(vehiculo_encontrado, campo_habilitacion_vehiculo):
            return {
                "error": (
                    f"El vehiculo {codigo_vehiculo} no esta habilitado "
                    f"para el destino {data.destino}"
                )
            }

    personal = obtener_personal_data()
    campo_habilitacion = obtener_campo_habilitacion(data.destino)

    if campo_habilitacion is not None:
        clave_proyecto = campo_habilitacion.replace("habilitacion_", "")
        nombres_validar = [data.chofer] + (data.acompanantes or [])
        for nombre in nombres_validar:
            nombre_limpio = (nombre or "").strip()
            if not nombre_limpio:
                continue

            persona = next(
                (p for p in personal if normalizar_texto(p.get("nombre", "")) == normalizar_texto(nombre_limpio)),
                None,
            )
            if persona is None:
                return {"error": f"No se encontro en personal: {nombre_limpio}"}
            if not esta_habilitado_personal(persona, clave_proyecto):
                return {"error": f"{nombre_limpio} no esta habilitado para el destino {data.destino}"}

    with get_sqlite_connection() as conn:
        chofer_ocupado = conn.execute(
            """
            SELECT 1
            FROM recursos_viaje rv
            JOIN viajes v ON v.id = rv.id_viaje
            WHERE rv.chofer = ?
              AND UPPER(COALESCE(v.estado, '')) = 'ASIGNADO'
              AND v.id <> ?
            """,
            (data.chofer, data.id_viaje),
        ).fetchone()
        if chofer_ocupado:
            return {"error": f"El chofer {data.chofer} ya está asignado a otro viaje"}

        viaje_row = conn.execute(
            "SELECT raw_json FROM viajes WHERE id = ?",
            (data.id_viaje,),
        ).fetchone()
        if viaje_row is None:
            return {"error": "Viaje no encontrado"}

        viaje = parse_json_dict(viaje_row["raw_json"], default={})
        if not viaje:
            return {"error": "Viaje no encontrado"}

        recursos_data = data.dict()
        recursos_data["vehiculo"] = vehiculo_ingresado
        recursos_data["vehiculo_fuera_flota"] = vehiculo_encontrado is None

        viaje["recursos"] = recursos_data
        viaje["estado"] = "ASIGNADO"
        viaje["orden_salida_generada"] = True

        nro_orden = generar_numero_orden_sql(conn)
        viaje["nro_orden_salida"] = nro_orden
        fecha_orden = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        guardar_viaje_sql(conn, viaje)
        guardar_recursos_sql(conn, data.id_viaje, recursos_data)

        orden = {
            "nro_orden": nro_orden,
            "fecha_orden": fecha_orden,
            "id_viaje": viaje.get("id"),
            "viaje": dict(viaje),
            "recursos": dict(recursos_data),
            "estado": "ASIGNADO",
        }
        conn.execute(
            """
            INSERT INTO ordenes_salida (nro_orden, fecha_orden, id_viaje, estado, cierre_logistica_json, raw_json)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(nro_orden) DO UPDATE SET
                fecha_orden = excluded.fecha_orden,
                id_viaje = excluded.id_viaje,
                estado = excluded.estado,
                cierre_logistica_json = excluded.cierre_logistica_json,
                raw_json = excluded.raw_json
            """,
            (
                nro_orden,
                fecha_orden,
                viaje.get("id"),
                "ASIGNADO",
                None,
                json.dumps(orden, ensure_ascii=False),
            ),
        )
        sincronizar_gestion_operativa_asignado(conn, orden, recursos_data)
        conn.commit()

    return {"mensaje": "Recursos asignados", "nro_orden": nro_orden}

@app.get("/imagenes/editadas/listado")
def imagenes_editadas_listado():
    editadas_dir = os.path.join(BASE_DIR, "Imagenes", "Editadas")
    extensiones_validas = (".jpg", ".jpeg", ".png", ".webp", ".gif")
    imagenes = []
    if os.path.isdir(editadas_dir):
        for nombre in sorted(os.listdir(editadas_dir)):
            if nombre.lower().endswith(extensiones_validas):
                imagenes.append({"url": f"/Imagenes/Editadas/{nombre}"})
    return {"imagenes": imagenes}


@app.get("/login", response_class=HTMLResponse)
def login_view(request: Request):
    return HTMLResponse(
        _leer_html(LOGIN_PATH),
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.post("/login")
def login_submit(request: Request, username: str = Form(""), password: str = Form("")):
    user = str(username or "").strip()
    pwd = str(password or "")

    if user == AUTH_USER and pwd == AUTH_PASS:
        response = RedirectResponse("/", status_code=302)
        response.set_cookie(AUTH_COOKIE_NAME, user, max_age=SESSION_IDLE_SECONDS, httponly=True, samesite="lax")
        return response

    # Usuarios administrados desde el módulo Administración. El nombre de
    # usuario es el correo y la contraseña se guarda como SHA-256, formato que
    # ya utiliza la tabla usuarios existente.
    try:
        with get_sqlite_connection() as conn:
            fila = conn.execute(
                """
                SELECT id, password_hash FROM usuarios
                WHERE lower(correo) = lower(?)
                  AND upper(COALESCE(estado, 'ACTIVO')) = 'ACTIVO'
                  AND COALESCE(bloqueado, 0) = 0
                """,
                (user,),
            ).fetchone()
            password_hash = hashlib.sha256(pwd.encode("utf-8")).hexdigest()
            if fila and fila["password_hash"] == password_hash:
                conn.execute(
                    """
                    UPDATE usuarios
                    SET intentos_fallidos = 0, ultimo_acceso = datetime('now'),
                        updated_at = datetime('now')
                    WHERE id = ?
                    """,
                    (fila["id"],),
                )
                conn.execute(
                    """
                    INSERT INTO historial_accesos
                    (usuario_id, username_input, evento, detalle, ip, user_agent)
                    VALUES (?, ?, 'LOGIN_EXITOSO', '', ?, ?)
                    """,
                    (fila["id"], user, request.client.host if request.client else "", request.headers.get("user-agent", "")),
                )
                conn.commit()
                response = RedirectResponse("/", status_code=302)
                response.set_cookie(AUTH_COOKIE_NAME, user, max_age=SESSION_IDLE_SECONDS, httponly=True, samesite="lax")
                return response
            if fila:
                conn.execute(
                    "UPDATE usuarios SET intentos_fallidos = COALESCE(intentos_fallidos, 0) + 1, updated_at = datetime('now') WHERE id = ?",
                    (fila["id"],),
                )
                conn.commit()
    except sqlite3.Error:
        # La respuesta no revela si el usuario existe ni detalles internos.
        pass
    return RedirectResponse("/login?error=1", status_code=302)


@app.get("/me")
def get_me(request: Request):
    perfil = _usuario_autenticado(request)
    if perfil is None:
        return {"autenticado": False}
    return {"autenticado": True, **perfil}


@app.get("/session/actividad")
def session_actividad(request: Request):
    if _usuario_autenticado(request) is None:
        raise HTTPException(status_code=401, detail="Sesión no válida")
    return {"ok": True}


def _admin_usuario_a_dict(conn, fila):
    usuario = dict(fila)
    usuario["modulos"] = parse_json_list(usuario.pop("modulos_json", ""))
    usuario["paneles"] = parse_json_dict(usuario.pop("paneles_json", ""), default={})
    usuario["acciones"] = parse_json_dict(usuario.pop("acciones_json", ""), default={})
    roles = conn.execute(
        """
        SELECT r.id, r.nombre FROM roles_funcionales r
        JOIN usuario_roles_funcionales ur ON ur.rol_id = r.id
        WHERE ur.usuario_id = ? ORDER BY r.nombre
        """,
        (usuario["id"],),
    ).fetchall()
    usuario["roles"] = [dict(rol) for rol in roles]
    # Las contraseñas nunca se devuelven ni se almacenan en texto plano.
    usuario["password_visible"] = ""
    return usuario


def _registrar_acceso_admin(conn, request: Request, evento: str, detalle: str = ""):
    perfil = _usuario_autenticado(request) or {}
    conn.execute(
        """
        INSERT INTO historial_accesos (username_input, evento, detalle, ip, user_agent)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            perfil.get("usuario", ""), evento, detalle,
            request.client.host if request.client else "",
            request.headers.get("user-agent", ""),
        ),
    )


@app.get("/administracion", response_class=HTMLResponse)
def administracion_view(request: Request):
    try:
        _requiere_administrador(request)
    except HTTPException:
        return RedirectResponse("/", status_code=302)
    return _leer_html(ADMINISTRACION_PATH)


@app.get("/mantenimiento", response_class=HTMLResponse)
def mantenimiento_view(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(FLEETCARE_PATH)


def _requiere_modulo(request: Request, modulo: str):
    perfil = _usuario_autenticado(request)
    if perfil is None:
        raise HTTPException(status_code=401, detail="Sesión no válida")
    if str(perfil.get("tipo_usuario", "")).upper() == "ADMINISTRADOR" or modulo in perfil.get("modulos", []):
        return perfil
    raise HTTPException(status_code=403, detail="No tiene acceso a este módulo")


@app.get("/base_datos", response_class=HTMLResponse)
def base_datos_view(request: Request):
    try:
        _requiere_modulo(request, "base_datos")
    except HTTPException:
        return RedirectResponse("/", status_code=302)
    return _leer_html(BASE_DATOS_PATH)


def _tablas_base_datos(conn):
    return [
        fila["name"] for fila in conn.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        ).fetchall()
    ]


@app.get("/base_datos/tablas")
def base_datos_tablas(request: Request):
    _requiere_modulo(request, "base_datos")
    with get_sqlite_connection() as conn:
        tablas = _tablas_base_datos(conn)
        resultado = []
        for tabla in tablas:
            tabla_sql = tabla.replace('"', '""')
            cantidad = conn.execute(f'SELECT COUNT(*) FROM "{tabla_sql}"').fetchone()[0]
            resultado.append({"nombre": tabla, "registros": cantidad})
        return resultado


@app.get("/base_datos/tabla/{tabla}")
def base_datos_tabla(tabla: str, request: Request):
    _requiere_modulo(request, "base_datos")
    with get_sqlite_connection() as conn:
        if tabla not in _tablas_base_datos(conn):
            raise HTTPException(status_code=404, detail="Tabla no encontrada")
        tabla_sql = tabla.replace('"', '""')
        columnas = [fila["name"] for fila in conn.execute(f'PRAGMA table_info("{tabla_sql}")').fetchall()]
        filas = []
        for fila in conn.execute(f'SELECT * FROM "{tabla_sql}"').fetchall():
            registro = dict(fila)
            # Las credenciales y material de autenticación nunca se exponen en una grilla.
            for campo in ("password_hash", "password_cifrada"):
                if campo in registro:
                    registro[campo] = "••••••••"
            filas.append(registro)
        return {"tabla": tabla, "columnas": columnas, "filas": filas, "total": len(filas)}


@app.get("/admin/roles_funcionales")
def admin_listar_roles(request: Request):
    _requiere_administrador(request)
    with get_sqlite_connection() as conn:
        return [dict(fila) for fila in conn.execute(
            "SELECT id, nombre, activo FROM roles_funcionales WHERE COALESCE(activo, 1) = 1 ORDER BY nombre"
        ).fetchall()]


@app.get("/admin/usuarios")
def admin_listar_usuarios(request: Request):
    _requiere_administrador(request)
    with get_sqlite_connection() as conn:
        filas = conn.execute(
            """
            SELECT id, nombre_apellido, dni, legajo, correo, estado, tipo_usuario,
                   modulos_json, paneles_json, acciones_json, bloqueado,
                   intentos_fallidos, password_temporal, ultimo_acceso, created_at
            FROM usuarios ORDER BY nombre_apellido, correo
            """
        ).fetchall()
        return [_admin_usuario_a_dict(conn, fila) for fila in filas]


@app.post("/admin/usuarios")
def admin_crear_usuario_endpoint(request: Request, payload: dict):
    _requiere_administrador(request)
    datos = dict(payload or {})
    password_temporal = ""
    if not str(datos.get("password", "")):
        password_temporal = secrets.token_urlsafe(9)
        datos["password"] = password_temporal
    resultado = admin_crear_usuario(datos)
    if not resultado.get("ok"):
        raise HTTPException(status_code=400, detail=resultado.get("error", "No se pudo crear el usuario"))
    with get_sqlite_connection() as conn:
        _registrar_acceso_admin(conn, request, "USUARIO_CREADO", f"usuario_id={resultado['id']}")
        conn.commit()
    return {**resultado, "password_temporal": password_temporal}


@app.put("/admin/usuarios/{usuario_id}")
def admin_actualizar_usuario(usuario_id: int, request: Request, payload: dict):
    _requiere_administrador(request)
    datos = dict(payload or {})
    nombre = str(datos.get("nombre_apellido", "")).strip()
    correo = str(datos.get("correo", "")).strip()
    if not nombre or not correo:
        raise HTTPException(status_code=400, detail="nombre_apellido y correo son obligatorios")
    modulos = datos.get("modulos", []) if isinstance(datos.get("modulos", []), list) else []
    paneles = datos.get("paneles", {}) if isinstance(datos.get("paneles", {}), dict) else {}
    acciones = datos.get("acciones", {}) if isinstance(datos.get("acciones", {}), dict) else {}
    roles = datos.get("roles", []) if isinstance(datos.get("roles", []), list) else []
    try:
        with get_sqlite_connection() as conn:
            existente = conn.execute("SELECT id FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
            if not existente:
                raise HTTPException(status_code=404, detail="Usuario no encontrado")
            campos = [
                "nombre_apellido = ?", "dni = ?", "legajo = ?", "correo = ?", "estado = ?",
                "tipo_usuario = ?", "modulos_json = ?", "paneles_json = ?", "acciones_json = ?",
                "updated_at = datetime('now')",
            ]
            valores = [
                nombre, str(datos.get("dni", "")).strip(), str(datos.get("legajo", "")).strip(), correo,
                str(datos.get("estado", "ACTIVO")).strip().upper() or "ACTIVO",
                str(datos.get("tipo_usuario", "CONSULTOR")).strip().upper() or "CONSULTOR",
                json.dumps(modulos, ensure_ascii=False), json.dumps(paneles, ensure_ascii=False), json.dumps(acciones, ensure_ascii=False),
            ]
            if str(datos.get("password", "")):
                nueva_password = str(datos["password"])
                campos.extend(["password_hash = ?", "password_cifrada = ?", "password_temporal = 0"])
                valores.append(hashlib.sha256(nueva_password.encode("utf-8")).hexdigest())
                valores.append(_cifrar_password(nueva_password))
            valores.append(usuario_id)
            conn.execute(f"UPDATE usuarios SET {', '.join(campos)} WHERE id = ?", valores)
            conn.execute("DELETE FROM usuario_roles_funcionales WHERE usuario_id = ?", (usuario_id,))
            for rol_id in roles:
                try:
                    conn.execute("INSERT OR IGNORE INTO usuario_roles_funcionales (usuario_id, rol_id) VALUES (?, ?)", (usuario_id, int(rol_id)))
                except (TypeError, ValueError):
                    continue
            _registrar_acceso_admin(conn, request, "USUARIO_ACTUALIZADO", f"usuario_id={usuario_id}")
            conn.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="El correo ya está asignado a otro usuario")
    return {"ok": True, "id": usuario_id}


@app.post("/admin/usuarios/{usuario_id}/estado")
def admin_cambiar_estado_usuario(usuario_id: int, request: Request, payload: dict):
    _requiere_administrador(request)
    estado = str((payload or {}).get("estado", "")).strip().upper()
    if estado not in {"ACTIVO", "INACTIVO"}:
        raise HTTPException(status_code=400, detail="Estado inválido")
    with get_sqlite_connection() as conn:
        cur = conn.execute("UPDATE usuarios SET estado = ?, updated_at = datetime('now') WHERE id = ?", (estado, usuario_id))
        if not cur.rowcount:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        _registrar_acceso_admin(conn, request, "USUARIO_ESTADO", f"usuario_id={usuario_id}; estado={estado}")
        conn.commit()
    return {"ok": True}


@app.post("/admin/usuarios/{usuario_id}/bloqueo")
def admin_cambiar_bloqueo_usuario(usuario_id: int, request: Request, payload: dict):
    _requiere_administrador(request)
    bloqueado = 1 if bool((payload or {}).get("bloqueado")) else 0
    with get_sqlite_connection() as conn:
        cur = conn.execute("UPDATE usuarios SET bloqueado = ?, updated_at = datetime('now') WHERE id = ?", (bloqueado, usuario_id))
        if not cur.rowcount:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        _registrar_acceso_admin(conn, request, "USUARIO_BLOQUEO", f"usuario_id={usuario_id}; bloqueado={bloqueado}")
        conn.commit()
    return {"ok": True}


@app.post("/admin/usuarios/{usuario_id}/reset_password")
def admin_resetear_password(usuario_id: int, request: Request, payload: dict):
    _requiere_administrador(request)
    password = str((payload or {}).get("password", "")) or secrets.token_urlsafe(9)
    with get_sqlite_connection() as conn:
        cur = conn.execute(
            """UPDATE usuarios SET password_hash = ?, password_cifrada = ?, password_temporal = 1,
               intentos_fallidos = 0, updated_at = datetime('now') WHERE id = ?""",
            (hashlib.sha256(password.encode("utf-8")).hexdigest(), _cifrar_password(password), usuario_id),
        )
        if not cur.rowcount:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        _registrar_acceso_admin(conn, request, "PASSWORD_RESTABLECIDA", f"usuario_id={usuario_id}")
        conn.commit()
    return {"ok": True, "password": password}


@app.get("/admin/usuarios/{usuario_id}/password")
def admin_ver_password(usuario_id: int, request: Request):
    _requiere_administrador(request)
    with get_sqlite_connection() as conn:
        fila = conn.execute("SELECT password_cifrada FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
        if not fila:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        password = _descifrar_password(fila["password_cifrada"] or "")
        _registrar_acceso_admin(conn, request, "PASSWORD_VISUALIZADA", f"usuario_id={usuario_id}")
        conn.commit()
    if password is None:
        return {
            "password": None,
            "mensaje": "No disponible: la contraseña se fijó antes de habilitar esta función, "
                       "o falta configurar PASSWORD_ENCRYPTION_KEY. Usá 'Restablecer clave' para fijar una nueva.",
        }
    return {"password": password}


@app.get("/admin/accesos")
def admin_listar_accesos(request: Request, limit: int = Query(100, ge=1, le=500)):
    _requiere_administrador(request)
    with get_sqlite_connection() as conn:
        filas = conn.execute(
            """
            SELECT h.created_at, h.evento, h.username_input, h.detalle, h.ip, u.nombre_apellido
            FROM historial_accesos h LEFT JOIN usuarios u ON u.id = h.usuario_id
            ORDER BY h.id DESC LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(fila) for fila in filas]


@app.get("/logout")
def logout(request: Request):
    response = RedirectResponse("/login", status_code=302)
    response.delete_cookie(AUTH_COOKIE_NAME)
    return response


@app.get("/", response_class=HTMLResponse)
def menu_principal(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(MENU_PRINCIPAL_PATH)


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(DASHBOARD_PATH)


@app.get("/compras", response_class=HTMLResponse)
def compras_view(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(COMPRAS_PATH)


@app.get("/compras/metadata")
def compras_metadata(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    with get_sqlite_connection() as conn:
        _ensure_prioridades_y_estados_compras(conn)
        ultimo_registro = conn.execute(
            "SELECT numero_solicitud FROM solicitud_compra ORDER BY id DESC LIMIT 1"
        ).fetchone()
        if ultimo_registro and ultimo_registro["numero_solicitud"]:
            ultimo_numero_texto = ultimo_registro["numero_solicitud"]
            ultimo_numero_digitos = ''.join(ch for ch in ultimo_numero_texto if ch.isdigit())
            try:
                ultimo_id = int(ultimo_numero_digitos)
            except ValueError:
                ultimo_id = 0
        else:
            ultimo_id = 0

        personal = [
            dict(r)
            for r in conn.execute(
                "SELECT legajo, nombre FROM personal WHERE UPPER(COALESCE(activo, 'ACTIVO')) = 'ACTIVO' ORDER BY nombre"
            ).fetchall()
        ]

        proyectos = [
            dict(r)
            for r in conn.execute(
                "SELECT id, nombre FROM proyectos WHERE COALESCE(activo, 1) = 1 ORDER BY nombre"
            ).fetchall()
        ]

        centros_costos = [
            dict(r)
            for r in conn.execute(
                "SELECT id, nombre FROM centros_costos WHERE COALESCE(activo, 1) = 1 ORDER BY nombre"
            ).fetchall()
        ]

        vehiculos = [
            {
                "codigo": r["codigo"],
                "nombre_equipo": ((r["marca"] or "").strip() + " " + (r["modelo"] or "").strip()).strip(),
            }
            for r in conn.execute(
                "SELECT codigo, marca, modelo FROM vehiculos ORDER BY codigo"
            ).fetchall()
        ]

        productos = [
            {
                "codigo": r["codigo"],
                "descripcion": r["descripcion"],
            }
            for r in conn.execute(
                "SELECT codigo, descripcion FROM productos ORDER BY codigo"
            ).fetchall()
        ]

        unidades_medida = [
            dict(r)
            for r in conn.execute(
                "SELECT id, nombre FROM unidades_medida WHERE COALESCE(activo, 1) = 1 ORDER BY nombre"
            ).fetchall()
        ]

        prioridades = [
            dict(r)
            for r in conn.execute(
                "SELECT id, nombre FROM prioridades_compra WHERE COALESCE(activo, 1) = 1 ORDER BY CASE nombre WHEN 'Baja' THEN 1 WHEN 'Media' THEN 2 WHEN 'Alta' THEN 3 WHEN 'Urgente' THEN 4 ELSE 5 END"
            ).fetchall()
        ]

        estados = [
            dict(r)
            for r in conn.execute(
                "SELECT id, nombre FROM estados_compra WHERE COALESCE(activo, 1) = 1 ORDER BY id"
            ).fetchall()
        ]

        if not prioridades:
            default_prioridades = [
                ("Baja", 1),
                ("Media", 1),
                ("Alta", 1),
                ("Urgente", 1),
            ]
            conn.executemany(
                "INSERT OR IGNORE INTO prioridades_compra (nombre, activo) VALUES (?, ?)",
                default_prioridades,
            )
            conn.commit()
            prioridades = [
                dict(r)
                for r in conn.execute(
                    "SELECT id, nombre FROM prioridades_compra WHERE COALESCE(activo, 1) = 1 ORDER BY id"
                ).fetchall()
            ]

        if not estados:
            default_estados = [
                ("Borrador", 1),
                ("En revisión", 1),
                ("Aprobada", 1),
                ("Rechazada", 1),
                ("Comprada", 1),
            ]
            conn.executemany(
                "INSERT OR IGNORE INTO estados_compra (nombre, activo) VALUES (?, ?)",
                default_estados,
            )
            conn.commit()
            estados = [
                dict(r)
                for r in conn.execute(
                    "SELECT id, nombre FROM estados_compra WHERE COALESCE(activo, 1) = 1 ORDER BY id"
                ).fetchall()
            ]

    numero_solicitud = f"SC-{ultimo_id + 1:05d}"
    return {
        "numero_solicitud": numero_solicitud,
        "personal": personal,
        "proyectos": proyectos,
        "centros_costos": centros_costos,
        "vehiculos": vehiculos,
        "productos": productos,
        "unidades_medida": unidades_medida,
        "prioridades": prioridades,
        "estados": estados,
    }


@app.get("/compras/solicitudes")
def compras_listar_solicitudes(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    with get_sqlite_connection() as conn:
        filas = conn.execute(
            """
            SELECT sc.id, sc.numero_solicitud, sc.fecha_solicitud,
                   COALESCE(p.nombre, '') AS solicitante,
                   COALESCE(pr.nombre, '') AS destino_compra,
                   COALESCE(pc.nombre, '') AS prioridad,
                   COALESCE(ec.nombre, '') AS estado
            FROM solicitud_compra sc
            LEFT JOIN personal p ON p.legajo = sc.id_solicitante
            LEFT JOIN proyectos pr ON pr.id = sc.id_proyecto
            LEFT JOIN prioridades_compra pc ON pc.id = sc.id_prioridad
            LEFT JOIN estados_compra ec ON ec.id = sc.id_estado
            ORDER BY sc.id DESC
            """
        ).fetchall()
    return [dict(fila) for fila in filas]


@app.get("/compras/dashboard")
def compras_dashboard(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    with get_sqlite_connection() as conn:
        _ensure_orden_compra_solicitudes(conn)
        solicitudes = conn.execute("SELECT COUNT(*) AS total FROM solicitud_compra").fetchone()["total"]
        ordenes = conn.execute("SELECT COUNT(*) AS total FROM OrdenCompra").fetchone()["total"]
        ordenes_anuladas = conn.execute("SELECT COUNT(*) AS total FROM OrdenCompra WHERE UPPER(COALESCE(estado, '')) = 'ANULADA'").fetchone()["total"]
        proveedores = conn.execute("SELECT COUNT(*) AS total FROM proveedores").fetchone()["total"]
        items_pendientes = conn.execute(
            "SELECT COUNT(*) AS total FROM solicitud_compra_detalle WHERE UPPER(COALESCE(estado_aprobacion, 'PENDIENTE')) IN ('PENDIENTE', 'APROBADA')"
        ).fetchone()["total"]
        total_neto = conn.execute(
            "SELECT COALESCE(SUM(ocd.subtotal), 0) AS total FROM OrdenCompra oc INNER JOIN OrdenCompraDetalle ocd ON ocd.orden_compra_id = oc.id WHERE UPPER(COALESCE(oc.estado, '')) <> 'ANULADA'"
        ).fetchone()["total"]
        solicitudes_aprobadas = conn.execute(
            "SELECT COUNT(*) AS total FROM solicitud_compra_detalle WHERE UPPER(COALESCE(estado_aprobacion, '')) = 'APROBADA'"
        ).fetchone()["total"]
        estados_solicitudes = conn.execute(
            '''
            SELECT COALESCE(ec.nombre, 'Sin estado') AS estado, COUNT(*) AS total
            FROM solicitud_compra sc
            LEFT JOIN estados_compra ec ON ec.id = sc.id_estado
            GROUP BY COALESCE(ec.nombre, 'Sin estado')
            ORDER BY total DESC, estado
            '''
        ).fetchall()
        estados_items = conn.execute(
            '''
            SELECT CASE
                WHEN vinculo.id IS NOT NULL THEN 'Gestionado'
                WHEN UPPER(COALESCE(d.estado_aprobacion, 'PENDIENTE')) = 'APROBADA' THEN 'Pendiente de gestionar'
                WHEN UPPER(COALESCE(d.estado_aprobacion, 'PENDIENTE')) = 'RECHAZADA' THEN 'Rechazado'
                ELSE 'Pendiente'
            END AS estado, COUNT(*) AS total
            FROM solicitud_compra_detalle d
            LEFT JOIN OrdenCompraDetalleSolicitud vinculo
                   ON vinculo.solicitud_compra_detalle_id = d.id
            GROUP BY estado
            ORDER BY total DESC, estado
            '''
        ).fetchall()
        solicitudes_plazo = conn.execute(
            '''
                 SELECT sc.id, sc.numero_solicitud, sc.fecha_solicitud,
                     COALESCE(pc.nombre, 'Sin prioridad') AS prioridad,
                     d.id AS detalle_id, d.descripcion_articulo, d.estado_aprobacion,
                     vinculo.id AS vinculo_id
            FROM solicitud_compra sc
            LEFT JOIN prioridades_compra pc ON pc.id = sc.id_prioridad
            LEFT JOIN solicitud_compra_detalle d ON d.id_solicitud = sc.id
            LEFT JOIN OrdenCompraDetalleSolicitud vinculo
                   ON vinculo.solicitud_compra_detalle_id = d.id
            ORDER BY sc.id
            '''
        ).fetchall()

    prioridad_config = {
        "BAJA": ("Baja", 30),
        "MEDIA": ("Media", 21),
        "ALTA": ("Alta", 14),
        "URGENTE": ("Urgente", 7),
        "URGENTES": ("Urgente", 7),
    }
    prioridad_resumen = {}
    solicitudes_agrupadas = {}
    for fila in solicitudes_plazo:
        solicitud_id = fila["id"]
        nombre_original = str(fila["prioridad"] or "Sin prioridad").strip()
        clave = nombre_original.upper()
        nombre, dias = prioridad_config.get(clave, (nombre_original, None))
        solicitud = solicitudes_agrupadas.setdefault(
            solicitud_id,
            {"prioridad": nombre, "dias": dias, "fecha": fila["fecha_solicitud"], "numero": fila["numero_solicitud"], "pendiente": False, "productos": []},
        )
        if fila["detalle_id"] is not None and fila["vinculo_id"] is None and str(fila["estado_aprobacion"] or "Pendiente").upper() != "RECHAZADA":
            solicitud["pendiente"] = True
            descripcion = str(fila["descripcion_articulo"] or "").strip()
            if descripcion and descripcion not in solicitud["productos"]:
                solicitud["productos"].append(descripcion)

    hoy = datetime.now().date()
    for solicitud in solicitudes_agrupadas.values():
        clave = solicitud["prioridad"]
        resumen = prioridad_resumen.setdefault(clave, {"prioridad": clave, "dias": solicitud["dias"], "total": 0, "fuera_plazo": 0})
        resumen["total"] += 1
        if solicitud["pendiente"] and solicitud["dias"] is not None:
            try:
                fecha_solicitud = datetime.strptime(str(solicitud["fecha"] or "")[:10], "%Y-%m-%d").date()
                if hoy > fecha_solicitud + timedelta(days=solicitud["dias"]):
                    resumen["fuera_plazo"] += 1
                    resumen.setdefault("solicitudes", []).append({
                        "numero_solicitud": solicitud["numero"],
                        "productos": solicitud["productos"],
                    })
            except ValueError:
                pass
    prioridades_ordenadas = ["Urgente", "Alta", "Media", "Baja"]
    plazos_prioridad = {"Urgente": 7, "Alta": 14, "Media": 21, "Baja": 30}
    for prioridad in prioridades_ordenadas:
        prioridad_resumen.setdefault(
            prioridad,
            {"prioridad": prioridad, "dias": plazos_prioridad[prioridad], "total": 0, "fuera_plazo": 0, "solicitudes": []},
        )
    prioridades_dashboard = [prioridad_resumen[key] for key in prioridades_ordenadas]
    solicitudes_fuera_plazo = sum(item["fuera_plazo"] for item in prioridades_dashboard)

    return {
        "solicitudes": solicitudes,
        "ordenes": ordenes,
        "ordenes_anuladas": ordenes_anuladas,
        "proveedores": proveedores,
        "items_pendientes": items_pendientes,
        "solicitudes_aprobadas": solicitudes_aprobadas,
        "total_neto": total_neto,
        "estados_solicitudes": [dict(item) for item in estados_solicitudes],
        "estados_items": [dict(item) for item in estados_items],
        "prioridades": prioridades_dashboard,
        "solicitudes_fuera_plazo": solicitudes_fuera_plazo,
    }


@app.get("/compras/proveedores")
def compras_listar_proveedores(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    with get_sqlite_connection() as conn:
        filas = conn.execute(
            '''
            SELECT id, "NOM_PROVEE" AS nombre, "DOMICILIO" AS domicilio,
                   "TELÉFONO_1" AS telefono_1, "TELÉFONO_2" AS telefono_2,
                   "DESC_COND" AS condicion_pago, "COND_IVA" AS condicion_iva,
                   "CUIT" AS cuit, "C_POSTAL" AS codigo_postal, "LOCALIDAD" AS localidad
            FROM proveedores
            ORDER BY "NOM_PROVEE"
            '''
        ).fetchall()
    return [dict(fila) for fila in filas]


def _ensure_orden_compra_historial(conn):
    conn.execute(
        '''
        CREATE TABLE IF NOT EXISTS OrdenCompraHistorial (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            orden_compra_id INTEGER NOT NULL,
            accion TEXT NOT NULL,
            detalle TEXT NOT NULL,
            usuario TEXT NOT NULL,
            usuario_nombre TEXT NOT NULL,
            fecha TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (orden_compra_id) REFERENCES OrdenCompra (id) ON DELETE CASCADE
        )
        '''
    )


def _registrar_orden_compra_historial(conn, orden_compra_id, accion, detalle, perfil):
    usuario = str(perfil.get("usuario") or "").strip()
    nombre_usuario = str(perfil.get("nombre_apellido") or "").strip() or usuario
    conn.execute(
        '''
        INSERT INTO OrdenCompraHistorial (orden_compra_id, accion, detalle, usuario, usuario_nombre)
        VALUES (?, ?, ?, ?, ?)
        ''',
        (
            orden_compra_id,
            accion,
            detalle,
            usuario,
            nombre_usuario,
        ),
    )


@app.get("/compras/ordenes-compra")
def compras_listar_ordenes_compra(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    with get_sqlite_connection() as conn:
        _ensure_orden_compra_historial(conn)
        filas = conn.execute(
            '''
            SELECT oc.id, oc.numero_oc, oc.fecha, oc.estado,
                   COALESCE(p."NOM_PROVEE", '') AS proveedor,
                   COALESCE(SUM(ocd.subtotal), 0) AS total_neto,
                   COUNT(ocd.id) AS cantidad_items
            FROM OrdenCompra oc
            LEFT JOIN proveedores p ON p.id = oc.proveedor_id
            LEFT JOIN OrdenCompraDetalle ocd ON ocd.orden_compra_id = oc.id
            GROUP BY oc.id, oc.numero_oc, oc.fecha, oc.estado, p."NOM_PROVEE"
            ORDER BY oc.id DESC
            ''',
        ).fetchall()
    return [dict(fila) for fila in filas]


@app.get("/compras/ordenes-compra/{orden_compra_id}")
def compras_obtener_orden_compra(orden_compra_id: int, request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    with get_sqlite_connection() as conn:
        _ensure_orden_compra_historial(conn)
        cabecera = conn.execute(
            '''
            SELECT oc.id, oc.numero_oc, oc.fecha, oc.proveedor_id, oc.cotizacion, oc.moneda,
                   oc.tipo_cambio, oc.forma_pago, oc.lugar_entrega, oc.transporte,
                   oc.validez_oferta, oc.observaciones, oc.solicitante, oc.aprobador,
                   oc.estado, oc.fecha_creacion,
                   COALESCE(p."NOM_PROVEE", '') AS proveedor,
                   COALESCE(p."DOMICILIO", '') AS domicilio,
                   COALESCE(p."TELÉFONO_1", '') AS telefono_1,
                   COALESCE(p."TELÉFONO_2", '') AS telefono_2,
                   COALESCE(p."COND_IVA", '') AS condicion_iva,
                   COALESCE(p."CUIT", '') AS cuit,
                   COALESCE(p."C_POSTAL", '') AS codigo_postal,
                   COALESCE(p."LOCALIDAD", '') AS localidad
            FROM OrdenCompra oc
            LEFT JOIN proveedores p ON p.id = oc.proveedor_id
            WHERE oc.id = ?
            ''',
            (orden_compra_id,),
        ).fetchone()
        if cabecera is None:
            raise HTTPException(status_code=404, detail="Orden de compra no encontrada.")

        detalles = conn.execute(
            '''
                        SELECT ocd.id, ocd.codigo_proveedor, ocd.descripcion, ocd.cantidad,
                                     ocd.centro_costo_id,
                   ocd.precio_unitario, ocd.subtotal,
                   COALESCE(cc.nombre, '') AS centro_costo,
                   COALESCE(GROUP_CONCAT(sc.numero_solicitud, ', '), '') AS solicitudes
            FROM OrdenCompraDetalle ocd
            LEFT JOIN centros_costos cc ON cc.id = ocd.centro_costo_id
            LEFT JOIN OrdenCompraDetalleSolicitud ocds ON ocds.orden_compra_detalle_id = ocd.id
            LEFT JOIN solicitud_compra_detalle scd ON scd.id = ocds.solicitud_compra_detalle_id
            LEFT JOIN solicitud_compra sc ON sc.id = scd.id_solicitud
            WHERE ocd.orden_compra_id = ?
            GROUP BY ocd.id, ocd.codigo_proveedor, ocd.descripcion, ocd.cantidad,
                     ocd.centro_costo_id, ocd.precio_unitario, ocd.subtotal, cc.nombre
            ORDER BY ocd.id
            ''',
            (orden_compra_id,),
        ).fetchall()
        historial = conn.execute(
            '''
            SELECT id, accion, detalle, usuario, usuario_nombre, fecha
            FROM OrdenCompraHistorial
            WHERE orden_compra_id = ?
            ORDER BY id DESC
            ''',
            (orden_compra_id,),
        ).fetchall()

    orden = dict(cabecera)
    orden["detalles"] = [dict(detalle) for detalle in detalles]
    orden["historial"] = [dict(item) for item in historial]
    return orden


@app.get("/compras/ordenes-compra/{orden_compra_id}/pdf")
def compras_descargar_orden_compra_pdf(orden_compra_id: int, request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    orden = compras_obtener_orden_compra(orden_compra_id, request)
    pdf_path = os.path.join(DOC_ALMACEN_DIR, f"{orden['numero_oc']}.pdf")
    proveedor = {
        "nombre": orden["proveedor"], "cuit": orden["cuit"], "domicilio": orden["domicilio"],
        "condicion_iva": orden["condicion_iva"], "codigo_postal": orden["codigo_postal"],
        "localidad": orden["localidad"], "telefono_1": orden["telefono_1"],
    }
    generar_pdf_orden_compra(pdf_path, orden, proveedor, orden["detalles"], [])
    return FileResponse(pdf_path, media_type="application/pdf", filename=os.path.basename(pdf_path))


@app.post("/compras/ordenes-compra/{orden_compra_id}/anular")
async def compras_anular_orden_compra(orden_compra_id: int, request: Request):
    perfil = _usuario_autenticado(request)
    if perfil is None:
        return RedirectResponse("/login", status_code=302)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    motivo = str(payload.get("motivo") or "").strip()
    if not motivo:
        raise HTTPException(status_code=400, detail="Debe indicar el motivo de la anulación.")

    with get_sqlite_connection() as conn:
        _ensure_orden_compra_historial(conn)
        orden = conn.execute("SELECT estado FROM OrdenCompra WHERE id = ?", (orden_compra_id,)).fetchone()
        if orden is None:
            raise HTTPException(status_code=404, detail="Orden de compra no encontrada.")
        if str(orden["estado"] or "").strip().upper() == "ANULADA":
            raise HTTPException(status_code=400, detail="La orden de compra ya está anulada.")
        conn.execute("UPDATE OrdenCompra SET estado = 'Anulada' WHERE id = ?", (orden_compra_id,))
        _registrar_orden_compra_historial(conn, orden_compra_id, "Anulada", motivo, perfil)
        conn.commit()
    return {"mensaje": "Orden de compra anulada"}


@app.put("/compras/ordenes-compra/{orden_compra_id}/editar")
async def compras_editar_orden_compra(orden_compra_id: int, request: Request):
    perfil = _usuario_autenticado(request)
    if perfil is None:
        return RedirectResponse("/login", status_code=302)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    motivo = str(payload.get("motivo") or "").strip()
    if not motivo:
        raise HTTPException(status_code=400, detail="Debe indicar el motivo de la edición.")
    detalles = payload.get("detalles")
    if not isinstance(detalles, list) or not detalles:
        raise HTTPException(status_code=400, detail="La orden debe conservar al menos un artículo.")

    with get_sqlite_connection() as conn:
        _ensure_orden_compra_historial(conn)
        orden = conn.execute("SELECT * FROM OrdenCompra WHERE id = ?", (orden_compra_id,)).fetchone()
        if orden is None:
            raise HTTPException(status_code=404, detail="Orden de compra no encontrada.")
        if str(orden["estado"] or "").strip().upper() == "ANULADA":
            raise HTTPException(status_code=400, detail="No se puede editar una orden anulada.")
        proveedor_id = payload.get("proveedor_id") or orden["proveedor_id"]
        if conn.execute("SELECT id FROM proveedores WHERE id = ?", (proveedor_id,)).fetchone() is None:
            raise HTTPException(status_code=400, detail="El proveedor seleccionado no existe.")

        originales = conn.execute(
            "SELECT id, codigo_proveedor, descripcion, cantidad, precio_unitario, centro_costo_id FROM OrdenCompraDetalle WHERE orden_compra_id = ? ORDER BY id",
            (orden_compra_id,),
        ).fetchall()
        if len(originales) != len(detalles) or {int(item.get("id") or 0) for item in detalles} != {int(item["id"]) for item in originales}:
            raise HTTPException(status_code=400, detail="No se pueden agregar ni quitar artículos al editar la orden.")

        anterior = {
            "proveedor_id": orden["proveedor_id"], "cotizacion": orden["cotizacion"], "moneda": orden["moneda"],
            "tipo_cambio": orden["tipo_cambio"], "forma_pago": orden["forma_pago"], "lugar_entrega": orden["lugar_entrega"],
            "transporte": orden["transporte"], "validez_oferta": orden["validez_oferta"], "observaciones": orden["observaciones"],
            "solicitante": orden["solicitante"], "aprobador": orden["aprobador"],
            "detalles": [dict(item) for item in originales],
        }
        conn.execute(
            '''
            UPDATE OrdenCompra SET proveedor_id = ?, cotizacion = ?, moneda = ?, tipo_cambio = ?,
                forma_pago = ?, lugar_entrega = ?, transporte = ?, validez_oferta = ?, observaciones = ?,
                solicitante = ?, aprobador = ?
            WHERE id = ?
            ''',
            (
                proveedor_id, str(payload.get("cotizacion") or ""), str(payload.get("moneda") or ""),
                payload.get("tipo_cambio") or None, str(payload.get("forma_pago") or ""),
                str(payload.get("lugar_entrega") or ""), str(payload.get("transporte") or ""),
                str(payload.get("validez_oferta") or ""), str(payload.get("observaciones") or ""),
                str(payload.get("solicitante") or ""), str(payload.get("aprobador") or ""), orden_compra_id,
            ),
        )
        for item in detalles:
            try:
                cantidad = float(item.get("cantidad") or 0)
                precio = float(item.get("precio_unitario") or 0)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Cantidad y precio unitario deben ser numéricos.")
            conn.execute(
                '''
                UPDATE OrdenCompraDetalle
                SET codigo_proveedor = ?, descripcion = ?, cantidad = ?, precio_unitario = ?, subtotal = ?, centro_costo_id = ?
                WHERE id = ? AND orden_compra_id = ?
                ''',
                (
                    str(item.get("codigo_proveedor") or ""), str(item.get("descripcion") or ""), cantidad,
                    precio, cantidad * precio, item.get("centro_costo_id") or None, item["id"], orden_compra_id,
                ),
            )
        actualizado = dict(payload)
        actualizado.pop("motivo", None)
        _registrar_orden_compra_historial(
            conn, orden_compra_id, "Editada",
            json.dumps({"motivo": motivo, "anterior": anterior, "actualizado": actualizado}, ensure_ascii=False, default=str),
            perfil,
        )
        conn.commit()
    return {"mensaje": "Orden de compra actualizada"}


def _proveedor_values(payload):
    nombre = str(payload.get("nombre") or "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre del proveedor es obligatorio.")
    return (
        nombre,
        str(payload.get("domicilio") or "").strip(),
        str(payload.get("telefono_1") or "").strip(),
        str(payload.get("telefono_2") or "").strip(),
        str(payload.get("condicion_pago") or "").strip(),
        str(payload.get("condicion_iva") or "").strip(),
        str(payload.get("cuit") or "").strip(),
        str(payload.get("codigo_postal") or "").strip(),
        str(payload.get("localidad") or "").strip(),
    )


@app.post("/compras/proveedores")
async def compras_crear_proveedor(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    values = _proveedor_values(await request.json())
    with get_sqlite_connection() as conn:
        conn.execute(
            '''
            INSERT INTO proveedores (
                "NOM_PROVEE", "DOMICILIO", "TELÉFONO_1", "TELÉFONO_2",
                "DESC_COND", "COND_IVA", "CUIT", "C_POSTAL", "LOCALIDAD"
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            values,
        )
        conn.commit()
    return {"mensaje": "Proveedor creado"}


@app.put("/compras/proveedores/{proveedor_id}")
async def compras_actualizar_proveedor(proveedor_id: int, request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    values = _proveedor_values(await request.json())
    with get_sqlite_connection() as conn:
        if conn.execute("SELECT id FROM proveedores WHERE id = ?", (proveedor_id,)).fetchone() is None:
            raise HTTPException(status_code=404, detail="Proveedor no encontrado.")
        conn.execute(
            '''
            UPDATE proveedores SET
                "NOM_PROVEE" = ?, "DOMICILIO" = ?, "TELÉFONO_1" = ?, "TELÉFONO_2" = ?,
                "DESC_COND" = ?, "COND_IVA" = ?, "CUIT" = ?, "C_POSTAL" = ?, "LOCALIDAD" = ?
            WHERE id = ?
            ''',
            (*values, proveedor_id),
        )
        conn.commit()
    return {"mensaje": "Proveedor actualizado"}


@app.delete("/compras/proveedores/{proveedor_id}")
def compras_eliminar_proveedor(proveedor_id: int, request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    with get_sqlite_connection() as conn:
        if conn.execute("SELECT id FROM proveedores WHERE id = ?", (proveedor_id,)).fetchone() is None:
            raise HTTPException(status_code=404, detail="Proveedor no encontrado.")
        conn.execute("DELETE FROM proveedores WHERE id = ?", (proveedor_id,))
        conn.commit()
    return {"mensaje": "Proveedor eliminado"}




def _ensure_orden_compra_solicitudes(conn):
    conn.executescript(
        '''
        CREATE TABLE IF NOT EXISTS OrdenCompraDetalleSolicitud (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            orden_compra_detalle_id INTEGER NOT NULL,
            solicitud_compra_detalle_id INTEGER NOT NULL UNIQUE,
            fecha_vinculacion TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (orden_compra_detalle_id) REFERENCES OrdenCompraDetalle (id) ON DELETE CASCADE,
            FOREIGN KEY (solicitud_compra_detalle_id) REFERENCES solicitud_compra_detalle (id) ON DELETE RESTRICT,
            UNIQUE (orden_compra_detalle_id, solicitud_compra_detalle_id)
        );
        CREATE INDEX IF NOT EXISTS idx_oc_detalle_solicitud_oc_detalle
            ON OrdenCompraDetalleSolicitud (orden_compra_detalle_id);
        '''
    )


@app.get("/compras/solicitudes-items")
def compras_listar_items_solicitud(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    with get_sqlite_connection() as conn:
        _ensure_estado_detalle_compra(conn)
        _ensure_orden_compra_solicitudes(conn)
        filas = conn.execute(
            '''
            SELECT d.id, sc.numero_solicitud, d.descripcion_articulo, d.cantidad,
                   COALESCE(d.unidad_medida, '') AS unidad_medida,
                   COALESCE(cc.nombre, '') AS centro_costo,
                   COALESCE(d.estado_aprobacion, 'Pendiente') AS estado_aprobacion,
                   CASE
                       WHEN vinculacion.id IS NOT NULL THEN 'Gestionado'
                       WHEN COALESCE(d.estado_aprobacion, 'Pendiente') = 'Aprobada' THEN 'Pendiente de gestionar'
                       ELSE COALESCE(d.estado_aprobacion, 'Pendiente')
                   END AS estado_gestion,
                   vinculacion.orden_compra_detalle_id
            FROM solicitud_compra_detalle d
            INNER JOIN solicitud_compra sc ON sc.id = d.id_solicitud
            LEFT JOIN centros_costos cc ON cc.id = d.id_centro_costo
            LEFT JOIN OrdenCompraDetalleSolicitud vinculacion
                   ON vinculacion.solicitud_compra_detalle_id = d.id
            WHERE COALESCE(d.estado_aprobacion, 'Pendiente') <> 'Rechazada'
            ORDER BY sc.fecha_solicitud DESC, d.id DESC
            '''
        ).fetchall()
    return [dict(fila) for fila in filas]


def generar_pdf_orden_compra(path_pdf, orden, proveedor, detalles, contactos):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader

    page_width, page_height = A4
    margin = 22
    content_width = page_width - (margin * 2)
    rows_per_page = 16
    pages = [detalles[index:index + rows_per_page] for index in range(0, len(detalles), rows_per_page)] or [[]]
    pdf = canvas.Canvas(path_pdf, pagesize=A4)
    pdf.setTitle(f"Orden de Compra {orden['numero_oc']}")

    def text(value):
        return str(value or "").replace("\n", " ")

    def draw_text(value, x, y, max_width, font="Helvetica", size=8, italic=False):
        value = text(value)
        font_name = f"{font}-Oblique" if italic and font == "Helvetica" else font
        while value and stringWidth(value, font_name, size) > max_width:
            value = value[:-1]
        pdf.setFont(font_name, size)
        pdf.drawString(x, y, value)

    def money(value):
        return f"$ {float(value or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def format_date(value):
        try:
            return datetime.strptime(text(value), "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            return text(value)

    for page_index, page_details in enumerate(pages, start=1):
        y = page_height - margin
        pdf.setStrokeColor(colors.black)
        pdf.setLineWidth(0.8)
        pdf.rect(margin, y - 48, content_width, 48)
        pdf.line(margin + 168, y, margin + 168, y - 48)
        pdf.line(margin + 348, y, margin + 348, y - 48)
        if os.path.exists(MEMBRETE_LOGO_PATH):
            pdf.drawImage(ImageReader(MEMBRETE_LOGO_PATH), margin + 12, y - 41, width=110, height=33, preserveAspectRatio=True, mask='auto')
        pdf.setFont("Helvetica-BoldOblique", 14)
        pdf.drawCentredString(margin + 258, y - 29, "ORDEN DE COMPRA")
        pdf.setFont("Helvetica", 7)
        pdf.drawString(margin + 360, y - 15, "Fecha de Rev.: 19/07/2019")
        pdf.drawString(margin + 360, y - 26, f"Página: {page_index} de {len(pages)}")
        pdf.drawString(margin + 360, y - 37, "Revisión: 2")
        y -= 55

        if page_index == 1:
            pdf.setFont("Helvetica-BoldOblique", 9)
            pdf.drawString(margin, y, EMPRESA_MEMBRETE["razon_social"])
            pdf.setFont("Helvetica-Oblique", 8)
            pdf.drawString(margin, y - 11, "Santa Catalina N° 551 S/Colectora Ruta 66, Palpalá CP 4612")
            pdf.drawString(margin, y - 22, "San Salvador de Jujuy - Teléfono/Fax: 0388-4052871")
            pdf.drawString(margin, y - 33, f"CUIT: {EMPRESA_MEMBRETE['cuit']}")
            pdf.rect(margin + 325, y - 42, content_width - 325, 48)
            pdf.setFont("Helvetica-BoldOblique", 8)
            pdf.drawString(margin + 338, y - 13, "N° Orden de Compra:")
            pdf.drawString(margin + 338, y - 29, "Fecha:")
            pdf.drawString(margin + 338, y - 40, "N° Cotización Proveedor:")
            pdf.setFont("Helvetica", 9)
            pdf.drawString(margin + 470, y - 13, text(orden["numero_oc"]))
            pdf.drawString(margin + 470, y - 29, format_date(orden["fecha"]))
            pdf.drawString(margin + 470, y - 40, text(orden["cotizacion"]))
            y -= 55

            pdf.setFont("Helvetica-BoldOblique", 9)
            pdf.drawString(margin, y, "Datos del Proveedor")
            provider_top = y - 3
            pdf.setFillColor(colors.HexColor("#eef4df"))
            pdf.rect(margin, provider_top - 58, content_width, 58, fill=1, stroke=1)
            pdf.setFillColor(colors.black)
            labels = [("Nombre", proveedor["nombre"], margin + 4, provider_top - 13), ("CUIT", proveedor["cuit"], margin + 4, provider_top - 26), ("Domicilio", proveedor["domicilio"], margin + 4, provider_top - 39), ("IVA", proveedor["condicion_iva"], margin + 4, provider_top - 52), ("Código Postal", proveedor["codigo_postal"], margin + 292, provider_top - 26), ("Localidad", proveedor["localidad"], margin + 292, provider_top - 39), ("Teléfono 1", proveedor["telefono_1"], margin + 292, provider_top - 52)]
            for label, value, x, row_y in labels:
                pdf.setFont("Helvetica-BoldOblique", 7)
                pdf.drawString(x, row_y, f"{label}:")
                draw_text(value, x + 55, row_y, 205, "Helvetica-Oblique", 7)
            y = provider_top - 70

        pdf.setFont("Helvetica-Oblique", 8)
        pdf.drawString(margin, y, "Agradeceremos nos provean de lo siguiente:")
        y -= 4
        columns = [("N°", 19), ("Código\nproveedor", 55), ("Cant.", 35), ("Descripción", 235), ("Precio unitario", 74), ("NETO S/IVA", 74), ("Centro de Costo", 72)]
        header_height = 23
        x = margin
        pdf.setFillColor(colors.HexColor("#dce6f1"))
        pdf.rect(margin, y - header_height, content_width, header_height, fill=1, stroke=1)
        pdf.setFillColor(colors.black)
        for label, width in columns:
            pdf.rect(x, y - header_height, width, header_height, fill=0, stroke=1)
            pdf.setFont("Helvetica-BoldOblique", 7)
            for line_index, line in enumerate(label.split("\n")):
                pdf.drawCentredString(x + width / 2, y - 10 - (line_index * 8), line)
            x += width
        y -= header_height
        row_height = 20
        for row_index, item in enumerate(page_details, start=(page_index - 1) * rows_per_page + 1):
            x = margin
            values = [str(row_index), item["codigo_proveedor"], f"{float(item['cantidad'] or 0):g}", item["descripcion"], money(item["precio_unitario"]), money(item["subtotal"]), item["centro_costo"]]
            for (label, width), value in zip(columns, values):
                pdf.rect(x, y - row_height, width, row_height, fill=0, stroke=1)
                if label in {"N°", "Cant.", "Precio unitario", "NETO S/IVA"}:
                    pdf.setFont("Helvetica", 7)
                    pdf.drawCentredString(x + width / 2, y - 13, text(value))
                else:
                    draw_text(value, x + 3, y - 13, width - 6, "Helvetica-Oblique", 7)
                x += width
            y -= row_height

        if page_index == len(pages):
            total = sum(float(item["subtotal"] or 0) for item in detalles)
            pdf.setFillColor(colors.HexColor("#f4dede"))
            pdf.rect(margin + content_width - 146, y - 20, 146, 20, fill=1, stroke=1)
            pdf.setFillColor(colors.black)
            pdf.setFont("Helvetica-BoldOblique", 8)
            pdf.drawCentredString(margin + content_width - 73, y - 13, money(total))
            y -= 34
            pdf.setFont("Helvetica-BoldOblique", 9)
            pdf.drawString(margin, y, "NOTA:")
            pdf.setFont("Helvetica-Oblique", 6.5)
            notes = ["(1) SMG SRL es agente de retención del Impuesto a las Ganancias; informar situación impositiva.", "(2) Los pagos se realizarán contra factura original o proforma, remito de entrega y OC autorizada.", "(3) En factura, hacer referencia al número de Orden de Compra."]
            for note in notes:
                y -= 10
                pdf.setFillColor(colors.red)
                pdf.drawString(margin, y, note)
            pdf.setFillColor(colors.black)
            y -= 16
            pdf.setFont("Helvetica-BoldOblique", 8)
            pdf.drawString(margin, y, "Observaciones:")
            draw_text(orden["observaciones"], margin + 82, y, content_width - 82, "Helvetica", 8)
            y -= 18
            for label, value in [("Moneda", orden["moneda"]), ("Tipo de cambio", orden["tipo_cambio"]), ("Forma de pago", orden["forma_pago"]), ("Lugar de entrega", orden["lugar_entrega"]), ("Transporte", orden["transporte"]), ("Validez de la oferta", orden["validez_oferta"])]:
                pdf.setFont("Helvetica-BoldOblique", 7.5)
                pdf.drawRightString(margin + 95, y, f"{label}:")
                pdf.rect(margin + 100, y - 4, content_width - 100, 14, fill=0, stroke=1)
                draw_text(value, margin + 104, y, content_width - 108, "Helvetica-Oblique", 7.5)
                y -= 14
            pdf.setFont("Helvetica-BoldOblique", 7.5)
            pdf.drawString(margin, y, "Contactos:")
            for contact in contactos[:2]:
                pdf.setFont("Helvetica-Oblique", 7.5)
                pdf.drawString(margin + 58, y, text(contact.get("nombre")))
                pdf.drawString(margin + 215, y, text(contact.get("correo")))
                pdf.drawString(margin + 390, y, text(contact.get("telefono")))
                y -= 11
            y -= 8
            for label, value in [("Solicitó", orden["solicitante"]), ("Aprobó", orden["aprobador"]), ("Recepción Proveedor", "")]:
                width = content_width / 3
                x = margin + (["Solicitó", "Aprobó", "Recepción Proveedor"].index(label) * width)
                pdf.rect(x, y - 43, width, 43, fill=0, stroke=1)
                pdf.setFont("Helvetica-BoldOblique", 7)
                pdf.drawString(x + 4, y - 9, f"{label}:")
                pdf.setFont("Helvetica-Oblique", 8)
                pdf.drawString(x + 5, y - 30, text(value))
                pdf.setFont("Helvetica-Oblique", 6.5)
                pdf.drawString(x + 5, y - 39, "Firma y Aclaración")
        pdf.showPage()
    pdf.save()


@app.post("/compras/ordenes-compra")
async def compras_guardar_orden_compra(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    perfil = _usuario_autenticado(request) or {}
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    proveedor_id = payload.get("proveedor_id")
    items = payload.get("items", [])
    if not proveedor_id:
        raise HTTPException(status_code=400, detail="Debe seleccionar un proveedor.")
    if not isinstance(items, list) or not items:
        raise HTTPException(status_code=400, detail="Debe cargar al menos un artículo.")
    if len(items) > 25:
        raise HTTPException(status_code=400, detail="La orden de compra admite hasta 25 artículos.")

    contactos = payload.get("contactos", [])
    if not isinstance(contactos, list):
        contactos = []
    contactos = [contacto for contacto in contactos if isinstance(contacto, dict)][:2]

    with get_sqlite_connection() as conn:
        _ensure_orden_compra_solicitudes(conn)
        _ensure_orden_compra_historial(conn)
        proveedor = conn.execute(
            '''
            SELECT id, "NOM_PROVEE" AS nombre, "DOMICILIO" AS domicilio,
                   "TELÉFONO_1" AS telefono_1, "TELÉFONO_2" AS telefono_2,
                   "COND_IVA" AS condicion_iva, "CUIT" AS cuit,
                   "C_POSTAL" AS codigo_postal, "LOCALIDAD" AS localidad
            FROM proveedores WHERE id = ?
            ''',
            (proveedor_id,),
        ).fetchone()
        if proveedor is None:
            raise HTTPException(status_code=400, detail="El proveedor seleccionado no existe.")
        ultimo = conn.execute("SELECT numero_oc FROM OrdenCompra ORDER BY id DESC LIMIT 1").fetchone()
        ultimo_numero = ''.join(ch for ch in str(ultimo[0] if ultimo else "") if ch.isdigit())
        numero_oc = f"OC-{int(ultimo_numero or 0) + 1:05d}"
        cur = conn.execute(
            '''
            INSERT INTO OrdenCompra (
                numero_oc, fecha, proveedor_id, cotizacion, moneda, tipo_cambio,
                forma_pago, lugar_entrega, transporte, validez_oferta, observaciones,
                solicitante, aprobador, estado
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                numero_oc, str(payload.get("fecha") or datetime.now().strftime("%Y-%m-%d")), proveedor_id,
                str(payload.get("cotizacion") or ""), str(payload.get("moneda") or "PESOS ARGENTINOS"),
                payload.get("tipo_cambio") or None, str(payload.get("forma_pago") or ""),
                str(payload.get("lugar_entrega") or ""), str(payload.get("transporte") or ""),
                str(payload.get("validez_oferta") or ""), str(payload.get("observaciones") or ""),
                str(payload.get("solicitante") or ""), str(payload.get("aprobador") or ""), "Emitida",
            ),
        )
        orden_compra_id = cur.lastrowid
        detalles_pdf = []
        for item in items:
            try:
                cantidad = float(item.get("cantidad") or 0)
                precio_unitario = float(item.get("precio_unitario") or 0)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Cantidad y precio unitario deben ser numéricos.")
            ids_solicitudes = item.get("solicitud_detalle_ids") or []
            if not isinstance(ids_solicitudes, list) or len(ids_solicitudes) > 2:
                raise HTTPException(status_code=400, detail="Cada artículo admite hasta dos ítems de solicitud.")
            ids_solicitudes = list({int(item_id) for item_id in ids_solicitudes if str(item_id).strip()})
            cur_detalle = conn.execute(
                '''
                INSERT INTO OrdenCompraDetalle (
                    orden_compra_id, codigo_proveedor, articulo_id, descripcion, cantidad,
                    precio_unitario, subtotal, centro_costo_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    orden_compra_id, str(item.get("codigo_proveedor") or ""), item.get("articulo_id") or None,
                    str(item.get("descripcion") or ""), cantidad, precio_unitario,
                    cantidad * precio_unitario, item.get("centro_costo_id") or None,
                ),
            )
            detalles_pdf.append({
                "codigo_proveedor": str(item.get("codigo_proveedor") or ""),
                "descripcion": str(item.get("descripcion") or ""),
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "subtotal": cantidad * precio_unitario,
                "centro_costo": str(item.get("centro_costo_id") or ""),
            })
            for solicitud_detalle_id in ids_solicitudes:
                disponible = conn.execute(
                    '''
                    SELECT d.id
                    FROM solicitud_compra_detalle d
                    LEFT JOIN OrdenCompraDetalleSolicitud vinculo
                           ON vinculo.solicitud_compra_detalle_id = d.id
                    WHERE d.id = ?
                      AND COALESCE(d.estado_aprobacion, 'Pendiente') = 'Aprobada'
                      AND vinculo.id IS NULL
                    ''',
                    (solicitud_detalle_id,),
                ).fetchone()
                if disponible is None:
                    raise HTTPException(status_code=400, detail="Uno de los ítems de solicitud ya fue gestionado o no está aprobado.")
                conn.execute(
                    "INSERT INTO OrdenCompraDetalleSolicitud (orden_compra_detalle_id, solicitud_compra_detalle_id) VALUES (?, ?)",
                    (cur_detalle.lastrowid, solicitud_detalle_id),
                )
                conn.execute(
                    "UPDATE solicitud_compra_detalle SET estado_aprobacion = 'Gestionado' WHERE id = ?",
                    (solicitud_detalle_id,),
                )
            _registrar_orden_compra_historial(conn, orden_compra_id, "Emitida", "Emisión inicial de la orden de compra.", perfil)
        conn.commit()
    orden_pdf = {
        "numero_oc": numero_oc,
        "fecha": str(payload.get("fecha") or datetime.now().strftime("%Y-%m-%d")),
        "cotizacion": str(payload.get("cotizacion") or ""),
        "moneda": str(payload.get("moneda") or "PESOS ARGENTINOS"),
        "tipo_cambio": str(payload.get("tipo_cambio") or ""),
        "forma_pago": str(payload.get("forma_pago") or ""),
        "lugar_entrega": str(payload.get("lugar_entrega") or ""),
        "transporte": str(payload.get("transporte") or ""),
        "validez_oferta": str(payload.get("validez_oferta") or ""),
        "observaciones": str(payload.get("observaciones") or ""),
        "solicitante": str(payload.get("solicitante") or ""),
        "aprobador": str(payload.get("aprobador") or ""),
    }
    pdf_path = os.path.join(DOC_ALMACEN_DIR, f"{numero_oc}.pdf")
    generar_pdf_orden_compra(pdf_path, orden_pdf, dict(proveedor), detalles_pdf, contactos)
    return FileResponse(pdf_path, media_type="application/pdf", filename=os.path.basename(pdf_path))


@app.get("/compras/solicitudes/{solicitud_id}")
def compras_obtener_solicitud(solicitud_id: int, request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    perfil = _usuario_autenticado(request) or {}
    with get_sqlite_connection() as conn:
        _ensure_estado_detalle_compra(conn)
        cabecera = conn.execute(
            """
            SELECT sc.id, sc.numero_solicitud, sc.fecha_solicitud,
                   COALESCE(p.nombre, '') AS solicitante,
                   COALESCE(pr.nombre, '') AS destino_compra,
                   COALESCE(pc.nombre, '') AS prioridad,
                   COALESCE(ec.nombre, '') AS estado,
                   COALESCE(sc.observaciones, '') AS observaciones
            FROM solicitud_compra sc
            LEFT JOIN personal p ON p.legajo = sc.id_solicitante
            LEFT JOIN proyectos pr ON pr.id = sc.id_proyecto
            LEFT JOIN prioridades_compra pc ON pc.id = sc.id_prioridad
            LEFT JOIN estados_compra ec ON ec.id = sc.id_estado
            WHERE sc.id = ?
            """,
            (solicitud_id,),
        ).fetchone()
        if cabecera is None:
            raise HTTPException(status_code=404, detail="Solicitud de compra no encontrada.")

        items = conn.execute(
            """
            SELECT d.id,
                   COALESCE(cc.nombre, '') AS centro_costo,
                   COALESCE(d.id_vehiculo, '') AS vehiculo,
                   d.descripcion_articulo, d.motivo_compra, d.cantidad,
                   COALESCE(d.unidad_medida, '') AS unidad_medida,
                   COALESCE(d.observacion, '') AS observacion,
                   COALESCE(d.estado_aprobacion, 'Pendiente') AS estado_aprobacion,
                   COALESCE(d.motivo_rechazo, '') AS motivo_rechazo,
                   COALESCE(d.usuario_resolucion, '') AS usuario_resolucion
            FROM solicitud_compra_detalle d
            LEFT JOIN centros_costos cc ON cc.id = d.id_centro_costo
            WHERE d.id_solicitud = ?
            ORDER BY d.id
            """,
            (solicitud_id,),
        ).fetchall()

    solicitud = dict(cabecera)
    solicitud["items"] = [dict(item) for item in items]
    solicitud["puede_aprobar"] = str(perfil.get("tipo_usuario", "")).upper() == "ADMINISTRADOR"
    return solicitud


@app.post("/compras/solicitudes/{solicitud_id}/resolver")
async def compras_resolver_solicitud(solicitud_id: int, request: Request):
    perfil = _requiere_administrador(request)
    usuario_resolucion = str(perfil.get("usuario") or perfil.get("nombre_apellido") or "").strip()
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    accion = str(payload.get("accion", "")).strip().lower()
    ids_items = payload.get("ids_items", [])
    motivo_rechazo = str(payload.get("motivo_rechazo", "")).strip()
    if accion not in {"aprobar", "rechazar"}:
        raise HTTPException(status_code=400, detail="Acción de aprobación inválida.")
    if not isinstance(ids_items, list) or not ids_items:
        raise HTTPException(status_code=400, detail="Seleccione al menos un ítem.")
    if accion == "rechazar" and not motivo_rechazo:
        raise HTTPException(status_code=400, detail="Debe indicar el motivo del rechazo.")

    try:
        ids_items = sorted({int(item_id) for item_id in ids_items})
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Los ítems seleccionados no son válidos.")

    with get_sqlite_connection() as conn:
        _ensure_prioridades_y_estados_compras(conn)
        _ensure_estado_detalle_compra(conn)
        solicitud = conn.execute("SELECT id FROM solicitud_compra WHERE id = ?", (solicitud_id,)).fetchone()
        if solicitud is None:
            raise HTTPException(status_code=404, detail="Solicitud de compra no encontrada.")

        placeholders = ", ".join("?" for _ in ids_items)
        items_validos = conn.execute(
            f"SELECT id FROM solicitud_compra_detalle WHERE id_solicitud = ? AND id IN ({placeholders})",
            (solicitud_id, *ids_items),
        ).fetchall()
        if len(items_validos) != len(ids_items):
            raise HTTPException(status_code=400, detail="Hay ítems que no pertenecen a esta solicitud.")

        nuevo_estado = "Aprobada" if accion == "aprobar" else "Rechazada"
        conn.execute(
            f"UPDATE solicitud_compra_detalle SET estado_aprobacion = ?, motivo_rechazo = ?, usuario_resolucion = ? WHERE id_solicitud = ? AND id IN ({placeholders})",
            (nuevo_estado, "" if accion == "aprobar" else motivo_rechazo, usuario_resolucion, solicitud_id, *ids_items),
        )

        resumen = conn.execute(
            """
            SELECT COUNT(*) AS total,
                   SUM(CASE WHEN estado_aprobacion = 'Aprobada' THEN 1 ELSE 0 END) AS aprobados,
                   SUM(CASE WHEN estado_aprobacion = 'Rechazada' THEN 1 ELSE 0 END) AS rechazados
            FROM solicitud_compra_detalle WHERE id_solicitud = ?
            """,
            (solicitud_id,),
        ).fetchone()
        total = int(resumen["total"] or 0)
        aprobados = int(resumen["aprobados"] or 0)
        rechazados = int(resumen["rechazados"] or 0)
        if total and aprobados == total:
            estado_general = "Aprobada"
        elif total and rechazados == total:
            estado_general = "Rechazada"
        elif aprobados:
            estado_general = "Aprobada parcialmente"
        else:
            estado_general = "En revisión"

        estado = conn.execute("SELECT id FROM estados_compra WHERE nombre = ?", (estado_general,)).fetchone()
        conn.execute(
            "UPDATE solicitud_compra SET id_estado = ?, fecha_modificacion = datetime('now') WHERE id = ?",
            (estado["id"], solicitud_id),
        )
        conn.commit()

    return {"estado": estado_general}


@app.get("/compras/solicitudes/{solicitud_id}/pdf")
def compras_descargar_solicitud_pdf(solicitud_id: int, request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    perfil = _usuario_autenticado(request) or {}
    with get_sqlite_connection() as conn:
        _ensure_estado_detalle_compra(conn)
        solicitud = conn.execute(
            """
            SELECT sc.numero_solicitud, sc.fecha_solicitud,
                   COALESCE(p.nombre, '') AS solicitante,
                   COALESCE(pr.nombre, '') AS destino_compra,
                   COALESCE(pc.nombre, '') AS prioridad,
                   COALESCE(ec.nombre, 'En revisión') AS estado,
                   COALESCE(sc.observaciones, '') AS observaciones
            FROM solicitud_compra sc
            LEFT JOIN personal p ON p.legajo = sc.id_solicitante
            LEFT JOIN proyectos pr ON pr.id = sc.id_proyecto
            LEFT JOIN prioridades_compra pc ON pc.id = sc.id_prioridad
            LEFT JOIN estados_compra ec ON ec.id = sc.id_estado
            WHERE sc.id = ?
            """,
            (solicitud_id,),
        ).fetchone()
        items = conn.execute(
            """
            SELECT d.descripcion_articulo, d.cantidad, COALESCE(d.unidad_medida, '') AS unidad_medida,
                   COALESCE(d.id_centro_costo, '') AS centro_costo, COALESCE(d.id_vehiculo, '') AS vehiculo,
                   d.motivo_compra, COALESCE(d.motivo_rechazo, '') AS motivo_rechazo,
                   COALESCE(d.estado_aprobacion, 'Pendiente') AS estado_aprobacion,
                   COALESCE(d.usuario_resolucion, '') AS usuario_resolucion
            FROM solicitud_compra_detalle d
            WHERE d.id_solicitud = ? ORDER BY d.id
            """,
            (solicitud_id,),
        ).fetchall()

    if solicitud is None:
        raise HTTPException(status_code=404, detail="Solicitud de compra no encontrada.")

    pdf_path = os.path.join(DOC_ALMACEN_DIR, f"{solicitud['numero_solicitud']}.pdf")
    if False and not os.path.isfile(pdf_path):
        raise HTTPException(status_code=404, detail="El PDF de la solicitud no está disponible.")

    motivos_rechazo, detalle_pdf = [], []
    for item in items:
        motivo_pdf = str(item["motivo_compra"] or "").strip()
        estado_item = str(item["estado_aprobacion"] or "Pendiente").strip()
        usuario_item = str(item["usuario_resolucion"] or "").strip()
        if estado_item in {"Aprobada", "Rechazada"} and usuario_item:
            accion_item = "Aprobado" if estado_item == "Aprobada" else "Rechazado"
            motivo_pdf = f"{motivo_pdf} | {accion_item} por: {usuario_item}".strip(" |")
        motivo = str(item["motivo_rechazo"] or "").strip()
        if motivo:
            detalle_rechazo = f"{motivo} — Rechazado por: {usuario_item}" if usuario_item else motivo
            motivos_rechazo.append(detalle_rechazo)
        detalle_pdf.append(
            (
                str(item["descripcion_articulo"] or ""), str(item["cantidad"] or ""),
                str(item["unidad_medida"] or ""), str(item["centro_costo"] or ""),
                str(item["vehiculo"] or ""), motivo_pdf,
            )
        )
    bloque_envio = [
        ("Destino de compra", solicitud["destino_compra"]),
        ("Observaciones", solicitud["observaciones"]),
    ]
    if motivos_rechazo:
        bloque_envio.append(("Motivos de rechazo", "; ".join(motivos_rechazo)))

    generar_pdf_remito(
        pdf_path,
        "SOLICITUD DE COMPRAS",
        [
            ("Solicitud Nro", solicitud["numero_solicitud"]),
            ("Fecha", solicitud["fecha_solicitud"]),
            ("Solicitante", solicitud["solicitante"]),
            ("Prioridad", solicitud["prioridad"]),
            ("Estado actual", solicitud["estado"]),
        ],
        detalle_pdf,
        columnas=[("Articulo", 40), ("Cantidad", 270), ("Unidad", 320), ("C.C.", 370), ("C.I.", 415), ("Motivo", 460)],
        incluir_firmas=True,
        firma_entrega_nombre="Solicitante",
        firma_recibe_nombre="Responsable",
        bloque_envio=bloque_envio,
        bloque_envio_titulo="Resumen de Compra",
        bloque_pie_izq=[("Estado", solicitud["estado"])],
        bloque_pie_der=[("Total de items", str(len(items)))],
        observaciones_pie=solicitud["observaciones"],
        departamento="Departamento de Compras",
        usuario_pie=f"Usuario {str(perfil.get('usuario') or perfil.get('nombre_apellido') or '').strip()}".strip(),
        columnas_multilinea={0, 5},
    )

    return FileResponse(
        pdf_path,
        media_type="application/pdf",
        filename=os.path.basename(pdf_path),
    )


def generar_numero_solicitud(conn):
    ultimo_registro = conn.execute(
        "SELECT numero_solicitud FROM solicitud_compra ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if ultimo_registro and ultimo_registro[0]:
        ultimo_numero_texto = ultimo_registro[0]
        ultimo_numero_digitos = ''.join(ch for ch in ultimo_numero_texto if ch.isdigit())
        try:
            ultimo_id = int(ultimo_numero_digitos)
        except ValueError:
            ultimo_id = 0
    else:
        ultimo_id = 0
    return f"SC-{ultimo_id + 1:05d}"


def _ensure_prioridades_y_estados_compras(conn):
    prioridades = [
        r[0]
        for r in conn.execute(
            "SELECT id FROM prioridades_compra WHERE COALESCE(activo, 1) = 1"
        ).fetchall()
    ]
    estados = [
        r[0]
        for r in conn.execute(
            "SELECT id FROM estados_compra WHERE COALESCE(activo, 1) = 1"
        ).fetchall()
    ]

    if not prioridades:
        conn.executemany(
            "INSERT OR IGNORE INTO prioridades_compra (nombre, activo) VALUES (?, ?)",
            [
                ("Baja", 1),
                ("Media", 1),
                ("Alta", 1),
                ("Urgente", 1),
            ],
        )
    if not estados:
        conn.executemany(
            "INSERT OR IGNORE INTO estados_compra (nombre, activo) VALUES (?, ?)",
            [
                ("Borrador", 1),
                ("En revisión", 1),
                ("Aprobada", 1),
                ("Rechazada", 1),
                ("Comprada", 1),
            ],
        )
    if not prioridades or not estados:
        conn.commit()
    conn.execute(
        "INSERT OR IGNORE INTO estados_compra (nombre, activo) VALUES (?, ?)",
        ("Aprobada parcialmente", 1),
    )
    conn.commit()


def _ensure_estado_detalle_compra(conn):
    columnas = {fila[1] for fila in conn.execute("PRAGMA table_info(solicitud_compra_detalle)").fetchall()}
    if "estado_aprobacion" not in columnas:
        conn.execute("ALTER TABLE solicitud_compra_detalle ADD COLUMN estado_aprobacion TEXT NOT NULL DEFAULT 'Pendiente'")
    if "motivo_rechazo" not in columnas:
        conn.execute("ALTER TABLE solicitud_compra_detalle ADD COLUMN motivo_rechazo TEXT")
    if "usuario_resolucion" not in columnas:
        conn.execute("ALTER TABLE solicitud_compra_detalle ADD COLUMN usuario_resolucion TEXT")
    conn.commit()


def _obtener_proyecto_id_por_nombre(conn, nombre):
    nombre_normalizado = str(nombre or "").strip()
    if not nombre_normalizado:
        return None
    row = conn.execute(
        "SELECT id FROM proyectos WHERE lower(nombre) = lower(?) LIMIT 1",
        (nombre_normalizado,),
    ).fetchone()
    if row:
        return row[0]
    codigo = ''.join(ch for ch in nombre_normalizado.upper() if ch.isalnum())[:8] or 'PROYECTO'
    codigo_base = codigo
    suffix = 1
    while conn.execute("SELECT 1 FROM proyectos WHERE codigo = ?", (codigo,)).fetchone():
        suffix += 1
        codigo = f"{codigo_base[:6]}{suffix}"
    cur = conn.execute(
        "INSERT INTO proyectos (codigo, nombre, activo) VALUES (?, ?, 1)",
        (codigo, nombre_normalizado),
    )
    return cur.lastrowid


def _ensure_detalle_compra_cc_nullable(conn):
    info = conn.execute("PRAGMA table_info(solicitud_compra_detalle)").fetchall()
    current_columns = {r[1]: r for r in info}
    fk_info = conn.execute("PRAGMA foreign_key_list(solicitud_compra_detalle)").fetchall()

    needs_migration = any(r[1] == 'id_centro_costo' and r[3] == 1 for r in info)
    if not needs_migration:
        id_maquinaria_info = current_columns.get('id_maquinaria')
        id_vehiculo_info = current_columns.get('id_vehiculo')
        if id_maquinaria_info is None or id_maquinaria_info[2].upper() != 'TEXT':
            needs_migration = True
        if id_vehiculo_info is None or id_vehiculo_info[2].upper() != 'TEXT':
            needs_migration = True
    if not needs_migration:
        for fk in fk_info:
            if fk[2] == 'vehiculos' and fk[3] in ('id_maquinaria', 'id_vehiculo') and fk[4] != 'codigo':
                needs_migration = True
                break

    if needs_migration:
        if conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='solicitud_compra_detalle_old'").fetchone():
            existing_rows = conn.execute("SELECT COUNT(*) FROM solicitud_compra_detalle_old").fetchone()[0]
            if existing_rows == 0:
                conn.execute("DROP TABLE solicitud_compra_detalle_old")
            else:
                raise RuntimeError(
                    "There is already a non-empty backup table solicitud_compra_detalle_old. "
                    "Please remove or rename it before retrying."
                )
        conn.execute("ALTER TABLE solicitud_compra_detalle RENAME TO solicitud_compra_detalle_old")
        conn.execute('''
            CREATE TABLE solicitud_compra_detalle (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_solicitud INTEGER NOT NULL,
                id_centro_costo INTEGER,
                id_vehiculo TEXT,
                id_maquinaria TEXT,
                descripcion_articulo TEXT NOT NULL,
                motivo_compra TEXT NOT NULL,
                cantidad REAL NOT NULL,
                unidad_medida TEXT,
                observacion TEXT,
                FOREIGN KEY (id_solicitud) REFERENCES solicitud_compra (id) ON DELETE CASCADE,
                FOREIGN KEY (id_centro_costo) REFERENCES centros_costos (id) ON DELETE RESTRICT,
                FOREIGN KEY (id_vehiculo) REFERENCES vehiculos (codigo) ON DELETE SET NULL,
                FOREIGN KEY (id_maquinaria) REFERENCES vehiculos (codigo) ON DELETE SET NULL
            )
        ''')
        conn.execute(
            '''
            INSERT INTO solicitud_compra_detalle (
                id, id_solicitud, id_centro_costo, id_vehiculo, id_maquinaria,
                descripcion_articulo, motivo_compra, cantidad, unidad_medida, observacion
            )
            SELECT 
                id, id_solicitud, id_centro_costo, id_vehiculo, id_maquinaria,
                descripcion_articulo, motivo_compra, cantidad, unidad_medida, observacion
            FROM solicitud_compra_detalle_old
            '''
        )
        conn.execute("DROP TABLE solicitud_compra_detalle_old")


@app.post("/compras/emitir")
async def compras_emitir(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect

    perfil = _usuario_autenticado(request) or {}
    usuario_logueado = str(perfil.get("usuario") or perfil.get("nombre_apellido") or "").strip()

    try:
        payload = await request.json()
    except Exception:
        payload = {}

    numero_solicitud = str(payload.get("numero_solicitud", "")).strip()
    fecha_solicitud = str(payload.get("fecha_solicitud", "")).strip() or datetime.now().strftime("%Y-%m-%d")
    id_solicitante = payload.get("id_solicitante")
    destino_compra = str(payload.get("destino_compra", "")).strip()
    prioridad_id = payload.get("prioridad_id")
    observaciones = str(payload.get("observaciones", "")).strip()
    items = payload.get("items", [])

    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="La solicitud debe contener al menos un artículo.")
    if not destino_compra:
        raise HTTPException(status_code=400, detail="Debe indicar el destino de compra.")
    if not id_solicitante:
        raise HTTPException(status_code=400, detail="Debe indicar quien solicita.")

    with get_sqlite_connection() as conn:
        _ensure_prioridades_y_estados_compras(conn)
        _ensure_detalle_compra_cc_nullable(conn)
        _ensure_estado_detalle_compra(conn)
        numero_solicitud = generar_numero_solicitud(conn)
        id_proyecto = _obtener_proyecto_id_por_nombre(conn, destino_compra)

        cur = conn.execute(
            "INSERT INTO solicitud_compra (numero_solicitud, fecha_solicitud, id_solicitante, id_proyecto, id_prioridad, observaciones, id_estado, fecha_creacion) VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))",
            (
                numero_solicitud,
                fecha_solicitud,
                id_solicitante,
                id_proyecto,
                prioridad_id,
                observaciones,
                2,
            ),
        )
        solicitud_id = cur.lastrowid

        for item in items:
            descripcion_articulo = str(item.get("detalle", "")).strip()
            motivo_compra = str(item.get("motivo", "")).strip()
            unidad_medida = str(item.get("unidad", "")).strip()
            cantidad = item.get("cantidad")
            try:
                cantidad_value = float(cantidad)
            except (TypeError, ValueError):
                cantidad_value = 0.0
            id_centro_costo = item.get("centroCosto") or None
            id_vehiculo = item.get("ci") or None

            conn.execute(
                "INSERT INTO solicitud_compra_detalle (id_solicitud, id_centro_costo, id_vehiculo, descripcion_articulo, motivo_compra, cantidad, unidad_medida, observacion) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    solicitud_id,
                    id_centro_costo,
                    id_vehiculo,
                    descripcion_articulo,
                    motivo_compra,
                    cantidad_value,
                    unidad_medida,
                    "",
                ),
            )

        conn.commit()

    with get_sqlite_connection() as conn:
        datos_pdf = conn.execute(
            """
            SELECT COALESCE(p.nombre, '') AS solicitante,
                   COALESCE(pc.nombre, '') AS prioridad,
                   COALESCE(pr.nombre, '') AS destino_compra
            FROM solicitud_compra sc
            LEFT JOIN personal p ON p.legajo = sc.id_solicitante
            LEFT JOIN prioridades_compra pc ON pc.id = sc.id_prioridad
            LEFT JOIN proyectos pr ON pr.id = sc.id_proyecto
            WHERE sc.id = ?
            """,
            (solicitud_id,),
        ).fetchone()

    solicitante_nombre = str(datos_pdf["solicitante"] or "").strip() if datos_pdf else ""
    prioridad_nombre = str(datos_pdf["prioridad"] or "").strip() if datos_pdf else ""
    destino_nombre = str(datos_pdf["destino_compra"] or "").strip() if datos_pdf else ""

    pdf_path = os.path.join(DOC_ALMACEN_DIR, f"{numero_solicitud}.pdf")
    cabecera = [
        ("Solicitud Nº", numero_solicitud),
        ("Fecha", fecha_solicitud),
        ("Solicitante", solicitante_nombre),
        ("Prioridad", prioridad_nombre),
    ]
    bloque_envio = [
        ("Destino de compra", destino_nombre),
        ("Observaciones", observaciones),
    ]
    detalle_pdf = []
    for item in items:
        detalle_pdf.append(
            (
                str(item.get("detalle", "")),
                str(item.get("cantidad", "")),
                str(item.get("unidad", "")),
                str(item.get("centroCosto", "")),
                str(item.get("ci", "")),
                str(item.get("motivo", "")),
            )
        )
    columnas = [
        ("Artículo", 40),
        ("Cantidad", 270),
        ("Unidad", 320),
        ("C.C.", 370),
        ("C.I.", 415),
        ("Motivo", 460),
    ]
    generar_pdf_remito(
        pdf_path,
        "SOLICITUD DE COMPRAS",
        cabecera,
        detalle_pdf,
        columnas=columnas,
        incluir_firmas=True,
        firma_entrega_nombre="Solicitante",
        firma_recibe_nombre="Responsable",
        bloque_envio=bloque_envio,
        bloque_envio_titulo="Resumen de Compra",
        bloque_pie_izq=[("Estado", "En revisión")],
        bloque_pie_der=[("Total de ítems", str(len(items)))],
        observaciones_pie=observaciones,
        departamento="Departamento de Compras",
        usuario_pie=f"Usuario {usuario_logueado}" if usuario_logueado else "Usuario",
        columnas_multilinea={0, 5},
    )

    return FileResponse(pdf_path, media_type="application/pdf", filename=os.path.basename(pdf_path))


@app.get("/dashboard.html", response_class=HTMLResponse)
def dashboard_html(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(DASHBOARD_PATH)


@app.get("/gestion_operativa", response_class=HTMLResponse)
def gestion_operativa_view(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(GESTION_OPERATIVA_PATH)
    
@app.put("/estado/{id_viaje}")
def cambiar_estado(id_viaje: int, estado: str):
    with get_sqlite_connection() as conn:
        row = conn.execute("SELECT raw_json FROM viajes WHERE id = ?", (id_viaje,)).fetchone()
        if row is None:
            return {"error": "Viaje no encontrado"}

        viaje = parse_json_dict(row["raw_json"], default={})
        viaje["estado"] = estado
        guardar_viaje_sql(conn, viaje)
        conn.commit()

    return {"mensaje": "Estado actualizado"}


@app.get("/form", response_class=HTMLResponse)
def form_viaje(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(FORM_VIAJE_PATH)


@app.get("/recursos_form", response_class=HTMLResponse)
def form_recursos(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(FORM_RECURSOS_PATH)


@app.get("/ordenes")
def obtener_ordenes():
    return obtener_ordenes_data()


@app.get("/gestion_operativa/filtros")
def gestion_operativa_filtros():
    with get_sqlite_connection() as conn:
        reconciliar_gestion_operativa_cerradas(conn)

        ordenes = [
            str(r["valor"])
            for r in conn.execute(
                """
                SELECT DISTINCT TRIM(COALESCE(nro_orden, '')) AS valor
                FROM gestion_operativa
                WHERE TRIM(COALESCE(nro_orden, '')) <> ''
                ORDER BY LOWER(TRIM(COALESCE(nro_orden, '')))
                """
            ).fetchall()
        ]

        proyectos = [
            str(r["valor"])
            for r in conn.execute(
                """
                SELECT DISTINCT TRIM(COALESCE(proyecto, '')) AS valor
                FROM gestion_operativa
                WHERE TRIM(COALESCE(proyecto, '')) <> ''
                ORDER BY LOWER(TRIM(COALESCE(proyecto, '')))
                """
            ).fetchall()
        ]

        centros = [
            str(r["valor"])
            for r in conn.execute(
                """
                SELECT DISTINCT TRIM(COALESCE(centro_costo, '')) AS valor
                FROM gestion_operativa
                WHERE TRIM(COALESCE(centro_costo, '')) <> ''
                ORDER BY LOWER(TRIM(COALESCE(centro_costo, '')))
                """
            ).fetchall()
        ]

        empleados = [
            str(r["valor"])
            for r in conn.execute(
                """
                SELECT DISTINCT TRIM(COALESCE(nombre, '')) AS valor
                FROM gestion_operativa
                WHERE TRIM(COALESCE(nombre, '')) <> ''
                ORDER BY LOWER(TRIM(COALESCE(nombre, '')))
                """
            ).fetchall()
        ]

    return {
        "ordenes": ordenes,
        "proyectos": proyectos,
        "centros_costo": centros,
        "empleados": empleados,
    }


@app.get("/gestion_operativa/resumen")
def gestion_operativa_resumen(
    desde: str = "",
    hasta: str = "",
    nro_orden: str = "",
    proyecto: str = "",
    centro_costo: str = "",
    empleado: str = "",
):
    with get_sqlite_connection() as conn:
        payload = _gestion_operativa_resumen_payload(
            conn,
            desde=desde,
            hasta=hasta,
            nro_orden=nro_orden,
            proyecto=proyecto,
            centro_costo=centro_costo,
            empleado=empleado,
        )
    return payload


def _gestion_operativa_resumen_payload(
    conn,
    desde: str = "",
    hasta: str = "",
    nro_orden: str = "",
    proyecto: str = "",
    centro_costo: str = "",
    empleado: str = "",
):
    where = []
    params = []
    if desde:
        where.append("date(substr(COALESCE(fecha_cierre, fecha_orden, ''), 1, 10)) >= date(?)")
        params.append(desde)
    if hasta:
        where.append("date(substr(COALESCE(fecha_cierre, fecha_orden, ''), 1, 10)) <= date(?)")
        params.append(hasta)
    if nro_orden:
        where.append("LOWER(TRIM(COALESCE(nro_orden, ''))) = ?")
        params.append(nro_orden.strip().lower())
    if proyecto:
        where.append("LOWER(COALESCE(proyecto, '')) LIKE ?")
        params.append(f"%{proyecto.strip().lower()}%")
    if centro_costo:
        where.append("LOWER(COALESCE(centro_costo, '')) LIKE ?")
        params.append(f"%{centro_costo.strip().lower()}%")
    if empleado:
        where.append("LOWER(COALESCE(nombre, '')) LIKE ?")
        params.append(f"%{empleado.strip().lower()}%")

    sql = "SELECT * FROM gestion_operativa"
    if where:
        sql += " WHERE " + " AND ".join(where)

    reconciliar_gestion_operativa_cerradas(conn)
    rows = [dict(r) for r in conn.execute(sql, tuple(params)).fetchall()]

    horas_hombre = round(sum(float(r.get("horas_totales") or 0) for r in rows), 2)
    horas_comp = round(sum(float(r.get("horas_compensables") or 0) for r in rows), 2)
    costo_mes = round(sum(float(r.get("costo_total") or 0) for r in rows), 2)
    jornadas_filtradas = round(sum(float(r.get("jornadas") or 0) for r in rows), 2)
    viajes_ejecutados = len(
        {
            str(r.get("nro_orden") or "").strip()
            for r in rows
            if str(r.get("nro_orden") or "").strip() and str(r.get("estado_orden") or "").upper() == "CERRADO"
        }
    )

    tabla_map = {}
    for r in rows:
        key = (
            str(r.get("nro_orden") or "").strip(),
            str(r.get("nombre") or "").strip(),
            str(r.get("rol") or "").strip(),
            str(r.get("proyecto") or "").strip() or "Sin proyecto",
            str(r.get("centro_costo") or "").strip() or "Sin centro",
            str(r.get("origen") or "").strip() or "Sin origen",
            str(r.get("destino") or "").strip() or "Sin destino",
        )
        if key not in tabla_map:
            tabla_map[key] = {
                "nro_orden": key[0],
                "empleado": key[1],
                "rol": key[2],
                "proyecto": key[3],
                "centro_costo": key[4],
                "origen": key[5],
                "destino": key[6],
                "jornadas": 0.0,
                "hh": 0.0,
                "compensables": 0.0,
                "viaticos": 0.0,
            }
        item = tabla_map[key]
        item["jornadas"] += float(r.get("jornadas") or 0)
        item["hh"] += float(r.get("horas_totales") or 0)
        item["compensables"] += float(r.get("horas_compensables") or 0)
        item["viaticos"] += float(r.get("viatico") or 0)

    tabla = []
    for item in tabla_map.values():
        tabla.append(
            {
                "nro_orden": item["nro_orden"],
                "empleado": item["empleado"],
                "rol": item["rol"],
                "proyecto": item["proyecto"],
                "centro_costo": item["centro_costo"],
                "origen": item["origen"],
                "destino": item["destino"],
                "jornadas": round(item["jornadas"], 2),
                "hh": round(item["hh"], 2),
                "compensables": round(item["compensables"], 2),
                "viaticos": round(item["viaticos"], 2),
            }
        )
    tabla.sort(key=lambda x: (x["hh"], x["compensables"]), reverse=True)

    def _acumular_por(campo, valor):
        m = {}
        for r in rows:
            k = str(r.get(campo) or "").strip() or f"Sin {campo}"
            m[k] = m.get(k, 0.0) + float(r.get(valor) or 0)
        return [{"label": k, "value": round(v, 2)} for k, v in sorted(m.items(), key=lambda it: it[1], reverse=True)]

    horas_proyecto = _acumular_por("proyecto", "horas_totales")
    horas_empleado = _acumular_por("nombre", "horas_totales")
    costos_proyecto = _acumular_por("proyecto", "costo_total")

    viajes_proyecto_map = {}
    for r in rows:
        proyecto_lbl = str(r.get("proyecto") or "").strip() or "Sin proyecto"
        nro = str(r.get("nro_orden") or "").strip()
        if not nro:
            continue
        viajes_proyecto_map.setdefault(proyecto_lbl, set()).add(nro)
    viajes_proyecto = [
        {"label": k, "value": len(v)} for k, v in sorted(viajes_proyecto_map.items(), key=lambda it: len(it[1]), reverse=True)
    ]

    comp_mes_map = {}
    for r in rows:
        fecha_base = str(r.get("fecha_cierre") or r.get("fecha_orden") or "")
        mes = fecha_base[:7] if len(fecha_base) >= 7 else "Sin fecha"
        comp_mes_map[mes] = comp_mes_map.get(mes, 0.0) + float(r.get("horas_compensables") or 0)
    compensables_mes = [
        {"label": k, "value": round(v, 2)} for k, v in sorted(comp_mes_map.items(), key=lambda it: it[0])
    ]

    return {
        "kpis": {
            "horas_hombre_mes": horas_hombre,
            "horas_compensables_mes": horas_comp,
            "viajes_ejecutados": viajes_ejecutados,
            "costo_operativo_mes": costo_mes,
            "jornadas_filtradas": jornadas_filtradas,
        },
        "tabla": tabla,
        "graficos": {
            "horas_por_proyecto": horas_proyecto,
            "horas_por_empleado": horas_empleado,
            "horas_compensables_por_mes": compensables_mes,
            "viajes_por_proyecto": viajes_proyecto,
            "costos_por_proyecto": costos_proyecto,
        },
    }


@app.get("/gestion_operativa/pdf")
def gestion_operativa_pdf(
    desde: str = "",
    hasta: str = "",
    nro_orden: str = "",
    proyecto: str = "",
    centro_costo: str = "",
    empleado: str = "",
):
    with get_sqlite_connection() as conn:
        payload = _gestion_operativa_resumen_payload(
            conn,
            desde=desde,
            hasta=hasta,
            nro_orden=nro_orden,
            proyecto=proyecto,
            centro_costo=centro_costo,
            empleado=empleado,
        )

    filas = payload.get("tabla", []) if isinstance(payload, dict) else []

    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        from reportlab.lib import colors
    except Exception as exc:
        return {"error": f"No se pudo generar el PDF: {exc}"}

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=landscape(A4))
    ancho, alto = landscape(A4)

    if os.path.exists(MEMBRETE_LOGO_PATH):
        try:
            logo = ImageReader(MEMBRETE_LOGO_PATH)
            c.drawImage(logo, ancho - 185, alto - 92, width=140, height=45, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    y = alto - 40
    c.setFont("Helvetica-Bold", 15)
    c.drawString(180, y, "Gestion Operativa")
    y -= 24
    c.setFont("Helvetica-Bold", 13)
    c.drawString(180, y, "Resumen Operativo")
    y -= 20

    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 14

    filtros = []
    if desde:
        filtros.append(f"Desde: {desde}")
    if hasta:
        filtros.append(f"Hasta: {hasta}")
    if nro_orden:
        filtros.append(f"Nro Orden: {nro_orden}")
    if proyecto:
        filtros.append(f"Proyecto: {proyecto}")
    if centro_costo:
        filtros.append(f"Centro: {centro_costo}")
    if empleado:
        filtros.append(f"Empleado: {empleado}")
    if filtros:
        c.drawString(40, y, " | ".join(filtros)[:165])
        y -= 16

    table_left = 30
    table_right = ancho - 30
    table_width = table_right - table_left
    col_specs = [
        ("Nro Orden", 62.0, "left"),
        ("Empleado", 126.0, "left"),
        ("Rol", 75.0, "left"),
        ("Proyecto", 85.0, "left"),
        ("Centro Costos", 78.0, "left"),
        ("Origen", 126.0, "left"),
        ("Destino", 78.0, "left"),
        ("HH", 45.0, "right"),
        ("Comp.", 58.0, "right"),
        ("Viaticos", 62.0, "right"),
    ]
    total_base = sum(width for _, width, _ in col_specs)
    scale = min(1.0, table_width / total_base) if total_base > 0 else 1.0
    columnas = []
    x_cursor = table_left
    for titulo_col, width_base, align in col_specs:
        ancho_col = width_base * scale
        columnas.append({"titulo": titulo_col, "x": x_cursor, "w": ancho_col, "align": align})
        x_cursor += ancho_col
    if columnas:
        usado = columnas[-1]["x"] + columnas[-1]["w"]
        columnas[-1]["w"] += (table_right - usado)

    def texto_ajustado(valor: str, max_width: float, font_name: str, font_size: int) -> str:
        texto = str(valor or "")
        if c.stringWidth(texto, font_name, font_size) <= max_width:
            return texto
        sufijo = "..."
        if c.stringWidth(sufijo, font_name, font_size) > max_width:
            return ""
        low, high = 0, len(texto)
        while low < high:
            mid = (low + high + 1) // 2
            candidato = texto[:mid] + sufijo
            if c.stringWidth(candidato, font_name, font_size) <= max_width:
                low = mid
            else:
                high = mid - 1
        return texto[:low] + sufijo

    def dibujar_cabecera_tabla(y_pos: float):
        c.setFillColor(colors.HexColor("#1d4ed8"))
        c.rect(table_left, y_pos - 2, table_width, 16, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        pad = 4
        for col in columnas:
            max_txt = max(col["w"] - (pad * 2), 8)
            titulo = texto_ajustado(col["titulo"], max_txt, "Helvetica-Bold", 8)
            if col["align"] == "right":
                c.drawRightString(col["x"] + col["w"] - pad, y_pos + 2, titulo)
            else:
                c.drawString(col["x"] + pad, y_pos + 2, titulo)
        c.setFillColor(colors.black)
        c.setStrokeColor(colors.HexColor("#cbd5e1"))
        c.line(table_left, y_pos - 3, table_right, y_pos - 3)

    dibujar_cabecera_tabla(y)
    y -= 16
    c.setFont("Helvetica", 8)

    if not filas:
        c.drawString(40, y, "Sin registros para los filtros aplicados.")
    else:
        for fila in filas:
            if y < 42:
                c.showPage()
                y = alto - 42
                dibujar_cabecera_tabla(y)
                y -= 16
                c.setFont("Helvetica", 8)
            valores = [
                fila.get("nro_orden") or "",
                fila.get("empleado") or "",
                fila.get("rol") or "",
                fila.get("proyecto") or "",
                fila.get("centro_costo") or "",
                fila.get("origen") or "",
                fila.get("destino") or "",
                format(float(fila.get("hh") or 0), ".2f"),
                format(float(fila.get("compensables") or 0), ".2f"),
                format(float(fila.get("viaticos") or 0), ".2f"),
            ]
            pad = 4
            for col, valor in zip(columnas, valores):
                max_txt = max(col["w"] - (pad * 2), 8)
                txt = texto_ajustado(valor, max_txt, "Helvetica", 8)
                if col["align"] == "right":
                    c.drawRightString(col["x"] + col["w"] - pad, y, txt)
                else:
                    c.drawString(col["x"] + pad, y, txt)
            y -= 12

    c.save()
    pdf_bytes = pdf_buffer.getvalue()
    pdf_buffer.close()
    nombre = f"Gestion_Operativa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@app.get("/gestion_operativa/excel")
def gestion_operativa_excel(
    desde: str = "",
    hasta: str = "",
    nro_orden: str = "",
    proyecto: str = "",
    centro_costo: str = "",
    empleado: str = "",
):
    with get_sqlite_connection() as conn:
        payload = _gestion_operativa_resumen_payload(
            conn,
            desde=desde,
            hasta=hasta,
            nro_orden=nro_orden,
            proyecto=proyecto,
            centro_costo=centro_costo,
            empleado=empleado,
        )

    filas = payload.get("tabla", []) if isinstance(payload, dict) else []

    try:
        from openpyxl import Workbook
    except Exception as exc:
        return {"error": f"No se pudo generar el XLSX: {exc}"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Operativo"
    headers = ["Nro Orden", "Empleado", "Rol", "Proyecto", "Centro de Costos", "Origen", "Destino", "HH", "Compensables", "Viaticos"]
    ws.append(headers)
    for fila in filas:
        ws.append([
            fila.get("nro_orden") or "",
            fila.get("empleado") or "",
            fila.get("rol") or "",
            fila.get("proyecto") or "",
            fila.get("centro_costo") or "",
            fila.get("origen") or "",
            fila.get("destino") or "",
            float(fila.get("hh") or 0),
            float(fila.get("compensables") or 0),
            float(fila.get("viaticos") or 0),
        ])

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(12, max_len + 2), 40)

    out = io.BytesIO()
    wb.save(out)
    contenido = out.getvalue()
    out.close()

    nombre = f"Gestion_Operativa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@app.post("/ordenes/{nro_orden}/cierre")
async def guardar_cierre_logistico(
    nro_orden: str,
    payload: str = Form("{}"),
    checklist_mantenimiento: UploadFile | None = File(default=None),
    formulario_logistica_viaje: UploadFile | None = File(default=None),
):
    with get_sqlite_connection() as conn:
        orden_row = conn.execute(
            "SELECT raw_json FROM ordenes_salida WHERE nro_orden = ?",
            (nro_orden,),
        ).fetchone()
        if orden_row is None:
            return {"error": f"No se encontró la orden {nro_orden}"}

        orden = parse_json_dict(orden_row["raw_json"], default={})
        if not orden:
            return {"error": f"No se encontró la orden {nro_orden}"}

    try:
        payload_data = json.loads(payload or "{}")
    except json.JSONDecodeError:
        return {"error": "Payload de cierre inválido"}

    solicitud = payload_data.get("solicitud", {}) if isinstance(payload_data, dict) else {}
    asignacion = payload_data.get("asignacion", {}) if isinstance(payload_data, dict) else {}

    viaje_data = orden.get("viaje", {}) if isinstance(orden.get("viaje", {}), dict) else {}
    recursos_data = orden.get("recursos", {}) if isinstance(orden.get("recursos", {}), dict) else {}

    if solicitud:
        viaje_data["fecha_salida"] = solicitud.get("fecha_salida", viaje_data.get("fecha_salida", ""))
        viaje_data["fecha_regreso"] = solicitud.get("fecha_llegada", viaje_data.get("fecha_regreso", ""))
        viaje_data["solicitante"] = solicitud.get("quien_solicita", viaje_data.get("solicitante", ""))
        viaje_data["area"] = solicitud.get("area_solicita", viaje_data.get("area", ""))
        viaje_data["origen"] = solicitud.get("partida", viaje_data.get("origen", ""))
        viaje_data["destino"] = solicitud.get("destino", viaje_data.get("destino", ""))
        viaje_data["motivo"] = solicitud.get("motivo", viaje_data.get("motivo", ""))

        recursos_data["fecha"] = solicitud.get("fecha_emision", recursos_data.get("fecha", ""))
        recursos_data["centro_costo"] = solicitud.get("centro_costos", recursos_data.get("centro_costo", ""))
        recursos_data["datos_solicitante"] = solicitud.get("quien_solicita", recursos_data.get("datos_solicitante", ""))
        recursos_data["area_solicitante"] = solicitud.get("area_solicita", recursos_data.get("area_solicitante", ""))
        recursos_data["partida"] = solicitud.get("partida", recursos_data.get("partida", ""))
        recursos_data["destino"] = solicitud.get("destino", recursos_data.get("destino", ""))
        recursos_data["motivo_viaje"] = solicitud.get("motivo", recursos_data.get("motivo_viaje", ""))
        recursos_data["fecha_salida_viaje"] = solicitud.get("fecha_salida", recursos_data.get("fecha_salida_viaje", ""))
        recursos_data["fecha_regreso_viaje"] = solicitud.get("fecha_llegada", recursos_data.get("fecha_regreso_viaje", ""))
        recursos_data["hora_salida"] = solicitud.get("hora_salida", recursos_data.get("hora_salida", ""))
        recursos_data["hora_regreso"] = solicitud.get("hora_regreso", recursos_data.get("hora_regreso", ""))
        recursos_data["duracion_jornadas"] = solicitud.get("duracion_jornadas", recursos_data.get("duracion_jornadas", ""))

    if asignacion:
        recursos_data["chofer"] = asignacion.get("chofer", recursos_data.get("chofer", ""))
        recursos_data["vehiculo"] = asignacion.get("vehiculo", recursos_data.get("vehiculo", ""))
        recursos_data["chofer_viatico"] = asignacion.get("viatico_chofer", recursos_data.get("chofer_viatico", 0))
        recursos_data["otros_gastos"] = asignacion.get("otros_gastos", recursos_data.get("otros_gastos", 0))
        recursos_data["viaticos"] = asignacion.get("viaticos_sin_otros", recursos_data.get("viaticos", 0))
        recursos_data["medio_pago"] = asignacion.get("medio_pago", recursos_data.get("medio_pago", "Caja"))
        recursos_data["alojamiento"] = asignacion.get("alojamiento", recursos_data.get("alojamiento", "NO"))
        recursos_data["acompanantes"] = asignacion.get("acompanantes", recursos_data.get("acompanantes", []))
        recursos_data["acompanantes_con_viatico"] = asignacion.get(
            "acompanantes_con_viatico", recursos_data.get("acompanantes_con_viatico", [])
        )
        recursos_data["comprobacion_operaciones_logistica"] = asignacion.get(
            "comprobaciones", recursos_data.get("comprobacion_operaciones_logistica", {})
        )
        viaje_data["chofer"] = recursos_data.get("chofer", viaje_data.get("chofer", ""))
        viaje_data["vehiculo"] = recursos_data.get("vehiculo", viaje_data.get("vehiculo", ""))

    viaje_data["recursos"] = recursos_data
    viaje_data["estado"] = "CERRADO"
    orden["viaje"] = viaje_data
    orden["recursos"] = recursos_data
    orden["estado"] = "CERRADO"

    archivo_checklist = await guardar_adjunto_logistico(
        checklist_mantenimiento,
        nro_orden,
        "Check List emitido por Mantenimiento del Equipo",
    )
    archivo_formulario = await guardar_adjunto_logistico(
        formulario_logistica_viaje,
        nro_orden,
        "Formulario Logistica de Viaje",
    )

    cierre = {
        "fecha_guardado": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "datos": payload_data,
        "cerrado": True,
        "adjuntos": {
            "checklist_mantenimiento": archivo_checklist,
            "formulario_logistica_viaje": archivo_formulario,
        },
    }

    orden["cierre_logistica"] = cierre

    with get_sqlite_connection() as conn:
        conn.execute(
            """
            UPDATE ordenes_salida
               SET estado = ?,
                   cierre_logistica_json = ?,
                   raw_json = ?
             WHERE nro_orden = ?
            """,
            (
                "CERRADO",
                json.dumps(cierre, ensure_ascii=False),
                json.dumps(orden, ensure_ascii=False),
                nro_orden,
            ),
        )

        id_viaje = orden.get("id_viaje")
        viaje_row = conn.execute("SELECT raw_json FROM viajes WHERE id = ?", (id_viaje,)).fetchone()
        if viaje_row is not None:
            viaje_db = parse_json_dict(viaje_row["raw_json"], default={})
            viaje_db["estado"] = "CERRADO"
            viaje_db["recursos"] = recursos_data
            viaje_db["cierre_logistica"] = cierre
            guardar_viaje_sql(conn, viaje_db)
            guardar_recursos_sql(conn, id_viaje, recursos_data)

        sincronizar_gestion_operativa_cierre(conn, nro_orden, orden, payload_data)

        conn.commit()

    return {
        "mensaje": "Cierre logístico guardado",
        "adjuntos": cierre["adjuntos"],
        "estado": "CERRADO",
    }


@app.get("/ordenes_view", response_class=HTMLResponse)
def ordenes_view(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(ORDENES_VIEW_PATH)

@app.get("/print_viaje", response_class=HTMLResponse)
def print_viaje(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(PRINT_VIAJE_PATH)

@app.get("/print_orden_salida", response_class=HTMLResponse)
def print_orden_salida(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(PRINT_ORDEN_SALIDA_PATH)

@app.get("/personal_form", response_class=HTMLResponse)
def personal_form(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(PERSONAL_FORM_PATH)
    
    

    
@app.get("/vehiculos")
def obtener_vehiculos():
    return obtener_vehiculos_data()

@app.get("/choferes")
def obtener_choferes():
    return obtener_choferes_data()

@app.post("/vehiculos")
def crear_vehiculo(data: dict):
    codigo = str(data.get("codigo", "")).strip()
    if not codigo:
        return {"error": "Debe indicar un codigo de vehiculo"}

    with get_sqlite_connection() as conn:
        existe = conn.execute("SELECT 1 FROM vehiculos WHERE codigo = ?", (codigo,)).fetchone()
        if existe:
            return {"error": "El codigo de vehiculo ya existe"}

        conn.execute(
            """
            INSERT INTO vehiculos (
                codigo, propiedad, marca, tipo, modelo, dominio, anio, motor, chasis,
                sector, proyecto, operativo,
                habilitacion_pirquitas, habilitacion_exar, habilitacion_sdj,
                habilitacion_rincon, habilitacion_arli, raw_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                codigo,
                data.get("propiedad"),
                data.get("marca"),
                data.get("tipo"),
                data.get("modelo"),
                data.get("dominio"),
                data.get("anio"),
                data.get("motor"),
                data.get("chasis"),
                data.get("sector"),
                data.get("proyecto"),
                data.get("operativo"),
                data.get("habilitacion_pirquitas"),
                data.get("habilitacion_exar"),
                data.get("habilitacion_sdj"),
                data.get("habilitacion_rincon"),
                data.get("habilitacion_arli"),
                json.dumps(data, ensure_ascii=False),
            ),
        )
        conn.commit()

    return {"mensaje": "Vehículo creado"}

@app.delete("/vehiculos/{codigo}")
def eliminar_vehiculo(codigo: str):
    with get_sqlite_connection() as conn:
        conn.execute("DELETE FROM vehiculos WHERE codigo = ?", (codigo,))
        conn.commit()

    return {"mensaje": "Vehículo eliminado"}

@app.get("/vehiculos_form", response_class=HTMLResponse)
def vehiculos_form(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    with open("vehiculos.html", "r", encoding="utf-8") as f:
        return f.read()

@app.put("/vehiculos/{codigo}")
def actualizar_vehiculo(codigo: str, data: dict):
    with get_sqlite_connection() as conn:
        row = conn.execute("SELECT raw_json FROM vehiculos WHERE codigo = ?", (codigo,)).fetchone()
        if row is None:
            return {"error": "Vehículo no encontrado"}

        base_data = {}
        try:
            base_data = json.loads(row["raw_json"] or "{}")
        except Exception:
            base_data = {}

        base_data.update(data)

        conn.execute(
            """
            UPDATE vehiculos
               SET propiedad = ?,
                   marca = ?,
                   tipo = ?,
                   modelo = ?,
                   dominio = ?,
                   anio = ?,
                   motor = ?,
                   chasis = ?,
                   sector = ?,
                   proyecto = ?,
                   operativo = ?,
                   habilitacion_pirquitas = ?,
                   habilitacion_exar = ?,
                   habilitacion_sdj = ?,
                   habilitacion_rincon = ?,
                   habilitacion_arli = ?,
                   raw_json = ?
             WHERE codigo = ?
            """,
            (
                base_data.get("propiedad"),
                base_data.get("marca"),
                base_data.get("tipo"),
                base_data.get("modelo"),
                base_data.get("dominio"),
                base_data.get("anio"),
                base_data.get("motor"),
                base_data.get("chasis"),
                base_data.get("sector"),
                base_data.get("proyecto"),
                base_data.get("operativo"),
                base_data.get("habilitacion_pirquitas"),
                base_data.get("habilitacion_exar"),
                base_data.get("habilitacion_sdj"),
                base_data.get("habilitacion_rincon"),
                base_data.get("habilitacion_arli"),
                json.dumps(base_data, ensure_ascii=False),
                codigo,
            ),
        )
        conn.commit()

    return {"mensaje": "Vehículo actualizado"}


# -------- PERSONAL --------


@app.get("/personal")
def obtener_personal():
    return obtener_personal_data()


@app.post("/personal")
def crear_personal(data: dict):
    legajo = str(data.get("legajo", "")).strip()
    if not legajo:
        return {"error": "El legajo es obligatorio"}

    with get_sqlite_connection() as conn:
        existe = conn.execute("SELECT 1 FROM personal WHERE legajo = ?", (legajo,)).fetchone()
        if existe:
            return {"error": "El legajo ya existe"}

        conn.execute(
            """
            INSERT INTO personal (
                legajo, nombre, cuil, fecha_nacimiento, activo,
                habilitacion_pirquitas, habilitacion_exar, habilitacion_sdj,
                habilitacion_rincon, habilitacion_arli, raw_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                legajo,
                data.get("nombre"),
                data.get("cuil"),
                data.get("fecha_nacimiento"),
                data.get("activo", "ACTIVO"),
                data.get("habilitacion_pirquitas"),
                data.get("habilitacion_exar"),
                data.get("habilitacion_sdj"),
                data.get("habilitacion_rincon"),
                data.get("habilitacion_arli"),
                json.dumps(data, ensure_ascii=False),
            ),
        )
        conn.commit()

    return {"mensaje": "Empleado agregado"}


@app.put("/personal/{legajo}")
def actualizar_personal(legajo: str, data: dict):
    with get_sqlite_connection() as conn:
        row = conn.execute("SELECT raw_json FROM personal WHERE legajo = ?", (str(legajo),)).fetchone()
        if row is None:
            return {"error": "Empleado no encontrado"}

        base_data = {}
        try:
            base_data = json.loads(row["raw_json"] or "{}")
        except Exception:
            base_data = {}

        base_data.update(data)

        conn.execute(
            """
            UPDATE personal
               SET nombre = ?,
                   cuil = ?,
                   fecha_nacimiento = ?,
                   activo = ?,
                   habilitacion_pirquitas = ?,
                   habilitacion_exar = ?,
                   habilitacion_sdj = ?,
                   habilitacion_rincon = ?,
                   habilitacion_arli = ?,
                   raw_json = ?
             WHERE legajo = ?
            """,
            (
                base_data.get("nombre"),
                base_data.get("cuil"),
                base_data.get("fecha_nacimiento"),
                base_data.get("activo", "ACTIVO"),
                base_data.get("habilitacion_pirquitas"),
                base_data.get("habilitacion_exar"),
                base_data.get("habilitacion_sdj"),
                base_data.get("habilitacion_rincon"),
                base_data.get("habilitacion_arli"),
                json.dumps(base_data, ensure_ascii=False),
                str(legajo),
            ),
        )
        conn.commit()

    return {"mensaje": "Empleado actualizado"}


@app.get("/personal/proyectos")
def listar_proyectos_habilitacion():
    with get_sqlite_connection() as conn:
        return obtener_proyectos_habilitacion(conn)


@app.get("/personal/{legajo}/habilitaciones")
def listar_habilitaciones_personal(legajo: str):
    with get_sqlite_connection() as conn:
        existe = conn.execute("SELECT 1 FROM personal WHERE legajo = ?", (str(legajo),)).fetchone()
        if not existe:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
        habilitaciones = obtener_habilitaciones_por_legajo(conn, legajo)
    return habilitaciones.get(str(legajo), {})


class PersonalHabilitacionPayload(BaseModel):
    proyecto_id: int
    fecha_vencimiento: str


@app.post("/personal/{legajo}/habilitaciones")
def guardar_habilitacion_personal(legajo: str, payload: PersonalHabilitacionPayload):
    fecha = (payload.fecha_vencimiento or "").strip()
    if not fecha:
        return {"error": "Debe indicar una fecha de vencimiento"}
    with get_sqlite_connection() as conn:
        existe = conn.execute("SELECT 1 FROM personal WHERE legajo = ?", (str(legajo),)).fetchone()
        if not existe:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
        proyecto = conn.execute(
            "SELECT id FROM proyectos_habilitacion WHERE id = ?", (payload.proyecto_id,)
        ).fetchone()
        if not proyecto:
            raise HTTPException(status_code=404, detail="Proyecto no encontrado")
        conn.execute(
            """
            INSERT INTO personal_habilitaciones (legajo, proyecto_id, fecha_vencimiento, updated_at)
            VALUES (?, ?, ?, datetime('now'))
            ON CONFLICT(legajo, proyecto_id) DO UPDATE SET
                fecha_vencimiento = excluded.fecha_vencimiento,
                updated_at = datetime('now')
            """,
            (str(legajo), payload.proyecto_id, fecha),
        )
        conn.commit()
    return {"mensaje": "Habilitación guardada"}


@app.delete("/personal/{legajo}/habilitaciones/{proyecto_id}")
def quitar_habilitacion_personal(legajo: str, proyecto_id: int):
    with get_sqlite_connection() as conn:
        conn.execute(
            "DELETE FROM personal_habilitaciones WHERE legajo = ? AND proyecto_id = ?",
            (str(legajo), proyecto_id),
        )
        conn.commit()
    return {"mensaje": "Habilitación quitada"}


# -------- ALMACEN --------


class AlmacenSimplePayload(BaseModel):
    nombre: str
    activo: int = 1


class AlmacenUbicacionPayload(BaseModel):
    nombre: str
    parent_id: int | None = None
    activo: int = 1


class AlmacenProductoPayload(BaseModel):
    codigo: str
    descripcion: str
    marca: str = ""
    modelo: str = ""
    categoria_id: int | None = None
    tipo_producto_id: int | None = None
    unidad_medida_id: int | None = None
    stock_ingreso: float = 0
    punto_reposicion: float = 0
    ubicacion_id: int | None = None
    observaciones: str = ""


class AlmacenRemitoItemPayload(BaseModel):
    producto_id: int
    cantidad: float
    series: list[str] = []
    vehiculo_codigo: str = ""
    proyecto: str = ""
    instalacion: str = ""


class AlmacenRIPayload(BaseModel):
    proveedor: str = ""
    nro_remito_referencia: str = ""
    responsable_legajo: str | None = None
    fecha: str | None = None
    observaciones: str = ""
    items: list[AlmacenRemitoItemPayload]


class AlmacenREPayload(BaseModel):
    destinatario: str = ""
    razon_social: str = ""
    cuit_dni: str = ""
    direccion: str = ""
    localidad: str = ""
    provincia: str = ""
    telefono: str = ""
    codigo_postal: str = ""
    transporte: str = ""
    dominio: str = ""
    entrega_legajo: str | None = None
    recibe_legajo: str | None = None
    entrega_nombre: str = ""
    recibe_nombre: str = ""
    observaciones: str = ""
    fecha: str | None = None
    items: list[AlmacenRemitoItemPayload]


class AlmacenAutorizacionREPayload(BaseModel):
    autorizado_por: str = ""
    observaciones: str = ""


class AlmacenInventarioPayload(BaseModel):
    responsable_legajo: str | None = None
    observaciones: str = ""


class AlmacenConteoPayload(BaseModel):
    stock_fisico: float
    observaciones: str = ""


class AlmacenRolPayload(BaseModel):
    nombre: str
    permisos_json: str = "{}"


class AlmacenAsignacionRolPayload(BaseModel):
    legajo: str
    rol_id: int


def obtener_siguiente_numero(conn, tabla: str, prefijo: str):
    row = conn.execute(f"SELECT numero FROM {tabla}").fetchall()
    max_num = 0
    for r in row:
        nro = str(r["numero"] or "")
        if nro.startswith(prefijo):
            try:
                val = int(nro.split("-")[1])
                max_num = max(max_num, val)
            except Exception:
                continue
    return f"{prefijo}-{max_num + 1:06d}"


def ajustar_stock_producto(conn, producto_id: int, delta: float):
    producto = conn.execute(
        """
         SELECT p.id, p.codigo, p.descripcion,
             f.id AS id_familia,
             m.id AS id_marca,
             md.id AS id_modelo,
             u.id AS id_unidad,
             ub.id AS id_ubicacion,
               COALESCE(f.nombre, '') AS familia_nombre,
               COALESCE(m.nombre, '') AS marca_nombre,
               COALESCE(md.nombre, '') AS modelo_nombre,
               COALESCE(u.nombre, '') AS unidad_nombre
        FROM productos p
        LEFT JOIN familias f ON f.id = p.id_familia
        LEFT JOIN marcas m ON m.id = p.id_marca
        LEFT JOIN modelos md ON md.id = p.id_modelo
        LEFT JOIN unidades_medida u ON u.id = COALESCE(p.id_unidad, p.unidad_medida_id)
         LEFT JOIN ubicaciones ub ON ub.id = p.ubicacion_id
        WHERE p.id = ?
        """,
        (producto_id,),
    ).fetchone()
    if producto is None:
        raise ValueError(f"Producto inexistente: {producto_id}")

    actual_row = conn.execute(
        "SELECT stock_actual FROM stock WHERE producto_id = ?",
        (producto_id,),
    ).fetchone()
    actual = float(actual_row["stock_actual"] or 0) if actual_row else 0.0
    nuevo = actual + float(delta)
    if nuevo < 0:
        raise ValueError(f"Stock insuficiente para producto {producto_id}")

    payload_json = json.dumps(
        {
            "producto_id": producto_id,
            "id_familia": producto["id_familia"],
            "id_marca": producto["id_marca"],
            "id_modelo": producto["id_modelo"],
            "id_unidad": producto["id_unidad"],
            "id_ubicacion": producto["id_ubicacion"],
            "producto_codigo": producto["codigo"] or "",
            "producto_descripcion": producto["descripcion"] or "",
            "familia_nombre": producto["familia_nombre"] or "",
            "marca_nombre": producto["marca_nombre"] or "",
            "modelo_nombre": producto["modelo_nombre"] or "",
            "unidad_nombre": producto["unidad_nombre"] or "",
            "stock_actual": nuevo,
        },
        ensure_ascii=False,
    )

    if actual_row is None:
        conn.execute(
            """
            INSERT INTO stock (
                producto_id, id_familia, id_marca, id_modelo, id_unidad, id_ubicacion,
                producto_codigo, producto_descripcion,
                familia_nombre, marca_nombre, modelo_nombre, unidad_nombre,
                stock_actual, updated_at, raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                producto_id,
                producto["id_familia"],
                producto["id_marca"],
                producto["id_modelo"],
                producto["id_unidad"],
                producto["id_ubicacion"],
                producto["codigo"] or "",
                producto["descripcion"] or "",
                producto["familia_nombre"] or "",
                producto["marca_nombre"] or "",
                producto["modelo_nombre"] or "",
                producto["unidad_nombre"] or "",
                nuevo,
                fecha_actual_iso(),
                payload_json,
            ),
        )
    else:
        conn.execute(
            """
            UPDATE stock
               SET id_familia = ?,
                   id_marca = ?,
                   id_modelo = ?,
                   id_unidad = ?,
                   id_ubicacion = ?,
                   producto_codigo = ?,
                   producto_descripcion = ?,
                   familia_nombre = ?,
                   marca_nombre = ?,
                   modelo_nombre = ?,
                   unidad_nombre = ?,
                   stock_actual = ?,
                   updated_at = ?,
                   raw_json = ?
             WHERE producto_id = ?
            """,
            (
                producto["id_familia"],
                producto["id_marca"],
                producto["id_modelo"],
                producto["id_unidad"],
                producto["id_ubicacion"],
                producto["codigo"] or "",
                producto["descripcion"] or "",
                producto["familia_nombre"] or "",
                producto["marca_nombre"] or "",
                producto["modelo_nombre"] or "",
                producto["unidad_nombre"] or "",
                nuevo,
                fecha_actual_iso(),
                payload_json,
                producto_id,
            ),
        )


def asegurar_tabla_stock(conn):
    tablas = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    if "stock" not in tablas:
        conn.execute(
            """
            CREATE TABLE stock (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                producto_id INTEGER NOT NULL UNIQUE,
                id_familia INTEGER,
                id_marca INTEGER,
                id_modelo INTEGER,
                id_unidad INTEGER,
                id_ubicacion INTEGER,
                producto_codigo TEXT,
                producto_descripcion TEXT,
                familia_nombre TEXT,
                marca_nombre TEXT,
                modelo_nombre TEXT,
                unidad_nombre TEXT,
                stock_actual REAL NOT NULL DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT,
                raw_json TEXT,
                FOREIGN KEY (producto_id) REFERENCES productos (id) ON DELETE RESTRICT,
                FOREIGN KEY (id_familia) REFERENCES familias (id) ON DELETE SET NULL,
                FOREIGN KEY (id_marca) REFERENCES marcas (id) ON DELETE SET NULL,
                FOREIGN KEY (id_modelo) REFERENCES modelos (id) ON DELETE SET NULL,
                FOREIGN KEY (id_unidad) REFERENCES unidades_medida (id) ON DELETE SET NULL,
                FOREIGN KEY (id_ubicacion) REFERENCES ubicaciones (id) ON DELETE SET NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_stock_producto ON stock (producto_id)")
        return

    columnas_stock = {r[1] for r in conn.execute("PRAGMA table_info(stock)").fetchall()}
    if "stock_actual" not in columnas_stock:
        conn.execute("DROP TABLE IF EXISTS stock_new")
        conn.execute(
            """
            CREATE TABLE stock_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                producto_id INTEGER NOT NULL UNIQUE,
                id_familia INTEGER,
                id_marca INTEGER,
                id_modelo INTEGER,
                id_unidad INTEGER,
                id_ubicacion INTEGER,
                producto_codigo TEXT,
                producto_descripcion TEXT,
                familia_nombre TEXT,
                marca_nombre TEXT,
                modelo_nombre TEXT,
                unidad_nombre TEXT,
                stock_actual REAL NOT NULL DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT,
                raw_json TEXT,
                FOREIGN KEY (producto_id) REFERENCES productos (id) ON DELETE RESTRICT,
                FOREIGN KEY (id_familia) REFERENCES familias (id) ON DELETE SET NULL,
                FOREIGN KEY (id_marca) REFERENCES marcas (id) ON DELETE SET NULL,
                FOREIGN KEY (id_modelo) REFERENCES modelos (id) ON DELETE SET NULL,
                FOREIGN KEY (id_unidad) REFERENCES unidades_medida (id) ON DELETE SET NULL,
                FOREIGN KEY (id_ubicacion) REFERENCES ubicaciones (id) ON DELETE SET NULL
            )
            """
        )

        filas = conn.execute(
            """
            SELECT p.id AS producto_id,
                     f.id AS id_familia,
                     m.id AS id_marca,
                     md.id AS id_modelo,
                     u.id AS id_unidad,
                     ub.id AS id_ubicacion,
                   p.codigo AS producto_codigo,
                   p.descripcion AS producto_descripcion,
                   COALESCE(f.nombre, '') AS familia_nombre,
                   COALESCE(m.nombre, '') AS marca_nombre,
                   COALESCE(md.nombre, '') AS modelo_nombre,
                   COALESCE(u.nombre, '') AS unidad_nombre,
                   COALESCE(ri.total_ingresado, 0) - COALESCE(re.total_salida, 0) AS stock_total
            FROM productos p
            LEFT JOIN (
                SELECT producto_id, SUM(cantidad) AS total_ingresado
                FROM remitos_ingreso_detalle
                GROUP BY producto_id
            ) ri ON ri.producto_id = p.id
            LEFT JOIN (
                SELECT d.producto_id, SUM(d.cantidad) AS total_salida
                FROM remitos_entrega_detalle d
                JOIN remitos_entrega r ON r.id = d.remito_id
                WHERE UPPER(COALESCE(r.estado_autorizacion, 'AUTORIZADO')) = 'AUTORIZADO'
                GROUP BY d.producto_id
            ) re ON re.producto_id = p.id
            LEFT JOIN familias f ON f.id = p.id_familia
            LEFT JOIN marcas m ON m.id = p.id_marca
            LEFT JOIN modelos md ON md.id = p.id_modelo
            LEFT JOIN unidades_medida u ON u.id = COALESCE(p.id_unidad, p.unidad_medida_id)
            LEFT JOIN ubicaciones ub ON ub.id = p.ubicacion_id
            WHERE COALESCE(ri.total_ingresado, 0) <> 0 OR COALESCE(re.total_salida, 0) <> 0
            """
        ).fetchall()

        for fila in filas:
            stock_total = float(fila["stock_total"] or 0)
            conn.execute(
                """
                INSERT INTO stock_new (
                    producto_id, id_familia, id_marca, id_modelo, id_unidad, id_ubicacion,
                    producto_codigo, producto_descripcion,
                    familia_nombre, marca_nombre, modelo_nombre, unidad_nombre,
                    stock_actual, updated_at, raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    fila["producto_id"],
                    fila["id_familia"],
                    fila["id_marca"],
                    fila["id_modelo"],
                    fila["id_unidad"],
                    fila["id_ubicacion"],
                    fila["producto_codigo"] or "",
                    fila["producto_descripcion"] or "",
                    fila["familia_nombre"] or "",
                    fila["marca_nombre"] or "",
                    fila["modelo_nombre"] or "",
                    fila["unidad_nombre"] or "",
                    stock_total,
                    fecha_actual_iso(),
                    json.dumps(
                        {
                            "producto_id": fila["producto_id"],
                            "stock_actual": stock_total,
                        },
                        ensure_ascii=False,
                    ),
                ),
            )

        conn.execute("DROP TABLE stock")
        conn.execute("ALTER TABLE stock_new RENAME TO stock")

    columnas_stock = {r[1] for r in conn.execute("PRAGMA table_info(stock)").fetchall()}
    columnas_extra = {
        "id_familia": "INTEGER",
        "id_marca": "INTEGER",
        "id_modelo": "INTEGER",
        "id_unidad": "INTEGER",
        "id_ubicacion": "INTEGER",
        "familia_nombre": "TEXT",
        "marca_nombre": "TEXT",
        "modelo_nombre": "TEXT",
        "unidad_nombre": "TEXT",
        "updated_at": "TEXT",
    }
    for nombre_col, tipo_col in columnas_extra.items():
        if nombre_col not in columnas_stock:
            conn.execute(f"ALTER TABLE stock ADD COLUMN {nombre_col} {tipo_col}")

    conn.execute(
        """
        UPDATE stock
           SET id_familia = (
                   SELECT f.id
                   FROM productos p
                   LEFT JOIN familias f ON f.id = p.id_familia
                   WHERE p.id = stock.producto_id
               ),
               id_marca = (
                   SELECT m.id
                   FROM productos p
                   LEFT JOIN marcas m ON m.id = p.id_marca
                   WHERE p.id = stock.producto_id
               ),
               id_modelo = (
                   SELECT md.id
                   FROM productos p
                   LEFT JOIN modelos md ON md.id = p.id_modelo
                   WHERE p.id = stock.producto_id
               ),
               id_unidad = (
                   SELECT u.id
                   FROM productos p
                   LEFT JOIN unidades_medida u ON u.id = COALESCE(p.id_unidad, p.unidad_medida_id)
                   WHERE p.id = stock.producto_id
               ),
               id_ubicacion = (
                   SELECT ub.id
                   FROM productos p
                   LEFT JOIN ubicaciones ub ON ub.id = p.ubicacion_id
                   WHERE p.id = stock.producto_id
               ),
               familia_nombre = COALESCE((
                   SELECT f.nombre
                   FROM productos p
                   LEFT JOIN familias f ON f.id = p.id_familia
                   WHERE p.id = stock.producto_id
               ), ''),
               marca_nombre = COALESCE((
                   SELECT m.nombre
                   FROM productos p
                   LEFT JOIN marcas m ON m.id = p.id_marca
                   WHERE p.id = stock.producto_id
               ), ''),
               modelo_nombre = COALESCE((
                   SELECT md.nombre
                   FROM productos p
                   LEFT JOIN modelos md ON md.id = p.id_modelo
                   WHERE p.id = stock.producto_id
               ), ''),
               unidad_nombre = COALESCE((
                   SELECT u.nombre
                   FROM productos p
                   LEFT JOIN unidades_medida u ON u.id = COALESCE(p.id_unidad, p.unidad_medida_id)
                   WHERE p.id = stock.producto_id
               ), '')
            WHERE id_familia IS NULL
                OR id_marca IS NULL
                OR id_modelo IS NULL
                OR id_unidad IS NULL
                OR id_ubicacion IS NULL
                OR COALESCE(familia_nombre, '') = ''
            OR COALESCE(marca_nombre, '') = ''
            OR COALESCE(modelo_nombre, '') = ''
            OR COALESCE(unidad_nombre, '') = ''
        """
    )

    conn.execute("CREATE INDEX IF NOT EXISTS idx_stock_producto ON stock (producto_id)")


def obtener_stock_actual(conn, producto_id: int):
    # Garantiza migracion de esquema/datos aun si existe una tabla stock de formato viejo.
    asegurar_tabla_stock(conn)
    migrar_tabla_stock(conn)
    row = conn.execute(
        "SELECT COALESCE(stock_actual, 0) AS stock_actual FROM stock WHERE producto_id = ?",
        (producto_id,),
    ).fetchone()
    return float(row["stock_actual"] or 0) if row else 0.0


def obtener_stock_reservado(conn, producto_id: int, excluir_remito_id: int | None = None):
    query = """
        SELECT COALESCE(SUM(d.cantidad), 0) AS reservado
        FROM remitos_entrega_detalle d
        JOIN remitos_entrega r ON r.id = d.remito_id
        WHERE d.producto_id = ?
          AND UPPER(COALESCE(r.estado_autorizacion, 'AUTORIZADO')) = 'PENDIENTE'
    """
    params: list = [producto_id]
    if excluir_remito_id is not None:
        query += " AND r.id <> ?"
        params.append(excluir_remito_id)
    row = conn.execute(query, tuple(params)).fetchone()
    return float(row["reservado"] or 0) if row else 0.0


def obtener_stock_disponible(conn, producto_id: int, excluir_remito_id: int | None = None):
    stock_actual = obtener_stock_actual(conn, producto_id)
    reservado = obtener_stock_reservado(conn, producto_id, excluir_remito_id=excluir_remito_id)
    disponible = stock_actual - reservado
    return round(disponible, 2)


def descontar_stock(conn, producto_id: int, cantidad: float):
    ajustar_stock_producto(conn, producto_id, -float(cantidad))


def migrar_tabla_stock(conn):
    asegurar_tabla_stock(conn)
    total = conn.execute("SELECT COUNT(*) AS total FROM stock").fetchone()["total"]
    if total:
        return

    filas = conn.execute(
        """
        SELECT p.id AS producto_id,
             f.id AS id_familia,
             m.id AS id_marca,
             md.id AS id_modelo,
             u.id AS id_unidad,
             ub.id AS id_ubicacion,
               p.codigo AS producto_codigo,
               p.descripcion AS producto_descripcion,
               COALESCE(f.nombre, '') AS familia_nombre,
               COALESCE(m.nombre, '') AS marca_nombre,
               COALESCE(md.nombre, '') AS modelo_nombre,
               COALESCE(u.nombre, '') AS unidad_nombre,
               COALESCE(ri.total_ingresado, 0) - COALESCE(re.total_salida, 0) AS stock_total
        FROM productos p
        LEFT JOIN (
            SELECT producto_id, SUM(cantidad) AS total_ingresado
            FROM remitos_ingreso_detalle
            GROUP BY producto_id
        ) ri ON ri.producto_id = p.id
        LEFT JOIN (
            SELECT d.producto_id, SUM(d.cantidad) AS total_salida
            FROM remitos_entrega_detalle d
            JOIN remitos_entrega r ON r.id = d.remito_id
            WHERE UPPER(COALESCE(r.estado_autorizacion, 'AUTORIZADO')) = 'AUTORIZADO'
            GROUP BY d.producto_id
        ) re ON re.producto_id = p.id
        LEFT JOIN familias f ON f.id = p.id_familia
        LEFT JOIN marcas m ON m.id = p.id_marca
        LEFT JOIN modelos md ON md.id = p.id_modelo
        LEFT JOIN unidades_medida u ON u.id = COALESCE(p.id_unidad, p.unidad_medida_id)
        LEFT JOIN ubicaciones ub ON ub.id = p.ubicacion_id
        WHERE COALESCE(ri.total_ingresado, 0) <> 0 OR COALESCE(re.total_salida, 0) <> 0
        """
    ).fetchall()

    for fila in filas:
        stock_total = float(fila["stock_total"] or 0)
        conn.execute(
            """
            INSERT INTO stock (
                producto_id, id_familia, id_marca, id_modelo, id_unidad, id_ubicacion,
                producto_codigo, producto_descripcion,
                familia_nombre, marca_nombre, modelo_nombre, unidad_nombre,
                stock_actual, updated_at, raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                fila["producto_id"],
                fila["id_familia"],
                fila["id_marca"],
                fila["id_modelo"],
                fila["id_unidad"],
                fila["id_ubicacion"],
                fila["producto_codigo"] or "",
                fila["producto_descripcion"] or "",
                fila["familia_nombre"] or "",
                fila["marca_nombre"] or "",
                fila["modelo_nombre"] or "",
                fila["unidad_nombre"] or "",
                stock_total,
                fecha_actual_iso(),
                json.dumps(
                    {
                        "producto_id": fila["producto_id"],
                        "stock_actual": stock_total,
                    },
                    ensure_ascii=False,
                ),
            ),
        )


def obtener_persona_por_legajo(conn, legajo: str | None):
    if not legajo:
        return None
    return conn.execute("SELECT legajo, nombre FROM personal WHERE legajo = ?", (str(legajo),)).fetchone()


def registrar_movimiento_stock(
    conn,
    *,
    fecha: str,
    producto_id: int,
    documento: str,
    tipo: str,
    cantidad: float,
    responsable_legajo: str | None,
    responsable_nombre: str,
    observaciones: str,
    remito_id: int | None = None,
    vehiculo_codigo: str = "",
    proyecto: str = "",
    instalacion: str = "",
):
    stock_anterior = obtener_stock_actual(conn, producto_id)
    if tipo.upper() == "SALIDA":
        stock_nuevo = stock_anterior - float(cantidad)
    else:
        stock_nuevo = stock_anterior + float(cantidad)

    conn.execute(
        """
        INSERT INTO movimientos_stock (
            fecha, producto_id, documento, tipo, cantidad,
            stock_anterior, stock_nuevo,
            responsable_legajo, responsable_nombre, observaciones,
            remito_id, vehiculo_codigo, proyecto, instalacion, raw_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            fecha,
            producto_id,
            documento,
            tipo,
            float(cantidad),
            stock_anterior,
            stock_nuevo,
            responsable_legajo,
            responsable_nombre,
            observaciones,
            remito_id,
            vehiculo_codigo,
            proyecto,
            instalacion,
            json.dumps(
                {
                    "fecha": fecha,
                    "producto_id": producto_id,
                    "documento": documento,
                    "tipo": tipo,
                    "cantidad": cantidad,
                    "stock_anterior": stock_anterior,
                    "stock_nuevo": stock_nuevo,
                },
                ensure_ascii=False,
            ),
        ),
    )
    return stock_anterior, stock_nuevo


def generar_pdf_remito(
    path_pdf: str,
    titulo: str,
    cabecera: list[tuple[str, str]],
    detalle: list[tuple[str, ...]],
    columnas: list[tuple[str, int]] | None = None,
    incluir_firmas: bool = False,
    firmas_layout: str = "simple",
    firma_entrega_nombre: str = "",
    firma_recibe_nombre: str = "",
    bloque_envio: list[tuple[str, str]] | None = None,
    bloque_envio_titulo: str = "Datos de envio",
    bloque_destinatarios_titulo: str = "",
    bloque_pie_izq: list[tuple[str, str]] | None = None,
    bloque_pie_der: list[tuple[str, str]] | None = None,
    observaciones_pie: str = "",
    departamento: str = "Departamento de Almacen",
    usuario_pie: str = "",
    columnas_multilinea: set[int] | None = None,
):
    columnas = columnas or [("Producto", 40), ("Cantidad", 250), ("Series", 330), ("Proyecto", 470)]
    columnas_multilinea = columnas_multilinea or set()
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
    except Exception:
        with open(path_pdf, "w", encoding="utf-8") as f:
            f.write(f"{titulo}\n")
            f.write("\nMEMBRETE EMPRESA\n")
            f.write(f"Razon Social: {EMPRESA_MEMBRETE['razon_social']}\n")
            f.write(f"Telefono: {EMPRESA_MEMBRETE['telefono']}\n")
            f.write(f"Cod.Postal: {EMPRESA_MEMBRETE['codigo_postal']}\n")
            f.write(f"Direccion: {EMPRESA_MEMBRETE['direccion']}\n")
            f.write(f"CUIT: {EMPRESA_MEMBRETE['cuit']}\n")
            for k, v in cabecera:
                f.write(f"{k}: {v}\n")
            if bloque_envio:
                f.write(f"\n{(bloque_envio_titulo or 'Datos').upper()}\n")
                for k, v in bloque_envio:
                    f.write(f"{k}: {v}\n")
            f.write("\nDETALLE\n")
            f.write(" | ".join(titulo_columna for titulo_columna, _ in columnas) + "\n")
            for fila in detalle:
                f.write(" | ".join(str(valor) for valor in fila) + "\n")
            if bloque_pie_izq or bloque_pie_der:
                f.write("\nDESTINATARIOS\n")
                for k, v in (bloque_pie_izq or []):
                    f.write(f"{k}: {v}\n")
                for k, v in (bloque_pie_der or []):
                    f.write(f"{k}: {v}\n")
            if incluir_firmas:
                if bloque_destinatarios_titulo:
                    f.write(f"\n{bloque_destinatarios_titulo.upper()}\n")
                if firmas_layout == "triple_entrega_recepcion":
                    f.write("\n\nENTREGA: _______________________________  Firma y Aclaracion: ____________________\n")
                    f.write("RECIBI CONFORME: _______________________  Firma y Aclaracion: ____________________\n")
                    f.write("RECIBI CONFORME: _______________________  Firma y Aclaracion: ____________________\n")
                else:
                    f.write("\n\nFirma de entrega: __________________________\n")
                    f.write("Firma de recepcion: ________________________\n")
            if usuario_pie:
                f.write(f"\n{usuario_pie}\n")
        return

    c = canvas.Canvas(path_pdf, pagesize=A4)
    ancho, alto = A4
    y = alto - 40

    def dibujar_usuario_pie():
        if usuario_pie:
            c.setFont("Helvetica-Oblique", 7)
            c.drawString(60, 18, usuario_pie)

    # Membrete fijo para todos los remitos
    if os.path.exists(MEMBRETE_LOGO_PATH):
        try:
            logo = ImageReader(MEMBRETE_LOGO_PATH)
            c.drawImage(logo, ancho - 185, alto - 92, width=140, height=45, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    c.setFont("Helvetica-Bold", 15)
    c.drawString(190, y, departamento)
    y -= 24
    c.setFont("Helvetica-Bold", 13)
    c.drawString(190, y, titulo)
    y -= 20

    c.setFont("Helvetica-Bold", 9)
    c.drawString(40, y, "Razon Social :")
    c.drawString(40, y - 13, "Telefono :")
    c.drawString(40, y - 26, "Cod.Postal :")
    c.drawString(40, y - 39, "Direccion :")
    c.drawString(40, y - 52, "CUIT :")
    c.setFont("Helvetica", 9)
    c.drawString(120, y, EMPRESA_MEMBRETE["razon_social"])
    c.drawString(120, y - 13, EMPRESA_MEMBRETE["telefono"])
    c.drawString(120, y - 26, EMPRESA_MEMBRETE["codigo_postal"])
    c.drawString(120, y - 39, EMPRESA_MEMBRETE["direccion"])
    c.drawString(120, y - 52, EMPRESA_MEMBRETE["cuit"])
    y -= 72

    c.setFont("Helvetica", 10)
    for k, v in cabecera:
        c.drawString(40, y, f"{k}: {v}")
        y -= 14

    if bloque_envio:
        y -= 2
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, bloque_envio_titulo or "Datos")
        y -= 14
        c.setFont("Helvetica", 9)
        x_izq, x_der = 40, 320
        y_base_envio = y
        filas_izq, filas_der = [], []
        for k, v in bloque_envio:
            etiqueta = (k or "").strip().lower()
            if etiqueta in {"telefono", "codigo postal", "transporte", "dominio", "sucursal destino"}:
                filas_der.append((k, v))
            else:
                filas_izq.append((k, v))
        for idx, (k, v) in enumerate(filas_izq):
            c.drawString(x_izq, y_base_envio - (idx * 12), f"{k}: {v}")
        for idx, (k, v) in enumerate(filas_der):
            c.drawString(x_der, y_base_envio - (idx * 12), f"{k}: {v}")
        y = y_base_envio - (max(len(filas_izq), len(filas_der)) * 12) - 10

    y -= 6
    c.setFont("Helvetica-Bold", 10)
    for titulo_columna, posicion in columnas:
        c.drawString(posicion, y, titulo_columna)
    y -= 12
    c.line(40, y, ancho - 40, y)
    y -= 12

    c.setFont("Helvetica", 9)

    def dividir_en_lineas(texto, ancho_disponible):
        palabras = str(texto or "").split()
        if not palabras:
            return [""]
        lineas, linea_actual = [], ""
        for palabra in palabras:
            candidata = f"{linea_actual} {palabra}".strip()
            if not linea_actual or c.stringWidth(candidata, "Helvetica", 9) <= ancho_disponible:
                linea_actual = candidata
            else:
                lineas.append(linea_actual)
                linea_actual = palabra
        if linea_actual:
            lineas.append(linea_actual)
        return lineas

    for fila in detalle:
        lineas_por_columna = []
        for idx, (_, posicion) in enumerate(columnas):
            valor = fila[idx] if idx < len(fila) else ""
            if idx in columnas_multilinea:
                siguiente_posicion = columnas[idx + 1][1] if idx + 1 < len(columnas) else ancho - 40
                lineas_por_columna.append(dividir_en_lineas(valor, siguiente_posicion - posicion - 6))
            else:
                limite = 14 if idx == 0 else 32
                lineas_por_columna.append([str(valor)[:limite]])

        alto_fila = max(len(lineas) for lineas in lineas_por_columna) * 12
        if y - alto_fila < 280:
            dibujar_usuario_pie()
            c.showPage()
            y = alto - 40
            c.setFont("Helvetica-Bold", 10)
            for titulo_columna, posicion in columnas:
                c.drawString(posicion, y, titulo_columna)
            y -= 12
            c.line(40, y, ancho - 40, y)
            y -= 12
            c.setFont("Helvetica", 9)
        for idx, (_, posicion) in enumerate(columnas):
            for linea_idx, linea in enumerate(lineas_por_columna[idx]):
                c.drawString(posicion, y - (linea_idx * 12), linea)
        y -= alto_fila

    y_bloque_pie = 185
    if bloque_pie_izq or bloque_pie_der:
        c.setFont("Helvetica", 11)
        fila_alt = 15
        for idx, (k, v) in enumerate(bloque_pie_izq or []):
            c.drawString(50, y_bloque_pie - (idx * fila_alt), f"{k}: {v}")
        for idx, (k, v) in enumerate(bloque_pie_der or []):
            c.drawString(320, y_bloque_pie - (idx * fila_alt), f"{k}: {v}")
        y = min(y, y_bloque_pie - (max(len(bloque_pie_izq or []), len(bloque_pie_der or [])) * fila_alt) - 8)

    if incluir_firmas:
        if firmas_layout == "triple_entrega_recepcion":
            y = min(y, 124)
            if bloque_destinatarios_titulo:
                c.setFont("Helvetica-Bold", 10)
                c.drawString(60, y, bloque_destinatarios_titulo)
                y -= 12
            filas_firma = [
                ("ENTREGA:", (firma_entrega_nombre or "").strip()),
                ("RECIBI CONFORME:", (firma_recibe_nombre or "").strip()),
                ("RECIBI CONFORME:", ""),
            ]
            for etiqueta, nombre in filas_firma:
                c.setFont("Helvetica-Bold", 9)
                c.drawString(60, y - 10, etiqueta)
                c.rect(185, y - 20, 160, 22, stroke=1, fill=0)
                if nombre:
                    c.setFont("Helvetica-Bold", 9)
                    c.drawCentredString(265, y - 7, nombre[:24])
                c.line(385, y - 5, 530, y - 5)
                c.setFont("Helvetica-Bold", 8)
                c.drawCentredString(457, y - 17, "Firma y Aclaracion")
                y -= 38
        else:
            if y < 120:
                dibujar_usuario_pie()
                c.showPage()
                y = alto - 70
            else:
                y -= 20
            c.setFont("Helvetica", 10)
            c.line(60, y, 250, y)
            c.drawString(60, y - 14, "Firma de entrega")
            c.line(340, y, 530, y)
            c.drawString(340, y - 14, "Firma de recepcion")
            y -= 30

    if observaciones_pie:
        c.setFont("Helvetica", 10)
        if incluir_firmas and firmas_layout == "triple_entrega_recepcion":
            y_obs = max(12, y - 4)
        else:
            y_obs = max(24, y - 10)
        prefijo = "Observaciones: "
        texto = (observaciones_pie or "").strip()
        ancho_max = 470

        palabras = texto.split()
        lineas = []
        actual = ""
        for palabra in palabras:
            candidato = f"{actual} {palabra}".strip()
            ancho_candidato = c.stringWidth((prefijo if not lineas else "") + candidato, "Helvetica", 10)
            if ancho_candidato <= ancho_max:
                actual = candidato
            else:
                if actual:
                    lineas.append(actual)
                actual = palabra
                if len(lineas) >= 2:
                    break
        if actual and len(lineas) < 2:
            lineas.append(actual)

        if len(lineas) > 2:
            lineas = lineas[:2]
        if len(lineas) == 2:
            restante = " ".join(palabras[len(" ".join(lineas).split()):]).strip()
            if restante:
                base = lineas[1]
                while base and c.stringWidth(base + "...", "Helvetica", 10) > ancho_max:
                    base = base[:-1]
                lineas[1] = (base.rstrip() + "...") if base else "..."

        if not lineas:
            lineas = [""]

        c.drawString(60, y_obs, f"{prefijo}{lineas[0]}")
        if len(lineas) > 1:
            c.drawString(60, y_obs - 12, lineas[1])

    dibujar_usuario_pie()

    c.save()


@app.get("/almacen", response_class=HTMLResponse)
def almacen_view(request: Request):
    redirect = _requiere_login(request)
    if redirect is not None:
        return redirect
    return _leer_html(ALMACEN_V2_PATH)


@app.get("/almacen/categorias")
def almacen_categorias():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM categorias ORDER BY nombre").fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/categorias")
def crear_categoria(payload: AlmacenSimplePayload):
    nombre = (payload.nombre or "").strip()
    if not nombre:
        return {"error": "Nombre requerido"}
    with get_sqlite_connection() as conn:
        conn.execute("INSERT INTO categorias (nombre) VALUES (?)", (nombre,))
        conn.commit()
    return {"mensaje": "Categoria creada"}


@app.delete("/almacen/categorias/{item_id}")
def eliminar_categoria(item_id: int):
    with get_sqlite_connection() as conn:
        conn.execute("DELETE FROM categorias WHERE id = ?", (item_id,))
        conn.commit()
    return {"mensaje": "Categoria eliminada"}


@app.get("/almacen/tipos_producto")
def almacen_tipos_producto():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM tipos_producto ORDER BY nombre").fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/tipos")
def crear_tipo_producto(payload: AlmacenSimplePayload):
    nombre = (payload.nombre or "").strip()
    if not nombre:
        return {"error": "Nombre requerido"}
    with get_sqlite_connection() as conn:
        conn.execute(
            "INSERT INTO tipos_producto (nombre, activo) VALUES (?, ?)",
            (nombre, 1 if payload.activo else 0),
        )
        conn.commit()
    return {"mensaje": "Tipo creado"}


@app.delete("/almacen/tipos_producto/{item_id}")
def eliminar_tipo_producto(item_id: int):
    with get_sqlite_connection() as conn:
        conn.execute("DELETE FROM tipos_producto WHERE id = ?", (item_id,))
        conn.commit()
    return {"mensaje": "Tipo eliminado"}


@app.get("/almacen/unidades_medida")
def almacen_unidades_medida():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM unidades_medida ORDER BY nombre").fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/unidades")
def crear_unidad_medida(payload: AlmacenSimplePayload):
    nombre = (payload.nombre or "").strip()
    if not nombre:
        return {"error": "Nombre requerido"}
    with get_sqlite_connection() as conn:
        conn.execute(
            "INSERT INTO unidades_medida (nombre, activo) VALUES (?, ?)",
            (nombre, 1 if payload.activo else 0),
        )
        conn.commit()
    return {"mensaje": "Unidad creada"}


@app.delete("/almacen/unidades_medida/{item_id}")
def eliminar_unidad_medida(item_id: int):
    with get_sqlite_connection() as conn:
        conn.execute("DELETE FROM unidades_medida WHERE id = ?", (item_id,))
        conn.commit()
    return {"mensaje": "Unidad eliminada"}


@app.get("/almacen/ubicaciones")
def almacen_ubicaciones():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM ubicaciones ORDER BY COALESCE(ruta, nombre), nombre").fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/ubicaciones")
def crear_ubicacion(payload: AlmacenUbicacionPayload):
    nombre = (payload.nombre or "").strip()
    if not nombre:
        return {"error": "Nombre requerido"}
    with get_sqlite_connection() as conn:
        ruta = nombre
        if payload.parent_id:
            parent = conn.execute("SELECT ruta, nombre FROM ubicaciones WHERE id = ?", (payload.parent_id,)).fetchone()
            if parent is None:
                return {"error": "Ubicacion padre inexistente"}
            ruta_padre = parent["ruta"] or parent["nombre"]
            ruta = f"{ruta_padre} / {nombre}"
        conn.execute(
            "INSERT INTO ubicaciones (nombre, parent_id, ruta, activo) VALUES (?, ?, ?, ?)",
            (nombre, payload.parent_id, ruta, 1 if payload.activo else 0),
        )
        conn.commit()
    return {"mensaje": "Ubicacion creada"}


@app.delete("/almacen/ubicaciones/{item_id}")
def eliminar_ubicacion(item_id: int):
    with get_sqlite_connection() as conn:
        conn.execute("DELETE FROM ubicaciones WHERE id = ?", (item_id,))
        conn.commit()
    return {"mensaje": "Ubicacion eliminada"}


@app.get("/almacen/productos")
def listar_productos_almacen():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM productos ORDER BY codigo").fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/productos")
def crear_producto_almacen(payload: AlmacenProductoPayload):
    codigo = (payload.codigo or "").strip()
    descripcion = (payload.descripcion or "").strip()
    if not codigo or not descripcion:
        return {"error": "Codigo y descripcion son obligatorios"}
    with get_sqlite_connection() as conn:
        existe = conn.execute("SELECT 1 FROM productos WHERE codigo = ?", (codigo,)).fetchone()
        if existe:
            return {"error": "El codigo ya existe"}
        cur = conn.execute(
            """
            INSERT INTO productos (
                codigo, descripcion, marca, modelo,
                categoria_id, tipo_producto_id, unidad_medida_id,
                stock_minimo, stock_maximo, punto_reposicion,
                ubicacion_id, observaciones, fecha_alta, usuario_alta,
                raw_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                codigo,
                descripcion,
                payload.marca,
                payload.modelo,
                payload.categoria_id,
                payload.tipo_producto_id,
                payload.unidad_medida_id,
                0,
                0,
                0,
                None,
                payload.observaciones,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                None,
                json.dumps(payload.model_dump(), ensure_ascii=False),
            ),
        )
        conn.commit()
    return {"mensaje": "Producto creado"}


@app.put("/almacen/productos/{producto_id}")
def actualizar_producto_almacen(producto_id: int, payload: AlmacenProductoPayload):
    codigo = (payload.codigo or "").strip()
    descripcion = (payload.descripcion or "").strip()
    if not codigo or not descripcion:
        return {"error": "Codigo y descripcion son obligatorios"}
    with get_sqlite_connection() as conn:
        existe = conn.execute("SELECT 1 FROM productos WHERE id = ?", (producto_id,)).fetchone()
        if existe is None:
            return {"error": "Producto inexistente"}
        existe_codigo = conn.execute(
            "SELECT 1 FROM productos WHERE codigo = ? AND id <> ?",
            (codigo, producto_id),
        ).fetchone()
        if existe_codigo:
            return {"error": "El codigo ya existe en otro producto"}
        conn.execute(
            """
            UPDATE productos
               SET codigo = ?,
                   descripcion = ?,
                   marca = ?,
                   modelo = ?,
                   categoria_id = ?,
                   tipo_producto_id = ?,
                   unidad_medida_id = ?,
                   stock_minimo = ?,
                   stock_maximo = ?,
                   punto_reposicion = ?,
                   ubicacion_id = ?,
                   observaciones = ?,
                   usuario_alta = ?,
                   raw_json = ?
             WHERE id = ?
            """,
            (
                codigo,
                descripcion,
                payload.marca,
                payload.modelo,
                payload.categoria_id,
                payload.tipo_producto_id,
                payload.unidad_medida_id,
                payload.stock_ingreso,
                0,
                payload.punto_reposicion,
                payload.ubicacion_id,
                payload.observaciones,
                None,
                json.dumps(payload.model_dump(), ensure_ascii=False),
                producto_id,
            ),
        )
        conn.commit()
    return {"mensaje": "Producto actualizado"}


@app.delete("/almacen/productos/{producto_id}")
def eliminar_producto_almacen(producto_id: int):
    with get_sqlite_connection() as conn:
        movimientos = conn.execute(
            "SELECT 1 FROM movimientos_stock WHERE producto_id = ? LIMIT 1",
            (producto_id,),
        ).fetchone()
        if movimientos:
            return {"error": "No se puede eliminar: el producto tiene movimientos"}
        conn.execute("DELETE FROM productos WHERE id = ?", (producto_id,))
        conn.commit()
    return {"mensaje": "Producto eliminado"}


@app.get("/almacen/stock")
def consulta_stock_almacen(
    codigo: str = "",
    descripcion: str = "",
    categoria_id: int | None = Query(default=None),
    ubicacion_id: int | None = Query(default=None),
):
    codigo_like = f"%{(codigo or '').strip()}%"
    descripcion_like = f"%{(descripcion or '').strip()}%"
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
                 SELECT p.id, p.codigo, p.descripcion, p.punto_reposicion,
                   c.nombre AS categoria,
                   u.ruta AS ubicacion
            FROM productos p
            LEFT JOIN categorias c ON c.id = p.categoria_id
            LEFT JOIN ubicaciones u ON u.id = p.ubicacion_id
            WHERE p.codigo LIKE ?
              AND p.descripcion LIKE ?
              AND (? IS NULL OR p.categoria_id = ?)
              AND (? IS NULL OR p.ubicacion_id = ?)
            ORDER BY p.codigo
            """,
            (codigo_like, descripcion_like, categoria_id, categoria_id, ubicacion_id, ubicacion_id),
        ).fetchall()

        result = []
        for r in rows:
            stock_actual = obtener_stock_actual(conn, r["id"])
            item = dict(r)
            item["stock_actual"] = stock_actual
            result.append(item)
    return result


@app.post("/almacen/remitos_ingreso")
def crear_remito_ingreso(payload: AlmacenRIPayload):
    if not payload.items:
        return {"error": "Debe indicar al menos un item"}
    fecha = payload.fecha or datetime.now().strftime("%Y-%m-%d")
    with get_sqlite_connection() as conn:
        numero = obtener_siguiente_numero(conn, "remitos_ingreso", "RI")
        responsable = obtener_persona_por_legajo(conn, payload.responsable_legajo)
        responsable_nombre = responsable["nombre"] if responsable else ""

        cur = conn.execute(
            """
            INSERT INTO remitos_ingreso (
                numero, proveedor, nro_remito_referencia, responsable_legajo, responsable_nombre,
                fecha, observaciones, raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                numero,
                payload.proveedor,
                payload.nro_remito_referencia,
                payload.responsable_legajo,
                responsable_nombre,
                fecha,
                payload.observaciones,
                json.dumps(payload.model_dump(), ensure_ascii=False),
            ),
        )
        remito_id = cur.lastrowid

        for item in payload.items:
            if item.cantidad <= 0:
                return {"error": "Cantidad invalida en items"}

            producto = conn.execute(
                """
                SELECT p.id, p.codigo, p.descripcion,
                       COALESCE(f.nombre, '') AS familia_nombre,
                       COALESCE(m.nombre, '') AS marca_nombre,
                       COALESCE(md.nombre, '') AS modelo_nombre,
                       COALESCE(u.nombre, '') AS unidad_nombre
                FROM productos p
                LEFT JOIN familias f ON f.id = p.id_familia
                LEFT JOIN marcas m ON m.id = p.id_marca
                LEFT JOIN modelos md ON md.id = p.id_modelo
                LEFT JOIN unidades_medida u ON u.id = COALESCE(p.id_unidad, p.unidad_medida_id)
                WHERE p.id = ?
                """,
                (item.producto_id,),
            ).fetchone()
            if producto is None:
                return {"error": f"Producto inexistente: {item.producto_id}"}

            conn.execute(
                """
                INSERT INTO remitos_ingreso_detalle (
                    remito_id, producto_id, cantidad, series_json, observaciones,
                    vehiculo_codigo, proyecto, instalacion
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    remito_id,
                    item.producto_id,
                    item.cantidad,
                    json.dumps(item.series, ensure_ascii=False),
                    payload.observaciones,
                    item.vehiculo_codigo,
                    item.proyecto,
                    item.instalacion,
                ),
            )

            registrar_movimiento_stock(
                conn,
                fecha=fecha,
                producto_id=item.producto_id,
                documento=numero,
                tipo="ENTRADA",
                cantidad=item.cantidad,
                responsable_legajo=payload.responsable_legajo,
                responsable_nombre=responsable_nombre,
                observaciones=payload.observaciones,
                remito_id=remito_id,
                vehiculo_codigo=item.vehiculo_codigo,
                proyecto=item.proyecto,
                instalacion=item.instalacion,
            )

            ajustar_stock_producto(conn, item.producto_id, float(item.cantidad))

        conn.commit()

        det_rows = conn.execute(
            """
            SELECT p.codigo, p.descripcion, d.cantidad, d.instalacion
            FROM remitos_ingreso_detalle d
            JOIN productos p ON p.id = d.producto_id
            WHERE d.remito_id = ?
            ORDER BY d.id
            """,
            (remito_id,),
        ).fetchall()
        ubicacion_ingreso = ""
        for r in det_rows:
            if (r["instalacion"] or "").strip():
                ubicacion_ingreso = (r["instalacion"] or "").strip()
                break
        detalle_pdf = [
            (
                str(idx),
                str(r["codigo"] or ""),
                str(r["descripcion"] or ""),
                str(r["cantidad"]),
            )
            for idx, r in enumerate(det_rows, start=1)
        ]
        pdf_path = os.path.join(DOC_ALMACEN_DIR, f"{numero}.pdf")
        generar_pdf_remito(
            pdf_path,
            f"Remito de Ingreso {numero}",
            [
                ("Fecha", fecha),
            ],
            detalle_pdf,
            [("N°", 40), ("Codigo", 85), ("Descripcion", 180), ("Cantidad", 470)],
            incluir_firmas=True,
            bloque_envio=[
                ("Proveedor", payload.proveedor or ""),
                ("Remito prov./interno", payload.nro_remito_referencia or ""),
                ("Responsable", responsable_nombre),
                ("Ubicacion", ubicacion_ingreso),
            ],
            bloque_envio_titulo="Datos de ingreso",
            observaciones_pie=payload.observaciones or "",
        )

    return {"mensaje": "Remito de ingreso creado", "numero": numero, "id": remito_id}


@app.get("/almacen/remitos_ingreso")
def listar_remitos_ingreso():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM remitos_ingreso ORDER BY id DESC").fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/remitos_ingreso/{remito_id}/pdf")
def pdf_remito_ingreso(remito_id: int):
    with get_sqlite_connection() as conn:
        row = conn.execute("SELECT * FROM remitos_ingreso WHERE id = ?", (remito_id,)).fetchone()
        det_rows = conn.execute(
            """
            SELECT p.codigo, p.descripcion, d.cantidad, d.instalacion
            FROM remitos_ingreso_detalle d
            JOIN productos p ON p.id = d.producto_id
            WHERE d.remito_id = ?
            ORDER BY d.id
            """,
            (remito_id,),
        ).fetchall()
    if row is None:
        return {"error": "Remito no encontrado"}

    detalle_pdf = [
        (
            str(idx),
            str(r["codigo"] or ""),
            str(r["descripcion"] or ""),
            str(r["cantidad"]),
        )
        for idx, r in enumerate(det_rows, start=1)
    ]
    ubicacion_ingreso = ""
    for r in det_rows:
        if (r["instalacion"] or "").strip():
            ubicacion_ingreso = (r["instalacion"] or "").strip()
            break

    path_pdf = os.path.join(DOC_ALMACEN_DIR, f"{row['numero']}.pdf")
    generar_pdf_remito(
        path_pdf,
        f"Remito de Ingreso {row['numero']}",
        [
            ("Fecha", row["fecha"] or ""),
        ],
        detalle_pdf,
        [("N°", 40), ("Codigo", 85), ("Descripcion", 180), ("Cantidad", 470)],
        incluir_firmas=True,
        bloque_envio=[
            ("Proveedor", row["proveedor"] or ""),
            ("Remito prov./interno", row["nro_remito_referencia"] or ""),
            ("Responsable", row["responsable_nombre"] or ""),
            ("Ubicacion", ubicacion_ingreso),
        ],
        bloque_envio_titulo="Datos de ingreso",
        observaciones_pie=row["observaciones"] or "",
    )

    if not os.path.exists(path_pdf):
        return {"error": "PDF no disponible"}
    return FileResponse(path_pdf, media_type="application/pdf", filename=os.path.basename(path_pdf))


def _estado_remito_entrega(estado: str | None):
    valor = str(estado or "AUTORIZADO").strip().upper()
    if valor in {"PENDIENTE", "AUTORIZADO", "NO_AUTORIZADO"}:
        return valor
    return "AUTORIZADO"


def _detalle_remito_entrega_rows(conn, remito_id: int):
    return conn.execute(
        """
        SELECT d.*, p.codigo, p.descripcion
        FROM remitos_entrega_detalle d
        JOIN productos p ON p.id = d.producto_id
        WHERE d.remito_id = ?
        ORDER BY d.id
        """,
        (remito_id,),
    ).fetchall()


def _generar_pdf_remito_entrega(conn, row, det_rows):
    estado = _estado_remito_entrega(row["estado_autorizacion"])
    numero = str(row["numero"] or "")
    detalle_pdf = [
        (
            str(idx),
            str(r["codigo"] or ""),
            str(r["descripcion"] or ""),
            str(r["cantidad"]),
        )
        for idx, r in enumerate(det_rows, start=1)
    ]
    sucursal_destino = ""
    for r in det_rows:
        if (r["instalacion"] or "").strip():
            sucursal_destino = (r["instalacion"] or "").strip()
            break

    datos_envio_pdf = [
        ("Razon social", row["razon_social"] or row["destinatario"] or ""),
        ("CUIT/DNI", row["cuit_dni"] or ""),
        ("Obra/Direccion", row["direccion"] or ""),
        ("Provincia", row["provincia"] or ""),
        ("Localidad", row["localidad"] or ""),
        ("Telefono", row["telefono"] or ""),
        ("Codigo postal", row["codigo_postal"] or ""),
    ]
    bloque_pie_izq = [
        ("Destinatario", row["destinatario"] or ""),
        ("Entrega", row["entrega_nombre"] or ""),
        ("Recibe", row["recibe_nombre"] or ""),
    ]
    bloque_pie_der = [
        ("Transporte", row["transporte"] or ""),
        ("Dominio", row["dominio"] or ""),
        ("Sucursal destino", sucursal_destino),
    ]

    estado_txt = {
        "PENDIENTE": "Pendiente de autorizacion",
        "AUTORIZADO": "Autorizado",
        "NO_AUTORIZADO": "No autorizado",
    }.get(estado, estado)
    aviso_estado = ""
    if estado == "PENDIENTE":
        aviso_estado = "MOVIMIENTO NO AUTORIZADO - Estado pendiente de autorizacion. Sin impacto en inventario."
    elif estado == "NO_AUTORIZADO":
        aviso_estado = "MOVIMIENTO NO AUTORIZADO."

    observaciones_parts = [x for x in [row["observaciones"] or "", aviso_estado] if str(x or "").strip()]

    campos_cabecera = [
        ("Fecha", row["fecha"] or ""),
        ("Estado", estado_txt),
    ]
    if estado == "AUTORIZADO" and (row["autorizado_por"] or "").strip():
        campos_cabecera.append(("Autorizado por", (row["autorizado_por"] or "").strip()))
    if estado == "AUTORIZADO" and (row["fecha_autorizacion"] or "").strip():
        campos_cabecera.append(("Fecha autorizacion", (row["fecha_autorizacion"] or "").strip()))

    pdf_path = os.path.join(DOC_ALMACEN_DIR, f"{numero}.pdf")
    generar_pdf_remito(
        pdf_path,
        f"Remito de Entrega {numero}",
        campos_cabecera,
        detalle_pdf,
        [("N°", 40), ("Codigo", 85), ("Detalle", 180), ("Cantidad", 470)],
        incluir_firmas=(estado == "AUTORIZADO"),
        firmas_layout="triple_entrega_recepcion",
        firma_entrega_nombre=row["entrega_nombre"] or "",
        firma_recibe_nombre=row["recibe_nombre"] or "",
        bloque_envio=datos_envio_pdf,
        bloque_destinatarios_titulo="Destinatarios",
        bloque_pie_izq=bloque_pie_izq,
        bloque_pie_der=bloque_pie_der,
        observaciones_pie=" | ".join(observaciones_parts),
    )
    return pdf_path


def _procesar_movimientos_remito_entrega(conn, row, det_rows):
    fecha = row["fecha"] or datetime.now().strftime("%Y-%m-%d")
    for d in det_rows:
        registrar_movimiento_stock(
            conn,
            fecha=fecha,
            producto_id=int(d["producto_id"]),
            documento=row["numero"],
            tipo="SALIDA",
            cantidad=float(d["cantidad"] or 0),
            responsable_legajo=row["entrega_legajo"],
            responsable_nombre=row["entrega_nombre"] or "",
            observaciones=row["observaciones"] or "",
            remito_id=int(row["id"]),
            vehiculo_codigo=d["vehiculo_codigo"] or "",
            proyecto=d["proyecto"] or "",
            instalacion=d["instalacion"] or "",
        )
        descontar_stock(conn, int(d["producto_id"]), float(d["cantidad"] or 0))


@app.post("/almacen/remitos_entrega")
def crear_remito_entrega(payload: AlmacenREPayload):
    if not payload.items:
        return {"error": "Debe indicar al menos un item"}
    if len(payload.items) > 25:
        return {"error": "La salida admite hasta 25 items"}

    fecha = payload.fecha or datetime.now().strftime("%Y-%m-%d")
    with get_sqlite_connection() as conn:
        migrar_remitos_entrega_autorizacion(conn)
        numero = obtener_siguiente_numero(conn, "remitos_entrega", "RE")
        entrega = obtener_persona_por_legajo(conn, payload.entrega_legajo)
        recibe = obtener_persona_por_legajo(conn, payload.recibe_legajo)
        entrega_nombre = (entrega["nombre"] if entrega else payload.entrega_nombre or "").strip()
        recibe_nombre = (recibe["nombre"] if recibe else payload.recibe_nombre or "").strip()

        for item in payload.items:
            producto = conn.execute(
                "SELECT id, codigo FROM productos WHERE id = ?",
                (item.producto_id,),
            ).fetchone()
            if producto is None:
                return {"error": f"Producto inexistente: {item.producto_id}"}
            stock_disponible = obtener_stock_disponible(conn, item.producto_id)
            if stock_disponible < float(item.cantidad):
                return {
                    "error": f"Stock insuficiente para {producto['codigo']} (disponible {stock_disponible}, requerido {item.cantidad})"
                }

        cur = conn.execute(
            """
            INSERT INTO remitos_entrega (
                numero, destinatario, razon_social, cuit_dni, direccion,
                localidad, provincia, telefono, codigo_postal,
                transporte, dominio,
                entrega_legajo, entrega_nombre, recibe_legajo, recibe_nombre,
                observaciones, fecha,
                estado_autorizacion, fecha_autorizacion, autorizado_por, observaciones_autorizacion,
                raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                numero,
                payload.destinatario,
                payload.razon_social,
                payload.cuit_dni,
                payload.direccion,
                payload.localidad,
                payload.provincia,
                payload.telefono,
                payload.codigo_postal,
                payload.transporte,
                payload.dominio,
                payload.entrega_legajo,
                entrega_nombre,
                payload.recibe_legajo,
                recibe_nombre,
                payload.observaciones,
                fecha,
                "PENDIENTE",
                None,
                "",
                "",
                json.dumps(payload.model_dump(), ensure_ascii=False),
            ),
        )
        remito_id = cur.lastrowid

        for item in payload.items:
            conn.execute(
                """
                INSERT INTO remitos_entrega_detalle (
                    remito_id, producto_id, cantidad, series_json, observaciones,
                    vehiculo_codigo, proyecto, instalacion
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    remito_id,
                    item.producto_id,
                    item.cantidad,
                    json.dumps(item.series, ensure_ascii=False),
                    payload.observaciones,
                    item.vehiculo_codigo,
                    item.proyecto,
                    item.instalacion,
                ),
            )

        row = conn.execute("SELECT * FROM remitos_entrega WHERE id = ?", (remito_id,)).fetchone()
        det_rows = _detalle_remito_entrega_rows(conn, remito_id)
        _generar_pdf_remito_entrega(conn, row, det_rows)
        conn.commit()

    return {"mensaje": "Remito de entrega registrado en estado PENDIENTE", "numero": numero, "id": remito_id, "estado_autorizacion": "PENDIENTE"}


@app.get("/almacen/remitos_entrega")
def listar_remitos_entrega():
    with get_sqlite_connection() as conn:
        migrar_remitos_entrega_autorizacion(conn)
        rows = conn.execute("SELECT * FROM remitos_entrega ORDER BY id DESC").fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/remitos_entrega/pendientes")
def listar_remitos_entrega_pendientes():
    with get_sqlite_connection() as conn:
        migrar_remitos_entrega_autorizacion(conn)
        rows = conn.execute(
            """
            SELECT r.id, r.numero, r.fecha, r.destinatario, r.observaciones,
                   COUNT(d.id) AS total_items,
                   COALESCE(SUM(d.cantidad), 0) AS cantidad_total
            FROM remitos_entrega r
            LEFT JOIN remitos_entrega_detalle d ON d.remito_id = r.id
            WHERE UPPER(COALESCE(r.estado_autorizacion, 'AUTORIZADO')) = 'PENDIENTE'
            GROUP BY r.id, r.numero, r.fecha, r.destinatario, r.observaciones
            ORDER BY r.id DESC
            """
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/remitos_entrega/{remito_id}/autorizar")
def autorizar_remito_entrega(remito_id: int, payload: AlmacenAutorizacionREPayload):
    with get_sqlite_connection() as conn:
        migrar_remitos_entrega_autorizacion(conn)
        row = conn.execute("SELECT * FROM remitos_entrega WHERE id = ?", (remito_id,)).fetchone()
        if row is None:
            return {"error": "Remito no encontrado"}

        estado = _estado_remito_entrega(row["estado_autorizacion"])
        if estado == "AUTORIZADO":
            return {"mensaje": "El remito ya esta autorizado", "estado_autorizacion": "AUTORIZADO"}
        if estado == "NO_AUTORIZADO":
            return {"error": "El remito fue marcado como NO AUTORIZADO"}

        det_rows = _detalle_remito_entrega_rows(conn, remito_id)
        for d in det_rows:
            stock_disponible = obtener_stock_disponible(conn, int(d["producto_id"]), excluir_remito_id=remito_id)
            if stock_disponible < float(d["cantidad"] or 0):
                return {
                    "error": f"Stock insuficiente para {d['codigo']} (disponible {stock_disponible}, requerido {d['cantidad']})"
                }

        _procesar_movimientos_remito_entrega(conn, row, det_rows)
        conn.execute(
            """
            UPDATE remitos_entrega
               SET estado_autorizacion = 'AUTORIZADO',
                   fecha_autorizacion = ?,
                   autorizado_por = ?,
                   observaciones_autorizacion = ?
             WHERE id = ?
            """,
            (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                (payload.autorizado_por or "").strip(),
                (payload.observaciones or "").strip(),
                remito_id,
            ),
        )

        row_new = conn.execute("SELECT * FROM remitos_entrega WHERE id = ?", (remito_id,)).fetchone()
        _generar_pdf_remito_entrega(conn, row_new, det_rows)
        conn.commit()

    return {"mensaje": "Remito autorizado y movimiento ejecutado", "estado_autorizacion": "AUTORIZADO"}


@app.post("/almacen/remitos_entrega/{remito_id}/no_autorizar")
def no_autorizar_remito_entrega(remito_id: int, payload: AlmacenAutorizacionREPayload):
    with get_sqlite_connection() as conn:
        migrar_remitos_entrega_autorizacion(conn)
        row = conn.execute("SELECT * FROM remitos_entrega WHERE id = ?", (remito_id,)).fetchone()
        if row is None:
            return {"error": "Remito no encontrado"}

        estado = _estado_remito_entrega(row["estado_autorizacion"])
        if estado == "AUTORIZADO":
            return {"error": "El remito ya fue autorizado y ejecutado"}

        conn.execute(
            """
            UPDATE remitos_entrega
               SET estado_autorizacion = 'NO_AUTORIZADO',
                   fecha_autorizacion = ?,
                   autorizado_por = ?,
                   observaciones_autorizacion = ?
             WHERE id = ?
            """,
            (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                (payload.autorizado_por or "").strip(),
                (payload.observaciones or "").strip(),
                remito_id,
            ),
        )
        row_new = conn.execute("SELECT * FROM remitos_entrega WHERE id = ?", (remito_id,)).fetchone()
        det_rows = _detalle_remito_entrega_rows(conn, remito_id)
        _generar_pdf_remito_entrega(conn, row_new, det_rows)
        conn.commit()

    return {"mensaje": "Remito marcado como NO AUTORIZADO", "estado_autorizacion": "NO_AUTORIZADO"}


@app.get("/almacen/remitos_entrega/{remito_id}/pdf")
def pdf_remito_entrega(remito_id: int):
    with get_sqlite_connection() as conn:
        migrar_remitos_entrega_autorizacion(conn)
        row = conn.execute("SELECT * FROM remitos_entrega WHERE id = ?", (remito_id,)).fetchone()
        if row is None:
            return {"error": "Remito no encontrado"}
        det_rows = _detalle_remito_entrega_rows(conn, remito_id)
        path_pdf = _generar_pdf_remito_entrega(conn, row, det_rows)
    if not os.path.exists(path_pdf):
        return {"error": "PDF no disponible"}
    return FileResponse(path_pdf, media_type="application/pdf", filename=os.path.basename(path_pdf))


@app.get("/almacen/movimientos")
def listar_movimientos_almacen(
    fecha_desde: str = "",
    fecha_hasta: str = "",
    producto_id: int | None = Query(default=None),
    tipo: str = "",
):
    tipo = (tipo or "").strip().upper()
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT m.*, p.codigo AS producto_codigo, p.descripcion AS producto_descripcion
            FROM movimientos_stock m
            JOIN productos p ON p.id = m.producto_id
            WHERE (? = '' OR m.fecha >= ?)
              AND (? = '' OR m.fecha <= ?)
              AND (? IS NULL OR m.producto_id = ?)
              AND (? = '' OR UPPER(m.tipo) = ?)
            ORDER BY m.id DESC
            LIMIT 1000
            """,
            (fecha_desde, fecha_desde, fecha_hasta, fecha_hasta, producto_id, producto_id, tipo, tipo),
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/dashboard_data")
def dashboard_almacen_data():
    with get_sqlite_connection() as conn:
        total_productos = conn.execute("SELECT COUNT(*) AS total FROM productos").fetchone()["total"]
        productos_activos = total_productos

        stock_rows = conn.execute(
            "SELECT id, punto_reposicion FROM productos"
        ).fetchall()
        sin_stock = 0
        stock_critico = 0
        for r in stock_rows:
            stock_actual = obtener_stock_actual(conn, r["id"])
            if stock_actual <= 0:
                sin_stock += 1
            if stock_actual <= float(r["punto_reposicion"] or 0):
                stock_critico += 1

        productos_por_categoria = conn.execute(
            """
            SELECT c.nombre, COUNT(*) AS total
            FROM productos p
            LEFT JOIN categorias c ON c.id = p.categoria_id
            GROUP BY c.nombre
            ORDER BY total DESC
            """
        ).fetchall()

        mov_mensuales = conn.execute(
            """
            SELECT SUBSTR(COALESCE(fecha, ''), 1, 7) AS mes, COUNT(*) AS total
            FROM movimientos_stock
            WHERE SUBSTR(COALESCE(fecha, ''), 1, 7) <> ''
            GROUP BY SUBSTR(COALESCE(fecha, ''), 1, 7)
            ORDER BY mes DESC
            LIMIT 12
            """
        ).fetchall()

        stock_por_ubicacion_raw = conn.execute(
            """
            SELECT p.id, u.ruta AS ubicacion
            FROM productos p
            LEFT JOIN ubicaciones u ON u.id = p.ubicacion_id
            """
        ).fetchall()
        acumulado = {}
        for r in stock_por_ubicacion_raw:
            key = r["ubicacion"] or "Sin ubicacion"
            acumulado[key] = acumulado.get(key, 0.0) + obtener_stock_actual(conn, r["id"])
        stock_por_ubicacion = [{"nombre": k, "total": v} for k, v in acumulado.items()]

        ultimos_mov = conn.execute(
            """
            SELECT m.fecha, m.documento, m.tipo, m.cantidad, m.responsable_nombre,
                   p.codigo AS producto_codigo, p.descripcion AS producto_descripcion
            FROM movimientos_stock m
            JOIN productos p ON p.id = m.producto_id
            ORDER BY m.id DESC
            LIMIT 25
            """
        ).fetchall()

    return {
        "total_productos": total_productos,
        "productos_activos": productos_activos,
        "stock_critico": stock_critico,
        "sin_stock": sin_stock,
        "productos_por_categoria": [dict(r) for r in productos_por_categoria],
        "movimientos_mensuales": [dict(r) for r in reversed(mov_mensuales)],
        "stock_por_ubicacion": stock_por_ubicacion,
        "ultimos_movimientos": [dict(r) for r in ultimos_mov],
    }


@app.post("/almacen/inventarios")
def crear_inventario(payload: AlmacenInventarioPayload):
    with get_sqlite_connection() as conn:
        numero = obtener_siguiente_numero(conn, "inventarios", "INV")
        fecha = datetime.now().strftime("%Y-%m-%d")
        responsable = obtener_persona_por_legajo(conn, payload.responsable_legajo)
        responsable_nombre = responsable["nombre"] if responsable else ""

        cur = conn.execute(
            """
            INSERT INTO inventarios (
                numero, fecha, estado, responsable_legajo, responsable_nombre,
                observaciones, raw_json
            ) VALUES (?, ?, 'PENDIENTE', ?, ?, ?, ?)
            """,
            (
                numero,
                fecha,
                payload.responsable_legajo,
                responsable_nombre,
                payload.observaciones,
                json.dumps(payload.model_dump(), ensure_ascii=False),
            ),
        )
        inventario_id = cur.lastrowid

        productos = conn.execute("SELECT id FROM productos ORDER BY codigo").fetchall()
        for p in productos:
            stock_sistema = obtener_stock_actual(conn, p["id"])
            conn.execute(
                """
                INSERT INTO inventarios_detalle (
                    inventario_id, producto_id, stock_sistema, estado
                ) VALUES (?, ?, ?, 'PENDIENTE')
                """,
                (inventario_id, p["id"], stock_sistema),
            )

        conn.commit()

    return {"mensaje": "Inventario generado", "id": inventario_id, "numero": numero}


@app.get("/almacen/inventarios")
def listar_inventarios():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM inventarios ORDER BY id DESC").fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/inventarios/{inventario_id}/detalle")
def detalle_inventario(inventario_id: int):
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT d.*, p.codigo AS producto_codigo, p.descripcion AS producto_descripcion
            FROM inventarios_detalle d
            JOIN productos p ON p.id = d.producto_id
            WHERE d.inventario_id = ?
            ORDER BY p.codigo
            """,
            (inventario_id,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.put("/almacen/inventarios/detalle/{detalle_id}")
def registrar_conteo_inventario(detalle_id: int, payload: AlmacenConteoPayload):
    with get_sqlite_connection() as conn:
        row = conn.execute(
            "SELECT * FROM inventarios_detalle WHERE id = ?",
            (detalle_id,),
        ).fetchone()
        if row is None:
            return {"error": "Detalle de inventario inexistente"}

        diferencia = float(payload.stock_fisico) - float(row["stock_sistema"] or 0)
        estado = "OK" if abs(diferencia) < 1e-9 else "CON_DIFERENCIA"
        conn.execute(
            """
            UPDATE inventarios_detalle
               SET stock_fisico = ?,
                   diferencia = ?,
                   estado = ?,
                   observaciones = ?
             WHERE id = ?
            """,
            (payload.stock_fisico, diferencia, estado, payload.observaciones, detalle_id),
        )

        if abs(diferencia) > 1e-9:
            inv_row = conn.execute(
                "SELECT id, responsable_legajo, responsable_nombre FROM inventarios WHERE id = ?",
                (row["inventario_id"],),
            ).fetchone()
            ajuste_existente = conn.execute(
                "SELECT id FROM ajustes_stock WHERE inventario_detalle_id = ?",
                (detalle_id,),
            ).fetchone()
            if ajuste_existente is None:
                conn.execute(
                    """
                    INSERT INTO ajustes_stock (
                        inventario_id, inventario_detalle_id, producto_id,
                        cantidad_ajuste, estado,
                        solicitado_por_legajo, solicitado_por_nombre,
                        fecha_solicitud, observaciones, raw_json
                    ) VALUES (?, ?, ?, ?, 'PENDIENTE_APROBACION', ?, ?, ?, ?, ?)
                    """,
                    (
                        row["inventario_id"],
                        detalle_id,
                        row["producto_id"],
                        diferencia,
                        inv_row["responsable_legajo"] if inv_row else None,
                        inv_row["responsable_nombre"] if inv_row else "",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        payload.observaciones,
                        json.dumps(
                            {
                                "detalle_id": detalle_id,
                                "stock_sistema": row["stock_sistema"],
                                "stock_fisico": payload.stock_fisico,
                                "diferencia": diferencia,
                            },
                            ensure_ascii=False,
                        ),
                    ),
                )

        conn.commit()
    return {"mensaje": "Conteo guardado", "diferencia": diferencia, "estado": estado}


@app.get("/almacen/ajustes")
def listar_ajustes_almacen():
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT a.*, p.codigo AS producto_codigo, p.descripcion AS producto_descripcion
            FROM ajustes_stock a
            JOIN productos p ON p.id = a.producto_id
            ORDER BY a.id DESC
            """
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/roles")
def listar_roles_almacen():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM roles_almacen ORDER BY nombre").fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/roles")
def crear_rol_almacen(payload: AlmacenRolPayload):
    nombre = (payload.nombre or "").strip()
    if not nombre:
        return {"error": "Nombre requerido"}
    with get_sqlite_connection() as conn:
        conn.execute(
            "INSERT INTO roles_almacen (nombre, permisos_json, activo) VALUES (?, ?, 1)",
            (nombre, payload.permisos_json or "{}"),
        )
        conn.commit()
    return {"mensaje": "Rol creado"}


@app.post("/almacen/roles/asignar")
def asignar_rol_almacen(payload: AlmacenAsignacionRolPayload):
    with get_sqlite_connection() as conn:
        persona = conn.execute("SELECT 1 FROM personal WHERE legajo = ?", (payload.legajo,)).fetchone()
        rol = conn.execute("SELECT 1 FROM roles_almacen WHERE id = ?", (payload.rol_id,)).fetchone()
        if persona is None:
            return {"error": "Legajo inexistente"}
        if rol is None:
            return {"error": "Rol inexistente"}
        conn.execute(
            "INSERT OR IGNORE INTO personal_roles_almacen (legajo, rol_id) VALUES (?, ?)",
            (payload.legajo, payload.rol_id),
        )
        conn.commit()
    return {"mensaje": "Rol asignado"}


# -------- ALMACEN V2 (Tomos I-IV) --------


ALMACEN_V2_CATALOGOS = {
    "categorias": "categorias",
    "familias": "familias",
    "marcas": "marcas",
    "modelos": "modelos",
    "unidades": "unidades_medida",
    "ubicaciones": "ubicaciones",
    "proyectos": "proyectos",
    "instalaciones": "instalaciones",
}

TIPOS_DOCUMENTO_VALIDOS = {"RI", "RE", "RD", "RT", "AJ", "INV"}


class CatalogoV2Payload(BaseModel):
    nombre: str
    codigo: str | None = None
    descripcion: str | None = None
    activo: int = 1
    id_categoria: int | None = None
    id_marca: int | None = None
    id_proyecto: int | None = None
    cliente: str | None = None
    id_padre: int | None = None
    abreviatura: str | None = None


class ProductoV2Payload(BaseModel):
    codigo: str
    descripcion: str
    id_categoria: int | None = None
    id_familia: int | None = None
    id_marca: int | None = None
    id_modelo: int | None = None
    id_unidad: int | None = None
    id_ubicacion: int | None = None
    tipo_control: str = "CONSUMIBLE"
    stock_ingreso: float = 0
    punto_reposicion: float = 0
    observaciones: str = ""


class DocumentoDetalleV2Payload(BaseModel):
    id_producto: int | None = None
    cantidad: float = 0
    id_vehiculo: str | None = None
    id_proyecto: int | None = None
    id_instalacion: int | None = None
    observaciones: str = ""


class DocumentoV2Payload(BaseModel):
    tipo: str
    fecha: str | None = None
    responsable: str | None = None
    observaciones: str = ""
    detalles: list[DocumentoDetalleV2Payload]


class InventarioConteoV2Payload(BaseModel):
    stock_fisico: float
    observaciones: str = ""


def fecha_actual_iso():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def numero_documento_v2(conn, tipo: str):
    prefijo = tipo.upper()
    rows = conn.execute("SELECT numero FROM documentos WHERE tipo = ?", (prefijo,)).fetchall()
    max_num = 0
    for r in rows:
        nro = str(r["numero"] or "")
        if nro.startswith(f"{prefijo}-"):
            try:
                max_num = max(max_num, int(nro.split("-")[1]))
            except Exception:
                pass
    return f"{prefijo}-{max_num + 1:06d}"


def stock_actual_v2(conn, id_producto: int):
    row = conn.execute(
        "SELECT stock_nuevo FROM movimientos WHERE id_producto = ? ORDER BY id DESC LIMIT 1",
        (id_producto,),
    ).fetchone()
    return float(row["stock_nuevo"] or 0) if row else 0.0


def auditar_v2(conn, usuario: str | None, accion: str, tabla: str, registro: int | None, anterior=None, nuevo=None):
    conn.execute(
        """
        INSERT INTO auditoria (fecha, usuario, accion, tabla, registro, valor_anterior, valor_nuevo)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            fecha_actual_iso(),
            usuario,
            accion,
            tabla,
            registro,
            json.dumps(anterior, ensure_ascii=False) if anterior is not None else None,
            json.dumps(nuevo, ensure_ascii=False) if nuevo is not None else None,
        ),
    )


def producto_v2_por_id(conn, id_producto: int):
    return conn.execute("SELECT * FROM productos WHERE id = ?", (id_producto,)).fetchone()


def movimiento_v2_registrar(
    conn,
    *,
    tipo: str,
    documento: str,
    id_documento: int | None,
    id_producto: int | None,
    cantidad: float,
    id_personal: str | None,
    id_vehiculo: str | None,
    id_proyecto: int | None,
    id_instalacion: int | None,
    observaciones: str,
):
    stock_anterior = 0.0
    stock_nuevo = 0.0
    if id_producto is not None:
        stock_anterior = stock_actual_v2(conn, id_producto)
        if tipo in {"RI", "RD"}:
            stock_nuevo = stock_anterior + float(cantidad)
        elif tipo in {"RE", "RT"}:
            stock_nuevo = stock_anterior - float(cantidad)
        elif tipo in {"AJ", "INV"}:
            stock_nuevo = stock_anterior + float(cantidad)
        else:
            stock_nuevo = stock_anterior

    cur = conn.execute(
        """
        INSERT INTO movimientos (
            fecha, tipo, documento, id_documento, id_producto,
            cantidad, stock_anterior, stock_nuevo,
            id_personal, id_vehiculo, id_proyecto, id_instalacion,
            observaciones, raw_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            fecha_actual_iso(),
            tipo,
            documento,
            id_documento,
            id_producto,
            cantidad,
            stock_anterior,
            stock_nuevo,
            id_personal,
            id_vehiculo,
            id_proyecto,
            id_instalacion,
            observaciones,
            json.dumps(
                {
                    "tipo": tipo,
                    "documento": documento,
                    "id_producto": id_producto,
                    "cantidad": cantidad,
                    "stock_anterior": stock_anterior,
                    "stock_nuevo": stock_nuevo,
                },
                ensure_ascii=False,
            ),
        ),
    )

    if id_producto is not None:
        conn.execute(
            """
            INSERT INTO stock_consumibles (id_producto, stock_actual, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(id_producto) DO UPDATE SET
                stock_actual = excluded.stock_actual,
                updated_at = excluded.updated_at
            """,
            (id_producto, stock_nuevo, fecha_actual_iso()),
        )

    return cur.lastrowid, stock_anterior, stock_nuevo


@app.get("/almacen/v2/dashboard")
def dashboard_almacen_v2(top_n: int = Query(default=10, ge=1, le=50)):
    mes_actual = datetime.now().strftime("%Y-%m")
    with get_sqlite_connection() as conn:
        total_productos = conn.execute("SELECT COUNT(*) AS n FROM productos").fetchone()["n"]
        total_consumibles = total_productos
        stock_base = conn.execute(
            """
            SELECT
                p.id,
                COALESCE(s.stock_actual, 0) AS stock_actual,
                COALESCE(p.punto_reposicion, 0) AS punto_reposicion
            FROM productos p
            LEFT JOIN stock s ON s.producto_id = p.id
            """
        ).fetchall()
        stock_critico = sum(
            1
            for r in stock_base
            if float(r["stock_actual"] or 0) > 0 and float(r["stock_actual"] or 0) <= float(r["punto_reposicion"] or 0)
        )
        sin_stock = sum(1 for r in stock_base if float(r["stock_actual"] or 0) <= 0)

        ingresos_mes = conn.execute(
            """
            SELECT COALESCE(SUM(cantidad), 0) AS total
            FROM movimientos_stock
            WHERE SUBSTR(COALESCE(fecha, ''), 1, 7) = ?
              AND UPPER(tipo) = 'ENTRADA'
            """,
            (mes_actual,),
        ).fetchone()["total"]
        entregas_mes = conn.execute(
            """
            SELECT COALESCE(SUM(cantidad), 0) AS total
            FROM movimientos_stock
            WHERE SUBSTR(COALESCE(fecha, ''), 1, 7) = ?
              AND UPPER(tipo) = 'SALIDA'
            """,
            (mes_actual,),
        ).fetchone()["total"]

        movimientos_mensuales = conn.execute(
            """
            SELECT
                SUBSTR(COALESCE(fecha, ''), 1, 7) AS mes,
                COALESCE(SUM(CASE WHEN UPPER(tipo) = 'ENTRADA' THEN cantidad ELSE 0 END), 0) AS ingresos,
                COALESCE(SUM(CASE WHEN UPPER(tipo) = 'SALIDA' THEN cantidad ELSE 0 END), 0) AS entregas
            FROM movimientos_stock
            WHERE SUBSTR(COALESCE(fecha, ''), 1, 7) <> ''
            GROUP BY SUBSTR(COALESCE(fecha, ''), 1, 7)
            ORDER BY mes DESC
            LIMIT 12
            """
        ).fetchall()
        por_categoria = conn.execute(
            """
            SELECT
                COALESCE(c.nombre, 'Sin categoria') AS categoria,
                COALESCE(SUM(COALESCE(s.stock_actual, 0)), 0) AS total
            FROM productos p
            LEFT JOIN categorias c ON c.id = p.categoria_id
            LEFT JOIN stock s ON s.producto_id = p.id
            GROUP BY COALESCE(c.nombre, 'Sin categoria')
            ORDER BY total DESC
            """
        ).fetchall()

        consumo_por_proyecto = conn.execute(
            """
            SELECT
                COALESCE(NULLIF(TRIM(proyecto), ''), 'Sin proyecto') AS proyecto,
                COALESCE(SUM(cantidad), 0) AS total
            FROM movimientos_stock
            WHERE UPPER(tipo) = 'SALIDA'
            GROUP BY COALESCE(NULLIF(TRIM(proyecto), ''), 'Sin proyecto')
            ORDER BY total DESC
            LIMIT 12
            """
        ).fetchall()

        top_materiales = conn.execute(
            """
            SELECT
                p.codigo,
                p.descripcion,
                COALESCE(SUM(m.cantidad), 0) AS total
            FROM movimientos_stock m
            JOIN productos p ON p.id = m.producto_id
            WHERE UPPER(m.tipo) = 'SALIDA'
            GROUP BY p.id
            ORDER BY total DESC, p.codigo ASC
            LIMIT ?
            """,
            (top_n,),
        ).fetchall()

        alertas_criticas = conn.execute(
            """
            SELECT
                p.codigo,
                p.descripcion,
                COALESCE(c.nombre, 'Sin categoria') AS categoria,
                COALESCE(u.ruta, 'Sin ubicacion') AS ubicacion,
                COALESCE(s.stock_actual, 0) AS stock_actual,
                COALESCE(p.punto_reposicion, 0) AS punto_reposicion
            FROM productos p
            LEFT JOIN categorias c ON c.id = p.categoria_id
            LEFT JOIN ubicaciones u ON u.id = p.ubicacion_id
            LEFT JOIN stock s ON s.producto_id = p.id
            WHERE COALESCE(s.stock_actual, 0) > 0
              AND COALESCE(s.stock_actual, 0) <= COALESCE(p.punto_reposicion, 0)
            ORDER BY (COALESCE(p.punto_reposicion, 0) - COALESCE(s.stock_actual, 0)) DESC, p.codigo ASC
            LIMIT 10
            """
        ).fetchall()

        sin_stock_items = conn.execute(
            """
            SELECT
                p.codigo,
                p.descripcion,
                COALESCE(c.nombre, 'Sin categoria') AS categoria,
                COALESCE(u.ruta, 'Sin ubicacion') AS ubicacion,
                COALESCE(s.stock_actual, 0) AS stock_actual,
                COALESCE(p.punto_reposicion, 0) AS punto_reposicion
            FROM productos p
            LEFT JOIN categorias c ON c.id = p.categoria_id
            LEFT JOIN ubicaciones u ON u.id = p.ubicacion_id
            LEFT JOIN stock s ON s.producto_id = p.id
            WHERE COALESCE(s.stock_actual, 0) <= 0
            ORDER BY p.codigo ASC
            LIMIT 10
            """
        ).fetchall()
    return {
        "total_productos": total_productos,
        "total_consumibles": total_consumibles,
        "ingresos_mes": ingresos_mes,
        "entregas_mes": entregas_mes,
        "stock_critico": stock_critico,
        "sin_stock": sin_stock,
        "movimientos_mensuales": [dict(r) for r in movimientos_mensuales],
        "stock_por_categoria": [dict(r) for r in por_categoria],
        "consumo_por_proyecto": [dict(r) for r in consumo_por_proyecto],
        "top_materiales": [dict(r) for r in top_materiales],
        "alertas_criticas": [dict(r) for r in alertas_criticas],
        "sin_stock_items": [dict(r) for r in sin_stock_items],
    }


@app.get("/almacen/v2/catalogos/{tipo}")
def listar_catalogo_v2(tipo: str):
    tabla = ALMACEN_V2_CATALOGOS.get(tipo)
    if not tabla:
        return {"error": "Catalogo no soportado"}
    with get_sqlite_connection() as conn:
        rows = conn.execute(f"SELECT * FROM {tabla} ORDER BY id DESC").fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/v2/catalogos/{tipo}")
def crear_catalogo_v2(tipo: str, payload: CatalogoV2Payload):
    tabla = ALMACEN_V2_CATALOGOS.get(tipo)
    if not tabla:
        return {"error": "Catalogo no soportado"}
    try:
        with get_sqlite_connection() as conn:
            if tabla == "categorias":
                cur = conn.execute("INSERT INTO categorias (nombre) VALUES (?)", (payload.nombre,))
            elif tabla == "familias":
                cur = conn.execute("INSERT INTO familias (nombre) VALUES (?)", (payload.nombre,))
            elif tabla == "modelos":
                cur = conn.execute(
                    "INSERT INTO modelos (id_marca, nombre, descripcion, activo) VALUES (?, ?, ?, ?)",
                    (payload.id_marca, payload.nombre, payload.descripcion, 1 if payload.activo else 0),
                )
            elif tabla == "proyectos":
                cur = conn.execute(
                    "INSERT INTO proyectos (codigo, nombre, cliente, activo) VALUES (?, ?, ?, ?)",
                    (payload.codigo or payload.nombre[:8].upper(), payload.nombre, payload.cliente, 1 if payload.activo else 0),
                )
            elif tabla == "instalaciones":
                cur = conn.execute(
                    "INSERT INTO instalaciones (id_proyecto, codigo, nombre, activo) VALUES (?, ?, ?, ?)",
                    (payload.id_proyecto, payload.codigo or payload.nombre[:6].upper(), payload.nombre, 1 if payload.activo else 0),
                )
            elif tabla == "ubicaciones":
                cur = conn.execute(
                    "INSERT INTO ubicaciones (parent_id, nombre, ruta, activo) VALUES (?, ?, ?, ?)",
                    (payload.id_padre, payload.nombre, payload.nombre, 1 if payload.activo else 0),
                )
            elif tabla == "unidades_medida":
                cur = conn.execute(
                    "INSERT INTO unidades_medida (nombre, activo) VALUES (?, ?)",
                    (payload.nombre, 1 if payload.activo else 0),
                )
            else:
                cur = conn.execute(
                    f"INSERT INTO {tabla} (nombre, activo) VALUES (?, ?)",
                    (payload.nombre, 1 if payload.activo else 0),
                )
            auditar_v2(conn, None, "CREATE", tabla, cur.lastrowid, None, payload.model_dump())
            conn.commit()
    except sqlite3.OperationalError as exc:
        if "database is locked" in str(exc).lower():
            return {"error": "Base de datos ocupada. Reintente en unos segundos."}
        raise
    return {"mensaje": "Registro creado"}


@app.delete("/almacen/v2/catalogos/{tipo}/{item_id}")
def eliminar_catalogo_v2(tipo: str, item_id: int):
    tabla = ALMACEN_V2_CATALOGOS.get(tipo)
    if not tabla:
        return {"error": "Catalogo no soportado"}
    with get_sqlite_connection() as conn:
        prev = conn.execute(f"SELECT * FROM {tabla} WHERE id = ?", (item_id,)).fetchone()
        if prev is None:
            return {"error": "Registro inexistente"}
        conn.execute(f"DELETE FROM {tabla} WHERE id = ?", (item_id,))
        auditar_v2(conn, None, "DELETE", tabla, item_id, dict(prev), None)
        conn.commit()
    return {"mensaje": "Registro eliminado"}


@app.get("/almacen/v2/productos")
def listar_productos_v2():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM productos ORDER BY codigo").fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/v2/productos/sugerencias-oc")
def sugerencias_productos_desde_ordenes_compra():
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            '''
            SELECT DISTINCT TRIM(ocd.codigo_proveedor) AS codigo,
                            TRIM(ocd.descripcion) AS descripcion
            FROM OrdenCompraDetalle ocd
            INNER JOIN OrdenCompra oc ON oc.id = ocd.orden_compra_id
            WHERE UPPER(COALESCE(oc.estado, '')) <> 'ANULADA'
              AND (TRIM(COALESCE(ocd.codigo_proveedor, '')) <> ''
                   OR TRIM(COALESCE(ocd.descripcion, '')) <> '')
            ORDER BY codigo, descripcion
            '''
        ).fetchall()
    return [dict(row) for row in rows]


@app.post("/almacen/v2/productos")
def crear_producto_v2(payload: ProductoV2Payload):
    if not payload.codigo.strip() or not payload.descripcion.strip():
        return {"error": "Codigo y descripcion son obligatorios"}
    try:
        with get_sqlite_connection() as conn:
            existe = conn.execute("SELECT 1 FROM productos WHERE codigo = ?", (payload.codigo.strip(),)).fetchone()
            if existe:
                return {"error": "Codigo duplicado"}
            cur = conn.execute(
                """
                INSERT INTO productos (
                    codigo, descripcion,
                    categoria_id, id_familia, id_marca, id_modelo,
                    id_unidad, unidad_medida_id,
                    tipo_control,
                    stock_minimo, stock_maximo, punto_reposicion,
                    ubicacion_id, observaciones,
                    fecha_alta, usuario_alta, raw_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    payload.codigo.strip(),
                    payload.descripcion.strip(),
                    payload.id_categoria,
                    payload.id_familia,
                    payload.id_marca,
                    payload.id_modelo,
                    payload.id_unidad,
                    payload.id_unidad,
                    "CONSUMIBLE",
                    0,
                    0,
                    0,
                    None,
                    payload.observaciones,
                    fecha_actual_iso(),
                    None,
                    json.dumps(payload.model_dump(), ensure_ascii=False),
                ),
            )
            auditar_v2(conn, None, "CREATE", "productos", cur.lastrowid, None, payload.model_dump())
            conn.commit()
    except sqlite3.OperationalError as exc:
        if "database is locked" in str(exc).lower():
            return {"error": "Base de datos ocupada. Reintente en unos segundos."}
        raise
    return {"mensaje": "Producto creado"}


@app.put("/almacen/v2/productos/{producto_id}")
def actualizar_producto_v2(producto_id: int, payload: ProductoV2Payload):
    with get_sqlite_connection() as conn:
        prev = conn.execute("SELECT * FROM productos WHERE id = ?", (producto_id,)).fetchone()
        if prev is None:
            return {"error": "Producto no encontrado"}
        conn.execute(
            """
            UPDATE productos
               SET codigo = ?,
                   descripcion = ?,
                   categoria_id = ?,
                   id_familia = ?,
                   id_marca = ?,
                   id_modelo = ?,
                   id_unidad = ?,
                   unidad_medida_id = ?,
                   tipo_control = ?,
                   stock_minimo = ?,
                   stock_maximo = ?,
                   punto_reposicion = ?,
                   ubicacion_id = ?,
                   observaciones = ?,
                   usuario_alta = ?,
                   raw_json = ?
             WHERE id = ?
            """,
            (
                payload.codigo.strip(),
                payload.descripcion.strip(),
                payload.id_categoria,
                payload.id_familia,
                payload.id_marca,
                payload.id_modelo,
                payload.id_unidad,
                payload.id_unidad,
                "CONSUMIBLE",
                payload.stock_ingreso,
                0,
                payload.punto_reposicion,
                payload.id_ubicacion,
                payload.observaciones,
                None,
                json.dumps(payload.model_dump(), ensure_ascii=False),
                producto_id,
            ),
        )
        auditar_v2(conn, None, "UPDATE", "productos", producto_id, dict(prev), payload.model_dump())
        conn.commit()
    return {"mensaje": "Producto actualizado"}


@app.delete("/almacen/v2/productos/{producto_id}")
def eliminar_producto_v2(producto_id: int):
    with get_sqlite_connection() as conn:
        prev = conn.execute("SELECT * FROM productos WHERE id = ?", (producto_id,)).fetchone()
        if prev is None:
            return {"error": "Producto no encontrado"}
        mov = conn.execute("SELECT 1 FROM movimientos WHERE id_producto = ? LIMIT 1", (producto_id,)).fetchone()
        if mov:
            return {"error": "No se puede eliminar: tiene movimientos"}
        conn.execute("DELETE FROM productos WHERE id = ?", (producto_id,))
        auditar_v2(conn, None, "DELETE", "productos", producto_id, dict(prev), None)
        conn.commit()
    return {"mensaje": "Producto eliminado"}


@app.get("/almacen/v2/stock")
def stock_v2(codigo: str = "", descripcion: str = "", id_categoria: int | None = Query(default=None), id_ubicacion: int | None = Query(default=None)):
    with get_sqlite_connection() as conn:
        return _stock_v2_rows(conn, codigo=codigo, descripcion=descripcion, id_categoria=id_categoria, id_ubicacion=id_ubicacion)


def _stock_v2_rows(
    conn,
    *,
    codigo: str = "",
    descripcion: str = "",
    id_categoria: int | None = None,
    id_ubicacion: int | None = None,
):
    codigo_like = f"%{(codigo or '').strip()}%"
    descripcion_like = f"%{(descripcion or '').strip()}%"
    # Inventarios v2 toma el stock directamente de la tabla stock.
    asegurar_tabla_stock(conn)
    migrar_tabla_stock(conn)
    rows = conn.execute(
        """
        SELECT
            s.producto_id AS id,
            COALESCE(NULLIF(s.producto_codigo, ''), p.codigo, '') AS codigo,
            COALESCE(NULLIF(s.producto_descripcion, ''), p.descripcion, '') AS descripcion,
            COALESCE(NULLIF(s.familia_nombre, ''), f.nombre, '') AS familia_nombre,
            COALESCE(NULLIF(s.marca_nombre, ''), m.nombre, '') AS marca_nombre,
            COALESCE(NULLIF(s.modelo_nombre, ''), md.nombre, '') AS modelo_nombre,
            COALESCE(NULLIF(s.unidad_nombre, ''), um.nombre, '') AS unidad_nombre,
            COALESCE(c.nombre, '') AS categoria,
            COALESCE(u.ruta, '') AS ubicacion,
            COALESCE(s.stock_actual, 0) AS stock_actual,
            COALESCE(p.punto_reposicion, 0) AS punto_reposicion
        FROM stock s
        LEFT JOIN productos p ON p.id = s.producto_id
        LEFT JOIN familias f ON f.id = COALESCE(s.id_familia, p.id_familia)
        LEFT JOIN marcas m ON m.id = COALESCE(s.id_marca, p.id_marca)
        LEFT JOIN modelos md ON md.id = COALESCE(s.id_modelo, p.id_modelo)
        LEFT JOIN unidades_medida um ON um.id = COALESCE(s.id_unidad, p.id_unidad, p.unidad_medida_id)
        LEFT JOIN categorias c ON c.id = p.categoria_id
        LEFT JOIN ubicaciones u ON u.id = COALESCE(s.id_ubicacion, p.ubicacion_id)
        WHERE COALESCE(NULLIF(s.producto_codigo, ''), p.codigo, '') LIKE ?
          AND COALESCE(NULLIF(s.producto_descripcion, ''), p.descripcion, '') LIKE ?
          AND (? IS NULL OR p.categoria_id = ?)
          AND (? IS NULL OR COALESCE(s.id_ubicacion, p.ubicacion_id) = ?)
        ORDER BY COALESCE(NULLIF(s.producto_codigo, ''), p.codigo, '')
        """,
        (codigo_like, descripcion_like, id_categoria, id_categoria, id_ubicacion, id_ubicacion),
    ).fetchall()

    out = []
    for r in rows:
        item = dict(r)
        reservado = obtener_stock_reservado(conn, int(item.get("id") or 0))
        item["stock_reservado"] = round(reservado, 2)
        item["stock_disponible"] = round(float(item.get("stock_actual") or 0) - reservado, 2)
        out.append(item)
    return out


def _filtrar_stock_v2_rows(
    filas,
    *,
    codigo: str = "",
    descripcion: str = "",
    categoria: str = "",
    ubicacion: str = "",
):
    codigo_q = (codigo or "").strip().lower()
    descripcion_q = (descripcion or "").strip().lower()
    categoria_q = (categoria or "").strip().lower()
    ubicacion_q = (ubicacion or "").strip().lower()

    out = []
    for fila in filas:
        codigo_v = str(fila.get("codigo") or "").lower()
        descripcion_v = str(fila.get("descripcion") or "").lower()
        categoria_v = str(fila.get("categoria") or "").lower()
        ubicacion_v = str(fila.get("ubicacion") or "").lower()
        if codigo_q and codigo_q not in codigo_v:
            continue
        if descripcion_q and descripcion_q not in descripcion_v:
            continue
        if categoria_q and categoria_q not in categoria_v:
            continue
        if ubicacion_q and ubicacion_q not in ubicacion_v:
            continue
        out.append(fila)
    return out


@app.get("/almacen/v2/inventarios/pdf")
def inventarios_v2_pdf(
    codigo: str = "",
    descripcion: str = "",
    categoria: str = "",
    ubicacion: str = "",
):
    with get_sqlite_connection() as conn:
        filas = _stock_v2_rows(conn)
    filas = _filtrar_stock_v2_rows(
        filas,
        codigo=codigo,
        descripcion=descripcion,
        categoria=categoria,
        ubicacion=ubicacion,
    )

    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        from reportlab.lib import colors
    except Exception as exc:
        return {"error": f"No se pudo generar el PDF: {exc}"}

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=landscape(A4))
    ancho, alto = landscape(A4)

    if os.path.exists(MEMBRETE_LOGO_PATH):
        try:
            logo = ImageReader(MEMBRETE_LOGO_PATH)
            c.drawImage(logo, ancho - 185, alto - 92, width=140, height=45, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    y = alto - 40
    c.setFont("Helvetica-Bold", 15)
    c.drawString(190, y, "Departamento de Almacen")
    y -= 24
    c.setFont("Helvetica-Bold", 13)
    c.drawString(190, y, "Reporte de Inventario")
    y -= 20

    c.setFont("Helvetica-Bold", 9)
    c.drawString(40, y, "Razon Social :")
    c.drawString(40, y - 13, "Telefono :")
    c.drawString(40, y - 26, "Cod.Postal :")
    c.drawString(40, y - 39, "Direccion :")
    c.drawString(40, y - 52, "CUIT :")
    c.setFont("Helvetica", 9)
    c.drawString(120, y, EMPRESA_MEMBRETE["razon_social"])
    c.drawString(120, y - 13, EMPRESA_MEMBRETE["telefono"])
    c.drawString(120, y - 26, EMPRESA_MEMBRETE["codigo_postal"])
    c.drawString(120, y - 39, EMPRESA_MEMBRETE["direccion"])
    c.drawString(120, y - 52, EMPRESA_MEMBRETE["cuit"])
    y -= 72

    filtros = []
    if codigo:
        filtros.append(f"Codigo: {codigo}")
    if descripcion:
        filtros.append(f"Descripcion: {descripcion}")
    if categoria:
        filtros.append(f"Categoria: {categoria}")
    if ubicacion:
        filtros.append(f"Ubicacion: {ubicacion}")
    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 14
    if filtros:
        c.drawString(40, y, " | ".join(filtros)[:165])
        y -= 16

    table_left = 36
    table_right = ancho - 44
    table_width = table_right - table_left
    col_specs = [
        ("N°", 36.0, "right"),
        ("Codigo", 68.0, "left"),
        ("Descripcion", 145.0, "left"),
        ("Familia", 78.0, "left"),
        ("Marca", 78.0, "left"),
        ("Modelo", 78.0, "left"),
        ("Unidad", 56.0, "left"),
        ("Categoria", 86.0, "left"),
        ("Ubicacion", 122.0, "left"),
        ("Stock actual", 66.0, "right"),
    ]
    total_base = sum(width for _, width, _ in col_specs)
    scale = min(1.0, table_width / total_base) if total_base > 0 else 1.0
    columnas = []
    x_cursor = table_left
    for titulo_col, width_base, align in col_specs:
        ancho_col = width_base * scale
        columnas.append({"titulo": titulo_col, "x": x_cursor, "w": ancho_col, "align": align})
        x_cursor += ancho_col
    # Compensa deriva por redondeo para que la ultima columna cierre exactamente en el borde derecho.
    if columnas:
        usado = columnas[-1]["x"] + columnas[-1]["w"]
        columnas[-1]["w"] += (table_right - usado)

    def texto_ajustado(valor: str, max_width: float, font_name: str, font_size: int) -> str:
        texto = str(valor or "")
        if c.stringWidth(texto, font_name, font_size) <= max_width:
            return texto
        sufijo = "..."
        if c.stringWidth(sufijo, font_name, font_size) > max_width:
            return ""
        low, high = 0, len(texto)
        while low < high:
            mid = (low + high + 1) // 2
            candidato = texto[:mid] + sufijo
            if c.stringWidth(candidato, font_name, font_size) <= max_width:
                low = mid
            else:
                high = mid - 1
        return texto[:low] + sufijo

    def dibujar_cabecera_tabla(y_pos: float):
        c.setFillColor(colors.HexColor("#1d4ed8"))
        c.rect(table_left, y_pos - 2, table_width, 16, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        pad = 5
        for col in columnas:
            max_txt = max(col["w"] - (pad * 2), 8)
            titulo = texto_ajustado(col["titulo"], max_txt, "Helvetica-Bold", 8)
            if col["align"] == "right":
                c.drawRightString(col["x"] + col["w"] - pad, y_pos + 2, titulo)
            else:
                c.drawString(col["x"] + pad, y_pos + 2, titulo)
        c.setFillColor(colors.black)
        c.setStrokeColor(colors.HexColor("#cbd5e1"))
        c.line(table_left, y_pos - 3, table_right, y_pos - 3)

    dibujar_cabecera_tabla(y)
    y -= 16

    c.setFont("Helvetica", 8)
    if not filas:
        c.drawString(40, y, "Sin registros para los filtros aplicados.")
    else:
        for idx, fila in enumerate(filas, start=1):
            if y < 42:
                c.showPage()
                y = alto - 42
                dibujar_cabecera_tabla(y)
                y -= 16
                c.setFont("Helvetica", 8)

            valores = [
                idx,
                fila.get("codigo") or "",
                fila.get("descripcion") or "",
                fila.get("familia_nombre") or "",
                fila.get("marca_nombre") or "",
                fila.get("modelo_nombre") or "",
                fila.get("unidad_nombre") or "",
                fila.get("categoria") or "",
                fila.get("ubicacion") or "",
                f"{float(fila.get('stock_actual') or 0):g}",
            ]
            pad = 5
            for col, valor in zip(columnas, valores):
                max_txt = max(col["w"] - (pad * 2), 8)
                txt = texto_ajustado(valor, max_txt, "Helvetica", 8)
                if col["align"] == "right":
                    c.drawRightString(col["x"] + col["w"] - pad, y, txt)
                else:
                    c.drawString(col["x"] + pad, y, txt)
            y -= 12

    c.save()
    pdf_bytes = pdf_buffer.getvalue()
    pdf_buffer.close()
    nombre = f"Inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@app.get("/almacen/v2/inventarios/excel")
def inventarios_v2_excel(
    codigo: str = "",
    descripcion: str = "",
    categoria: str = "",
    ubicacion: str = "",
):
    with get_sqlite_connection() as conn:
        filas = _stock_v2_rows(conn)
    filas = _filtrar_stock_v2_rows(
        filas,
        codigo=codigo,
        descripcion=descripcion,
        categoria=categoria,
        ubicacion=ubicacion,
    )

    try:
        from openpyxl import Workbook
    except Exception as exc:
        return {"error": f"No se pudo generar el XLSX: {exc}"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    headers = ["N°", "Codigo", "Descripcion", "Familia", "Marca", "Modelo", "Unidad", "Categoria", "Ubicacion", "Stock actual"]
    ws.append(headers)
    for idx, fila in enumerate(filas, start=1):
        ws.append([
            idx,
            fila.get("codigo") or "",
            fila.get("descripcion") or "",
            fila.get("familia_nombre") or "",
            fila.get("marca_nombre") or "",
            fila.get("modelo_nombre") or "",
            fila.get("unidad_nombre") or "",
            fila.get("categoria") or "",
            fila.get("ubicacion") or "",
            float(fila.get("stock_actual") or 0),
        ])

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(12, max_len + 2), 50)

    out = io.BytesIO()
    wb.save(out)
    contenido = out.getvalue()
    out.close()

    nombre = f"Inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@app.post("/almacen/v2/documentos")
def crear_documento_v2(payload: DocumentoV2Payload):
    tipo = (payload.tipo or "").strip().upper()
    if tipo not in TIPOS_DOCUMENTO_VALIDOS:
        return {"error": "Tipo de documento invalido"}
    if not payload.detalles:
        return {"error": "El documento requiere detalle"}

    fecha_doc = payload.fecha or datetime.now().strftime("%Y-%m-%d")
    with get_sqlite_connection() as conn:
        numero = numero_documento_v2(conn, tipo)
        cur_doc = conn.execute(
            """
            INSERT INTO documentos (tipo, numero, fecha, estado, responsable, observaciones, payload_json)
            VALUES (?, ?, ?, 'CONFIRMADO', ?, ?, ?)
            """,
            (tipo, numero, fecha_doc, payload.responsable, payload.observaciones, json.dumps(payload.model_dump(), ensure_ascii=False)),
        )
        id_documento = cur_doc.lastrowid

        for d in payload.detalles:
            if d.id_producto is None:
                return {"error": "Cada detalle requiere id_producto"}

            id_producto = d.id_producto

            prod = producto_v2_por_id(conn, id_producto)
            if prod is None:
                return {"error": f"Producto inexistente: {id_producto}"}

            cantidad = float(d.cantidad or 0)

            if tipo in {"RE", "RT"}:
                stock_now = stock_actual_v2(conn, id_producto)
                if stock_now < cantidad:
                    return {"error": f"Stock insuficiente para {prod['codigo']}"}

            cur_det = conn.execute(
                """
                INSERT INTO documentos_detalle (
                    id_documento, id_producto, cantidad,
                    id_vehiculo, id_proyecto, id_instalacion, observaciones
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    id_documento,
                    id_producto,
                    cantidad,
                    d.id_vehiculo,
                    d.id_proyecto,
                    d.id_instalacion,
                    d.observaciones,
                ),
            )

            movimiento_v2_registrar(
                conn,
                tipo=tipo,
                documento=numero,
                id_documento=id_documento,
                id_producto=id_producto,
                cantidad=cantidad,
                id_personal=payload.responsable,
                id_vehiculo=d.id_vehiculo,
                id_proyecto=d.id_proyecto,
                id_instalacion=d.id_instalacion,
                observaciones=d.observaciones or payload.observaciones,
            )

            auditar_v2(conn, payload.responsable, "CREATE", "documentos_detalle", cur_det.lastrowid, None, d.model_dump())

        auditar_v2(conn, payload.responsable, "CREATE", "documentos", id_documento, None, payload.model_dump())
        conn.commit()

        # PDF profesional reutilizando generador existente
        detalle_rows = conn.execute(
            """
            SELECT p.codigo, p.descripcion, d.cantidad, d.id_proyecto
            FROM documentos_detalle d
            LEFT JOIN productos p ON p.id = d.id_producto
            WHERE d.id_documento = ?
            ORDER BY d.id
            """,
            (id_documento,),
        ).fetchall()
        proy_map = {
            r["id"]: r["nombre"]
            for r in conn.execute("SELECT id, nombre FROM proyectos").fetchall()
        }
        detalle_pdf = [
            (
                f"{r['codigo'] or ''} - {r['descripcion'] or ''}",
                str(r["cantidad"]),
                "",
                proy_map.get(r["id_proyecto"], "") if r["id_proyecto"] else "",
            )
            for r in detalle_rows
        ]
        pdf_path = os.path.join(DOC_ALMACEN_DIR, f"{numero}.pdf")
        generar_pdf_remito(
            pdf_path,
            f"Documento {tipo} {numero}",
            [
                ("Fecha", fecha_doc),
                ("Responsable", payload.responsable or ""),
                ("Observaciones", payload.observaciones or ""),
            ],
            detalle_pdf,
        )

    return {"mensaje": "Documento creado", "id": id_documento, "numero": numero, "tipo": tipo}


@app.get("/almacen/v2/documentos")
def listar_documentos_v2(tipo: str = ""):
    tipo = (tipo or "").strip().upper()
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT * FROM documentos
            WHERE (? = '' OR tipo = ?)
            ORDER BY id DESC
            """,
            (tipo, tipo),
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/v2/documentos/{id_documento}/detalle")
def detalle_documento_v2(id_documento: int):
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT d.*, p.codigo, p.descripcion
            FROM documentos_detalle d
            LEFT JOIN productos p ON p.id = d.id_producto
            WHERE d.id_documento = ?
            ORDER BY d.id
            """,
            (id_documento,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/v2/documentos/{id_documento}/pdf")
def documento_pdf_v2(id_documento: int):
    with get_sqlite_connection() as conn:
        row = conn.execute("SELECT numero FROM documentos WHERE id = ?", (id_documento,)).fetchone()
    if row is None:
        return {"error": "Documento inexistente"}
    path_pdf = os.path.join(DOC_ALMACEN_DIR, f"{row['numero']}.pdf")
    if not os.path.exists(path_pdf):
        return {"error": "PDF no disponible"}
    return FileResponse(path_pdf, media_type="application/pdf", filename=os.path.basename(path_pdf))


@app.get("/almacen/v2/movimientos")
def listar_movimientos_v2(
    fecha_desde: str = "",
    fecha_hasta: str = "",
    id_producto: int | None = Query(default=None),
    tipo: str = "",
):
    with get_sqlite_connection() as conn:
        return _listar_movimientos_v2_rows(conn, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, id_producto=id_producto, tipo=tipo)


def _listar_movimientos_v2_rows(
    conn,
    *,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    id_producto: int | None = None,
    tipo: str = "",
):
    tipo = (tipo or "").strip().upper()
    tipo_db = tipo
    if tipo == "RI":
        tipo_db = "ENTRADA"
    elif tipo == "RE":
        tipo_db = "SALIDA"

    rows = conn.execute(
        """
        SELECT
            m.*,
            p.codigo AS producto_codigo,
            p.descripcion AS producto_descripcion,
            re.destinatario AS re_destinatario
        FROM movimientos_stock m
        LEFT JOIN productos p ON p.id = m.producto_id
        LEFT JOIN remitos_entrega re ON re.id = m.remito_id
        WHERE (? = '' OR SUBSTR(m.fecha,1,10) >= ?)
          AND (? = '' OR SUBSTR(m.fecha,1,10) <= ?)
          AND (? IS NULL OR m.producto_id = ?)
          AND (? = '' OR m.tipo = ?)
        ORDER BY m.id DESC
        LIMIT 1200
        """,
        (fecha_desde, fecha_desde, fecha_hasta, fecha_hasta, id_producto, id_producto, tipo_db, tipo_db),
    ).fetchall()

    out = []
    for r in rows:
        reg = dict(r)
        tipo_mov = (reg.get("tipo") or "").strip().upper()
        nombre_personal = (reg.get("responsable_nombre") or "").strip()

        if tipo_mov == "SALIDA":
            reg["empleado"] = (
                (reg.get("re_destinatario") or "").strip()
                or nombre_personal
            )
        else:
            reg["empleado"] = nombre_personal

        reg.pop("re_destinatario", None)
        out.append(reg)

    return out


def _filtrar_movimientos_v2_rows(
    filas,
    *,
    documento: str = "",
    producto: str = "",
    tipo: str = "",
):
    documento_filtro = (documento or "").strip().lower()
    producto_filtro = (producto or "").strip().lower()
    tipo_filtro = (tipo or "").strip().upper()

    if not documento_filtro and not producto_filtro and not tipo_filtro:
        return filas

    filtradas = []
    for fila in filas:
        tipo_mov = (fila.get("tipo") or "").strip().upper()
        if tipo_filtro:
            if tipo_filtro in {"ENTRADA", "RI"} and tipo_mov not in {"ENTRADA", "RI"}:
                continue
            if tipo_filtro in {"SALIDA", "RE"} and tipo_mov not in {"SALIDA", "RE"}:
                continue
            if tipo_filtro not in {"ENTRADA", "RI", "SALIDA", "RE"} and tipo_mov != tipo_filtro:
                continue

        documento_txt = (fila.get("documento") or "").strip().lower()
        producto_txt = f"{fila.get('producto_codigo') or ''} {fila.get('producto_descripcion') or ''}".strip().lower()
        if documento_filtro and documento_filtro not in documento_txt:
            continue
        if producto_filtro and producto_filtro not in producto_txt:
            continue
        filtradas.append(fila)

    return filtradas


def _formato_cantidad_movimiento(tipo_mov: str, cantidad: float) -> str:
    tipo_norm = (tipo_mov or "").strip().upper()
    signo = "+" if tipo_norm in {"RI", "ENTRADA"} else "-" if tipo_norm in {"RE", "SALIDA"} else ""
    return f"{signo}{cantidad:g}"


@app.get("/almacen/v2/movimientos/pdf")
def movimientos_v2_pdf(
    fecha_desde: str = "",
    fecha_hasta: str = "",
    documento: str = "",
    producto: str = "",
    tipo: str = "",
):
    with get_sqlite_connection() as conn:
        filas = _listar_movimientos_v2_rows(
            conn,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            tipo=tipo,
        )
    filas = _filtrar_movimientos_v2_rows(filas, documento=documento, producto=producto, tipo=tipo)

    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        from reportlab.lib import colors
    except Exception as exc:
        return {"error": f"No se pudo generar el PDF: {exc}"}

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=landscape(A4))
    ancho, alto = landscape(A4)

    if os.path.exists(MEMBRETE_LOGO_PATH):
        try:
            logo = ImageReader(MEMBRETE_LOGO_PATH)
            c.drawImage(logo, ancho - 185, alto - 92, width=140, height=45, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    y = alto - 40
    c.setFont("Helvetica-Bold", 15)
    c.drawString(190, y, "Departamento de Almacen")
    y -= 24
    c.setFont("Helvetica-Bold", 13)
    c.drawString(190, y, "Reporte de Movimientos")
    y -= 20

    c.setFont("Helvetica-Bold", 9)
    c.drawString(40, y, "Razon Social :")
    c.drawString(40, y - 13, "Telefono :")
    c.drawString(40, y - 26, "Cod.Postal :")
    c.drawString(40, y - 39, "Direccion :")
    c.drawString(40, y - 52, "CUIT :")
    c.setFont("Helvetica", 9)
    c.drawString(120, y, EMPRESA_MEMBRETE["razon_social"])
    c.drawString(120, y - 13, EMPRESA_MEMBRETE["telefono"])
    c.drawString(120, y - 26, EMPRESA_MEMBRETE["codigo_postal"])
    c.drawString(120, y - 39, EMPRESA_MEMBRETE["direccion"])
    c.drawString(120, y - 52, EMPRESA_MEMBRETE["cuit"])
    y -= 72

    filtros = []
    if fecha_desde:
        filtros.append(f"Desde: {fecha_desde}")
    if fecha_hasta:
        filtros.append(f"Hasta: {fecha_hasta}")
    if tipo:
        filtros.append(f"Tipo: {tipo}")
    if documento:
        filtros.append(f"Documento: {documento}")
    if producto:
        filtros.append(f"Producto: {producto}")

    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 14
    if filtros:
        c.drawString(40, y, " | ".join(filtros)[:165])
        y -= 16

    table_left = 36
    table_right = ancho - 44
    table_width = table_right - table_left
    # Anchos base por columna (en puntos), luego se escalan si exceden el area util.
    col_specs = [
        ("Fecha", 52.0, "left"),
        ("Tipo", 44.0, "left"),
        ("Documento", 74.0, "left"),
        ("Producto", 190.0, "left"),
        ("Empleado", 150.0, "left"),
        ("Cantidad", 62.0, "right"),
        ("Stock ant", 62.0, "right"),
        ("Stock nuevo", 74.0, "right"),
    ]
    total_base = sum(width for _, width, _ in col_specs)
    scale = min(1.0, table_width / total_base) if total_base > 0 else 1.0
    columnas = []
    x_cursor = table_left
    for titulo_col, width_base, align in col_specs:
        ancho_col = width_base * scale
        columnas.append({"titulo": titulo_col, "x": x_cursor, "w": ancho_col, "align": align})
        x_cursor += ancho_col

    def texto_ajustado(valor: str, max_width: float, font_name: str, font_size: int) -> str:
        texto = str(valor or "")
        if c.stringWidth(texto, font_name, font_size) <= max_width:
            return texto
        sufijo = "..."
        if c.stringWidth(sufijo, font_name, font_size) > max_width:
            return ""
        low, high = 0, len(texto)
        while low < high:
            mid = (low + high + 1) // 2
            candidato = texto[:mid] + sufijo
            if c.stringWidth(candidato, font_name, font_size) <= max_width:
                low = mid
            else:
                high = mid - 1
        return texto[:low] + sufijo

    def dibujar_cabecera_tabla(y_pos: float):
        c.setFillColor(colors.HexColor("#1d4ed8"))
        c.rect(table_left, y_pos - 2, table_width, 16, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        pad = 4
        for col in columnas:
            max_txt = max(col["w"] - (pad * 2), 8)
            titulo = texto_ajustado(col["titulo"], max_txt, "Helvetica-Bold", 8)
            if col["align"] == "right":
                c.drawRightString(col["x"] + col["w"] - pad, y_pos + 2, titulo)
            else:
                c.drawString(col["x"] + pad, y_pos + 2, titulo)
        c.setFillColor(colors.black)
        c.setStrokeColor(colors.HexColor("#cbd5e1"))
        c.line(table_left, y_pos - 3, table_right, y_pos - 3)

    dibujar_cabecera_tabla(y)
    y -= 16

    c.setFont("Helvetica", 8)
    if not filas:
        c.drawString(40, y, "Sin registros para los filtros aplicados.")
    else:
        for fila in filas:
            if y < 42:
                c.showPage()
                y = alto - 42
                dibujar_cabecera_tabla(y)
                y -= 16
                c.setFont("Helvetica", 8)

            fecha_txt = (fila.get("fecha") or "")[:10]
            tipo_txt = fila.get("tipo") or ""
            doc_txt = fila.get("documento") or ""
            prod_txt = f"{fila.get('producto_codigo') or ''} {fila.get('producto_descripcion') or ''}".strip()
            emp_txt = fila.get("empleado") or ""
            cant_txt = _formato_cantidad_movimiento(fila.get("tipo") or "", float(fila.get("cantidad") or 0))
            stock_ant = f"{float(fila.get('stock_anterior') or 0):g}"
            stock_nuevo = f"{float(fila.get('stock_nuevo') or 0):g}"

            fila_valores = [fecha_txt, tipo_txt, doc_txt, prod_txt, emp_txt, cant_txt, stock_ant, stock_nuevo]
            pad = 4
            for col, valor in zip(columnas, fila_valores):
                max_txt = max(col["w"] - (pad * 2), 8)
                txt = texto_ajustado(valor, max_txt, "Helvetica", 8)
                if col["align"] == "right":
                    c.drawRightString(col["x"] + col["w"] - pad, y, txt)
                else:
                    c.drawString(col["x"] + pad, y, txt)
            y -= 12

    c.save()
    pdf_bytes = pdf_buffer.getvalue()
    pdf_buffer.close()
    nombre = f"Movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@app.get("/almacen/v2/movimientos/excel")
def movimientos_v2_excel(
    fecha_desde: str = "",
    fecha_hasta: str = "",
    documento: str = "",
    producto: str = "",
    tipo: str = "",
):
    with get_sqlite_connection() as conn:
        filas = _listar_movimientos_v2_rows(
            conn,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            tipo=tipo,
        )
    filas = _filtrar_movimientos_v2_rows(filas, documento=documento, producto=producto, tipo=tipo)

    try:
        from openpyxl import Workbook
    except Exception as exc:
        return {"error": f"No se pudo generar el XLSX: {exc}"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    headers = ["Fecha", "Tipo", "Documento", "Producto", "Empleado", "Cantidad", "Stock ant", "Stock nuevo"]
    ws.append(headers)

    for fila in filas:
        ws.append([
            (fila.get("fecha") or "")[:10],
            fila.get("tipo") or "",
            fila.get("documento") or "",
            f"{fila.get('producto_codigo') or ''} {fila.get('producto_descripcion') or ''}".strip(),
            fila.get("empleado") or "",
            _formato_cantidad_movimiento(fila.get("tipo") or "", float(fila.get("cantidad") or 0)),
            float(fila.get("stock_anterior") or 0),
            float(fila.get("stock_nuevo") or 0),
        ])

    # Ajuste rapido de ancho por legibilidad en Excel.
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(12, max_len + 2), 48)

    out = io.BytesIO()
    wb.save(out)
    contenido = out.getvalue()
    out.close()
    nombre = f"Movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@app.post("/almacen/v2/inventarios")
def crear_inventario_v2(responsable: str | None = None):
    with get_sqlite_connection() as conn:
        numero = numero_documento_v2(conn, "INV")
        cur = conn.execute(
            """
            INSERT INTO inventarios (numero, fecha, estado, responsable_legajo, responsable_nombre, observaciones, raw_json)
            VALUES (?, ?, 'PENDIENTE', ?, (SELECT nombre FROM personal WHERE legajo = ?), '', '{}')
            """,
            (numero, datetime.now().strftime("%Y-%m-%d"), responsable, responsable),
        )
        id_inv = cur.lastrowid
        productos = conn.execute("SELECT id FROM productos").fetchall()
        for p in productos:
            conn.execute(
                "INSERT INTO inventarios_detalle (inventario_id, producto_id, stock_sistema, estado) VALUES (?, ?, ?, 'PENDIENTE')",
                (id_inv, p["id"], stock_actual_v2(conn, p["id"])),
            )
        auditar_v2(conn, responsable, "CREATE", "inventarios", id_inv, None, {"numero": numero})
        conn.commit()
    return {"mensaje": "Inventario creado", "id": id_inv, "numero": numero}


@app.get("/almacen/v2/inventarios")
def listar_inventarios_v2():
    with get_sqlite_connection() as conn:
        rows = conn.execute("SELECT * FROM inventarios ORDER BY id DESC").fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/v2/inventarios/{inventario_id}/detalle")
def detalle_inventarios_v2(inventario_id: int):
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT d.*, p.codigo AS producto_codigo, p.descripcion AS producto_descripcion
            FROM inventarios_detalle d
            JOIN productos p ON p.id = d.producto_id
            WHERE d.inventario_id = ?
            ORDER BY p.codigo
            """,
            (inventario_id,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.put("/almacen/v2/inventarios/detalle/{detalle_id}")
def conteo_inventario_v2(detalle_id: int, payload: InventarioConteoV2Payload):
    with get_sqlite_connection() as conn:
        row = conn.execute("SELECT * FROM inventarios_detalle WHERE id = ?", (detalle_id,)).fetchone()
        if row is None:
            return {"error": "Detalle inexistente"}
        diferencia = float(payload.stock_fisico) - float(row["stock_sistema"] or 0)
        estado = "OK" if abs(diferencia) < 1e-9 else "CON_DIFERENCIA"
        conn.execute(
            "UPDATE inventarios_detalle SET stock_fisico = ?, diferencia = ?, estado = ?, observaciones = ? WHERE id = ?",
            (payload.stock_fisico, diferencia, estado, payload.observaciones, detalle_id),
        )
        if abs(diferencia) > 1e-9:
            conn.execute(
                """
                INSERT INTO ajustes_stock (
                    inventario_id, inventario_detalle_id, producto_id, cantidad_ajuste,
                    estado, fecha_solicitud, observaciones, raw_json
                ) VALUES (?, ?, ?, ?, 'PENDIENTE', ?, ?, ?)
                """,
                (
                    row["inventario_id"],
                    detalle_id,
                    row["producto_id"],
                    diferencia,
                    fecha_actual_iso(),
                    payload.observaciones,
                    json.dumps(payload.model_dump(), ensure_ascii=False),
                ),
            )
        auditar_v2(conn, None, "UPDATE", "inventarios_detalle", detalle_id, dict(row), payload.model_dump())
        conn.commit()
    return {"mensaje": "Conteo guardado", "diferencia": diferencia, "estado": estado}


@app.get("/almacen/v2/ajustes")
def listar_ajustes_v2():
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT a.*, p.codigo AS producto_codigo, p.descripcion AS producto_descripcion
            FROM ajustes_stock a
            LEFT JOIN productos p ON p.id = a.producto_id
            ORDER BY a.id DESC
            """
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/almacen/v2/auditoria")
def listar_auditoria_v2(tabla: str = "", registro: int | None = Query(default=None), limit: int = 300):
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT * FROM auditoria
            WHERE (? = '' OR tabla = ?)
              AND (? IS NULL OR registro = ?)
            ORDER BY id DESC
            LIMIT ?
            """,
            (tabla, tabla, registro, registro, max(1, min(limit, 2000))),
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/almacen/v2/adjuntos")
async def subir_adjunto_v2(
    archivo: UploadFile = File(...),
    tipo: str = Form(...),
    id_documento: int | None = Form(default=None),
    id_inventario: int | None = Form(default=None),
    id_ajuste: int | None = Form(default=None),
):
    ext = os.path.splitext(archivo.filename or "")[1].lower()
    permitidas = {".pdf", ".jpg", ".jpeg", ".png", ".docx", ".xlsx"}
    if ext not in permitidas:
        return {"error": "Extension no permitida"}

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    nombre = f"{stamp}_{(archivo.filename or 'adjunto').replace(' ', '_')}"
    destino = os.path.join(DOC_ALMACEN_ADJ_DIR, nombre)
    contenido = await archivo.read()
    with open(destino, "wb") as f:
        f.write(contenido)

    with get_sqlite_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO adjuntos (fecha, tipo, archivo, id_documento, id_inventario, id_ajuste)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (fecha_actual_iso(), tipo, nombre, id_documento, id_inventario, id_ajuste),
        )
        auditar_v2(conn, None, "CREATE", "adjuntos", cur.lastrowid, None, {"archivo": nombre, "tipo": tipo})
        conn.commit()

    return {
        "mensaje": "Adjunto cargado",
        "id": cur.lastrowid,
        "archivo": nombre,
        "url": f"/Doc_Almacen_Adjuntos/{nombre}",
    }


@app.get("/almacen/v2/adjuntos")
def listar_adjuntos_v2(
    id_documento: int | None = Query(default=None),
    id_inventario: int | None = Query(default=None),
    id_ajuste: int | None = Query(default=None),
):
    with get_sqlite_connection() as conn:
        rows = conn.execute(
            """
            SELECT * FROM adjuntos
            WHERE (? IS NULL OR id_documento = ?)
              AND (? IS NULL OR id_inventario = ?)
              AND (? IS NULL OR id_ajuste = ?)
            ORDER BY id DESC
            """,
                        (id_documento, id_documento, id_inventario, id_inventario, id_ajuste, id_ajuste),
        ).fetchall()
    salida = []
    for r in rows:
        d = dict(r)
        d["url"] = f"/Doc_Almacen_Adjuntos/{d['archivo']}"
        salida.append(d)
    return salida

